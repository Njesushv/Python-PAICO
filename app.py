from flask import Flask, render_template, request, redirect, url_for, send_file, abort
import sqlite3
import os
import time
import qrcode
import io
from flask import send_from_directory, abort
import socket
from flask import send_file
import pandas as pd
from io import BytesIO
from flask import jsonify
from PIL import Image, ImageDraw, ImageFont
from flask import json
import re
from flask import render_template, request, redirect, url_for
from datetime import datetime, timedelta
from typing import Tuple, Optional
from flask import session, flash
from waitress import serve
import webbrowser
from threading import Timer
import threading
import os, sys, sqlite3
from flask import Flask, render_template, request, redirect, url_for, session
import logging
logging.basicConfig(level=logging.DEBUG)
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from flask import request, render_template
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
#==jhonatan hola
#====================================================================
# FUNCIONES AUXILIARES
def obtener_ip_local():
    """Obtiene la IP local de la red (no 127.0.0.1) - Funciona SIN internet"""
    try:
        # Obtener todas las IPs de las interfaces de red
        hostname = socket.gethostname()
        ips = socket.gethostbyname_ex(hostname)[2]
        
        # Filtrar IPs: buscar la que NO sea 127.0.0.1
        for ip in ips:
            if not ip.startswith('127.'):
                return ip
        
        # Si solo hay 127.0.0.1, devolverla como fallback
        return '127.0.0.1'
    except Exception:
        return '127.0.0.1'

#====================================================================
# ABRIR NAVEGADOR AUTOMÁTICAMENTE
def abrir_navegador():
    ip_local = obtener_ip_local()
    url = f"http://{ip_local}:5000"
    print(f"🌐 Abriendo navegador en: {url}")
    print(f"📱 Acceso desde otros dispositivos en la red: {url}")
    webbrowser.open(url)

#====================================================================
# CONFIGURACIÓN DE FLASK Y RUTAS
def ruta(rel):
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.abspath(".")
    return os.path.join(base, rel)


# Configurar rutas para PyInstaller
if getattr(sys, 'frozen', False):
    # Ejecutando como .exe empaquetado
    BASE_DIR = os.path.dirname(sys.executable)  # Carpeta donde está el .exe
    TEMPLATE_DIR = ruta("templates")  # templates está dentro de _internal
    STATIC_DIR = os.path.join(BASE_DIR, "static")  # static está junto al .exe
else:
    # Ejecutando como script Python normal
    BASE_DIR = os.path.abspath(".")
    TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
    STATIC_DIR = os.path.join(BASE_DIR, "static")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = "clave_super_secreta"
DB_PATH = os.path.join(BASE_DIR, "base_datos.db")
QR_FOLDER = os.path.join(STATIC_DIR, "qrcodes")
os.makedirs(QR_FOLDER, exist_ok=True)
app.debug = True

def get_db_connection():
    con = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON;")
    return con

#====================================================================
# RUTA SHUTDOWN (para detener el servidor desde el navegador)
@app.route("/shutdown")
def shutdown():
    func = request.environ.get("werkzeug.server.shutdown")
    if func:
        func()
    return "Servidor detenido"

#====================================================================
# INICIALIZAR BASE DE DATOS
def inicializar_db():
    """Crea la base de datos según el modelo lógico, solo si no existe el archivo."""
    if os.path.exists(DB_PATH):
        return

    with sqlite3.connect(DB_PATH) as con:
        con.execute("PRAGMA foreign_keys = ON;")
        cur = con.cursor()

        # ============ TABLA RESPONSABLES ============
        cur.execute("""
            CREATE TABLE responsables (
                id_responsable      INTEGER PRIMARY KEY AUTOINCREMENT,
                nombres             TEXT,
                dni                 TEXT,
                celular             TEXT,
                correo              TEXT,
                cargo               TEXT
            );
        """)

        # ============ TABLA SEDES ============
        cur.execute("""
            CREATE TABLE sedes (
                id_sede                     INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre_sede                 TEXT,
                direccion                   TEXT,
                telefono                    TEXT,
                id_responsable              INTEGER,
                id_responsable_secundario   INTEGER,
                FOREIGN KEY (id_responsable) 
                    REFERENCES responsables(id_responsable),
                FOREIGN KEY (id_responsable_secundario) 
                    REFERENCES responsables(id_responsable)
            );
        """)

        # ============ TABLA DOCENTES ============
        cur.execute("""
            CREATE TABLE docentes (
                id_docente      INTEGER PRIMARY KEY AUTOINCREMENT,
                nombres         TEXT,
                dni             TEXT UNIQUE,
                celular         TEXT,
                correo          TEXT,
                especialidad    TEXT,
                profesion       TEXT,
                codigo          TEXT,
                estado          TEXT
            );
        """)

        # ============ TABLA HORARIOS ============
        cur.execute("""
            CREATE TABLE horarios (
                id_horario          INTEGER PRIMARY KEY AUTOINCREMENT,
                descripcion         TEXT,
                dias                TEXT,
                hora_inicio         TEXT,
                hora_fin            TEXT,
                tolerancia          INTEGER DEFAULT 10,
                tolerancia_inicio   INTEGER DEFAULT 0,
                tolerancia_fin      INTEGER DEFAULT 0,
                activo              INTEGER DEFAULT 1,
                nombre_horario      TEXT,
                nota                TEXT,
                entrada_temprana    INTEGER DEFAULT 0,
                salida_temprana     INTEGER DEFAULT 0
            );
        """)

        # ============ TABLA CURSOS ============
        cur.execute("""
            CREATE TABLE cursos (
                id_curso        INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre          TEXT,
                descripcion     TEXT,
                id_docente      INTEGER,
                id_sede         INTEGER,
                aula            TEXT,
                nombre_horario  TEXT,
                id_horario      INTEGER,
                estado          TEXT,
                fecha_inicio    TEXT,
                fecha_fin       TEXT,
                duracion_curso  INTEGER DEFAULT 0,
                FOREIGN KEY (id_docente) REFERENCES docentes(id_docente),
                FOREIGN KEY (id_sede)    REFERENCES sedes(id_sede),
                FOREIGN KEY (id_horario) REFERENCES horarios(id_horario)
            );
        """)

        # ============ TABLA ALUMNOS ============
        cur.execute("""
            CREATE TABLE alumnos (
                id_alumno       INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo          TEXT UNIQUE,
                dni             TEXT UNIQUE,
                nombres         TEXT,
                fecha_nacimiento DATE,
                celular         TEXT,
                correo          TEXT,
                estado          TEXT,
                institucion     TEXT,
                id_sede         INTEGER,
                grado_academico TEXT DEFAULT '',
                FOREIGN KEY (id_sede) REFERENCES sedes(id_sede)
            );
        """)

        # ============ TABLA MATRICULAS ============
        cur.execute("""
            CREATE TABLE matriculas (
                id_matricula    INTEGER PRIMARY KEY AUTOINCREMENT,
                id_alumno       INTEGER,
                id_curso        INTEGER,
                fecha_matricula DATE,
                deuda           REAL,
                estado          TEXT,
                monto           TEXT,
                observacion     TEXT,
                tipo_pago       TEXT,
                id_sede         INTEGER,
                FOREIGN KEY (id_alumno) REFERENCES alumnos(id_alumno) ON DELETE CASCADE,
                FOREIGN KEY (id_curso)  REFERENCES cursos(id_curso),
                FOREIGN KEY (id_sede)   REFERENCES sedes(id_sede)
            );
        """)

        # ============ TABLA ASISTENCIAS (ALUMNOS) ============
        cur.execute("""
            CREATE TABLE asistencias (
                id_asistencia   INTEGER PRIMARY KEY AUTOINCREMENT,
                id_matricula    INTEGER,
                fecha           DATE,
                hora            TIME,
                observacion     TEXT,
                FOREIGN KEY (id_matricula) 
                    REFERENCES matriculas(id_matricula) ON DELETE CASCADE
            );
        """)

        # ============ TABLA USUARIOS ============
        cur.execute("""
            CREATE TABLE usuarios (
                id_usuario      INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre_usuario  TEXT,
                clave           TEXT,
                rol             TEXT,
                id_sede         INTEGER,
                FOREIGN KEY (id_sede) REFERENCES sedes(id_sede)
            );
        """)

        # ============ TABLA ASISTENCIAS_DOCENTES ============
        cur.execute("""
            CREATE TABLE asistencias_docentes (
                id_asistencia_doc   INTEGER PRIMARY KEY AUTOINCREMENT,
                id_docente          INTEGER,
                fecha               TEXT,
                hora                TEXT,
                observacion         TEXT,
                FOREIGN KEY (id_docente) 
                    REFERENCES docentes(id_docente) ON DELETE CASCADE
            );
        """)

        # ============ TABLA ASIGNACIONES_DOCENTES ============
        cur.execute("""
            CREATE TABLE asignaciones_docentes (
                id_asignacion   INTEGER PRIMARY KEY AUTOINCREMENT,
                id_horario      INTEGER,
                id_curso        INTEGER,
                id_docente      INTEGER,
                FOREIGN KEY (id_horario) REFERENCES horarios(id_horario),
                FOREIGN KEY (id_curso)   REFERENCES cursos(id_curso),
                FOREIGN KEY (id_docente) REFERENCES docentes(id_docente)
            );
        """)

        # ============ TABLA PAGOS ============
        cur.execute("""
            CREATE TABLE pagos (
                id_pago         INTEGER PRIMARY KEY AUTOINCREMENT,
                id_alumno       INTEGER,
                mes             TEXT,
                monto           REAL,
                metodo_pago     TEXT,
                fecha_pago      DATE,
                observacion     TEXT,
                fecha_registro  TEXT,
                curso           TEXT,
                proximo_pago    TEXT,
                tipo_pago       TEXT,
                FOREIGN KEY (id_alumno) 
                    REFERENCES alumnos(id_alumno) ON DELETE CASCADE
            );
        """)

        # ============ TABLA FOTOCHECK_LOGS ============
        cur.execute("""
            CREATE TABLE fotocheck_logs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                id_alumno   INTEGER,
                fecha       TEXT,
                FOREIGN KEY (id_alumno) 
                    REFERENCES alumnos(id_alumno) ON DELETE CASCADE
            );
        """)

        con.commit()

# Llamada al inicio del sistema
inicializar_db()

#====================================================================
# RUTA LOGIN
@app.route("/login", methods=["GET", "POST"])
def login():
    # Si ya está logueado → prohibir volver al login
    if "usuario" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        usuario = (request.form.get('usuario') or '').strip()
        clave   = (request.form.get('clave')   or '').strip()

        con = sqlite3.connect("base_datos.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # Asegúrate de seleccionar id_usuario para poder buscar sus sedes
        cur.execute("""
            SELECT id_usuario, nombre_usuario, clave, rol
            FROM usuarios
            WHERE TRIM(nombre_usuario) = TRIM(?)
              AND TRIM(clave) = TRIM(?)
        """, (usuario, clave))

        user = cur.fetchone()

        if user:
            # 👉 OJO: aquí usamos una tabla intermedia usuario_sedes
            cur.execute("""
                SELECT id_sede 
                FROM usuario_sedes
                WHERE id_usuario = ?
            """, (user["id_usuario"],))
            
            # lista de sedes que puede ver este usuario
            sedes_ids = [fila["id_sede"] for fila in cur.fetchall()]
            con.close()

            # Limpiamos la sesión anterior, por si acaso
            session.clear()
            session["usuario"] = user["nombre_usuario"]
            session["rol"] = user["rol"]

            # 🔹 Lista de sedes que el usuario puede ver (multi-sede)
            session["sedes"] = sedes_ids

            # 🔹 Compatibilidad con tu código actual (que usa solo id_sede)
            #     → tomamos la primera sede si existe
            session["id_sede"] = sedes_ids[0] if sedes_ids else None

            return redirect(url_for("dashboard"))

        # Si no encontró usuario:
        con.close()
        flash("Usuario o contraseña incorrectos.")
        return redirect(url_for("login"))

    return render_template("login.html")


#====================================================================
# RUTA LOGOUT
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


#====================================================================

@app.route('/whoami', methods=['GET'])
def whoami():
    """
    Endpoint ligero usado para descubrimiento en LAN.
    Devuelve JSON con información básica del servidor.
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(('8.8.8.8', 80))
            local_ip = s.getsockname()[0]
        except Exception:
            local_ip = '127.0.0.1'
        finally:
            s.close()

        info = {
            "name": "ProyectoAsistenciaQR",
            "ip": local_ip,
            "port": 5000,
            "version": "1.0"
        }
        return jsonify(info)
    except Exception as e:
        app.logger.exception("Error en whoami: %s", e)
        return jsonify({"error": "internal"}), 500



#====================================================================
# RUTA DASHBOARD (PRINCIPAL)
@app.route('/')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('login'))  # Si no ha iniciado sesión → Login

    # ---- Datos de sesión ----
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes para usuarios multi-sede

    print("### DASHBOARD DEBUG ###")
    print("ROL        :", rol)
    print("SEDES_IDS  :", sedes_ids)
    print("ID_SEDE_USU:", id_sede_usuario)

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # =====================================================
    # 1) RESUMEN ALUMNOS x CURSO (filtrado por sede)
    # =====================================================
    where_cursos = "WHERE 1=1"
    params_cursos = []

    if rol != 'admin':
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            where_cursos += f" AND c.id_sede IN ({placeholders})"
            params_cursos.extend(sedes_ids)
        elif id_sede_usuario:
            where_cursos += " AND c.id_sede = ?"
            params_cursos.append(id_sede_usuario)

    cur.execute(f"""
        SELECT 
            c.nombre AS curso, 
            COUNT(m.id_matricula) AS cantidad
        FROM cursos c
        LEFT JOIN matriculas m ON c.id_curso = m.id_curso
        {where_cursos}
        GROUP BY c.id_curso, c.nombre
        ORDER BY c.nombre
    """, params_cursos)
    resumen_cursos = [dict(row) for row in cur.fetchall()]

    print("RESUMEN_CURSOS:", resumen_cursos)

    # =====================================================
    # 2) RESUMEN ALUMNOS x SEDE (admin = todas, usuario = sus sedes)
    # =====================================================
    where_sedes = "WHERE 1=1"
    params_sedes = []

    if rol != 'admin':
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            where_sedes += f" AND c.id_sede IN ({placeholders})"
            params_sedes.extend(sedes_ids)
        elif id_sede_usuario:
            where_sedes += " AND c.id_sede = ?"
            params_sedes.append(id_sede_usuario)

    cur.execute(f"""
        SELECT 
            IFNULL(s.nombre_sede, 'SIN SEDE') AS sede,
            COUNT(m.id_matricula) AS cantidad
        FROM sedes s
        LEFT JOIN cursos c ON c.id_sede = s.id_sede
        LEFT JOIN matriculas m ON c.id_curso = m.id_curso
        {where_sedes}
        GROUP BY s.id_sede, s.nombre_sede
        ORDER BY sede
    """, params_sedes)
    resumen_sedes = [dict(row) for row in cur.fetchall()]

    print("RESUMEN_SEDES:", resumen_sedes)

    # =====================================================
    # 3) RESUMEN POR INSTITUCIÓN (filtrado por sede)
    # =====================================================
    where_inst = "WHERE a.institucion IS NOT NULL"
    params_inst = []

    if rol != 'admin':
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            where_inst += f"""
                AND a.id_alumno IN (
                    SELECT m.id_alumno
                    FROM matriculas m
                    JOIN cursos c ON m.id_curso = c.id_curso
                    WHERE c.id_sede IN ({placeholders})
                )
            """
            params_inst.extend(sedes_ids)
        elif id_sede_usuario:
            where_inst += """
                AND a.id_alumno IN (
                    SELECT m.id_alumno
                    FROM matriculas m
                    JOIN cursos c ON m.id_curso = c.id_curso
                    WHERE c.id_sede = ?
                )
            """
            params_inst.append(id_sede_usuario)

    cur.execute(f"""
        SELECT 
            a.institucion AS institucion, 
            COUNT(a.id_alumno) AS cantidad
        FROM alumnos a
        {where_inst}
        GROUP BY a.institucion
        ORDER BY a.institucion
    """, params_inst)
    resumen_institucion = [dict(row) for row in cur.fetchall()]

    print("RESUMEN_INSTITUCION:", resumen_institucion)

    con.close()

    return render_template(
        "alumnos.html",          # 👈 si tu HTML se llama dashboard.html, cámbialo aquí
        rol=rol,
        resumen_cursos=resumen_cursos,
        resumen_sedes=resumen_sedes,
        resumen_institucion=resumen_institucion
    )




#====================================================================
# RUTA DESCARGAR QR ALUMNO
@app.route('/descargar_qr/<codigo>')
def descargar_qr(codigo):
    # Ruta donde guardas tus QR (ajústala si es distinta)
    qr_path = os.path.join(STATIC_DIR, 'qrcodes', f'{codigo}.png')

    if not os.path.exists(qr_path):
        return f"❌ No se encontró el QR del alumno con código {codigo}.", 404

    # Descarga el archivo directamente
    return send_file(qr_path, as_attachment=True, download_name=f"QR_{codigo}.png")

#====================================================================
# RUTA CONTROL DE ALUMNOS
@app.route('/control_alumnos')
def control_alumnos():

    if 'usuario' not in session:
        return redirect(url_for('login'))

    # ---- Datos de sesión ----
    rol = (session.get('rol') or '').strip().lower()
    id_sede_usuario = session.get('id_sede')
    sedes_ids = session.get('sedes') or []   # lista de sedes para usuarios multi-sede

    con = get_db_connection()
    cur = con.cursor()

    curso_filtro = request.args.get("curso")
    buscar_alumno = (request.args.get("buscar_alumno") or "").strip()

    # ==============================
    # 1) Lista de cursos (según sede)
    # ==============================
    if rol == 'admin':
        cur.execute("SELECT DISTINCT nombre FROM cursos ORDER BY nombre ASC")
        cursos = [row[0] for row in cur.fetchall()]
    else:
        # Usuario normal
        if sedes_ids:
            # Tiene varias sedes asignadas
            placeholders = ",".join("?" * len(sedes_ids))
            cur.execute(f"""
                SELECT DISTINCT nombre
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre ASC
            """, sedes_ids)
        else:
            # Una sola sede
            cur.execute("""
                SELECT DISTINCT nombre 
                FROM cursos 
                WHERE id_sede = ?
                ORDER BY nombre ASC
            """, (id_sede_usuario,))

        cursos = [row[0] for row in cur.fetchall()]

    # ==============================
    # 2) Consulta base de alumnos
    # ==============================
    query = """
        SELECT 
            a.id_alumno,
            a.codigo,
            a.dni,
            a.nombres,
            a.fecha_nacimiento,
            a.celular,
            a.correo,
            a.institucion,
            a.grado_academico,   
            a.estado AS estado_alumno,
            m.estado AS estado_matricula,
            m.deuda,
            c.nombre AS curso,
            h.nombre_horario AS turno,
            h.descripcion AS descripcion_horario,
            h.hora_inicio,
            h.hora_fin
        FROM alumnos a
        LEFT JOIN matriculas m ON a.id_alumno = m.id_alumno
        LEFT JOIN cursos c ON m.id_curso = c.id_curso
        LEFT JOIN horarios h ON c.id_horario = h.id_horario
        WHERE 1=1
    """

    params = []

    # --- Filtro por curso ---
    if curso_filtro:
        query += " AND IFNULL(c.nombre, '') = ?"
        params.append(curso_filtro)

    # --- Filtro por alumno (nombre, DNI o código) ---
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # --- Filtro por sede (no admin) ---
    if rol != 'admin':
        if sedes_ids:
            # Varias sedes
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Una sola sede
            query += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    query += " ORDER BY a.id_alumno ASC"

    cur.execute(query, tuple(params))
    alumnos = cur.fetchall()
    con.close()

    return render_template(
        'control_alumnos.html',
        alumnos=alumnos,
        cursos=cursos,
        curso_filtro=curso_filtro,
        buscar_alumno=buscar_alumno
    )

#====================================================================
# RUTA EXPORTAR ALUMNOS A PDF
@app.route('/exportar_alumnos_pdf')
def exportar_alumnos_pdf():
    # 🔐 Validar sesión
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol            = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids      = session.get("sedes") or []   # lista de sedes si el usuario tiene varias

    curso = request.args.get("curso")
    buscar_alumno = (request.args.get("buscar_alumno") or "").strip()

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    query = """
    SELECT 
        m.fecha_matricula,
        a.nombres,
        a.dni,
        a.celular,
        a.correo,
        a.institucion,
        s.nombre_sede AS sede,
        c.nombre AS curso,
        h.descripcion AS horario,
        m.monto
    FROM alumnos a
    LEFT JOIN matriculas m ON a.id_alumno = m.id_alumno
    LEFT JOIN cursos c ON m.id_curso = c.id_curso
    LEFT JOIN horarios h ON c.id_horario = h.id_horario
    LEFT JOIN sedes s ON c.id_sede = s.id_sede
    LEFT JOIN docentes d ON c.id_docente = d.id_docente
    WHERE 1=1
    """

    params = []

    # Filtro por curso
    if curso:
        query += " AND IFNULL(c.nombre, '') = ?"
        params.append(curso)

    # Filtro por alumno (nombre, DNI o código)
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # 🔥 Filtro por sede según rol
    if rol != "admin":
        if sedes_ids:  # usuario con varias sedes
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:  # usuario con una sola sede
            query += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    query += " ORDER BY a.id_alumno ASC"

    cur.execute(query, params)
    alumnos = cur.fetchall()
    con.close()

    # ===== PDF EN MEMORIA EN HORIZONTAL =====
    buffer = BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=landscape(letter))

    elements = []
    styles = getSampleStyleSheet()

    cell_style = styles["BodyText"]
    cell_style.fontSize = 6
    cell_style.leading = 8

    titulo = Paragraph("<b>Reporte Completo de Alumnos Matriculados</b>", styles["Title"])
    elements.append(titulo)

    data = [[
        "Fecha Matrícula", 
        "Nombres y Apellidos", 
        "DNI", 
        "Celular",
        "Correo", 
        "Institución",
        "Sede",
        "Curso",
        "Horario",
        "Monto De Matricula"
    ]]

    for a in alumnos:
        # Validar / convertir monto
        monto_raw = a["monto"]
        try:
            monto = float(monto_raw) if monto_raw not in (None, "", " ") else 0.0
        except:
            monto = 0.0

        data.append([
            Paragraph(a["fecha_matricula"] or "-", cell_style),
            Paragraph(a["nombres"] or "-", cell_style),
            Paragraph(a["dni"] or "-", cell_style),
            Paragraph(a["celular"] or "-", cell_style),
            Paragraph(a["correo"] or "-", cell_style),
            Paragraph(a["institucion"] or "-", cell_style),
            Paragraph(a["sede"] or "-", cell_style),
            Paragraph(a["curso"] or "-", cell_style),
            Paragraph(a["horario"] or "-", cell_style),
            Paragraph(f"S/ {monto:.2f}", cell_style)
        ])

    table = Table(data, colWidths=[
        65, 80, 55, 55, 60, 60, 95, 80, 65, 75
    ])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1d3557")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightgrey]),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,1), (-1,-1), 3),
    ]))

    elements.append(table)
    pdf.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_alumnos.pdf",
        mimetype="application/pdf"
    )

#====================================================================
# RUTA ASISTENCIAS DOCENTES
@app.route('/asistencias_docentes')
def asistencias_docentes():
    # 🔐 Solo usuarios logueados
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    sedes_ids = session.get("sedes") or []
    id_sede_usuario = session.get("id_sede")

    # Fallback: si todavía usas solo id_sede
    if not sedes_ids and id_sede_usuario:
        sedes_ids = [id_sede_usuario]

    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    buscar_docente = (request.args.get('buscar_docente') or '').strip()

    if not fecha_desde or not fecha_hasta:
        hoy = datetime.now().strftime("%Y-%m-%d")
        fecha_desde = hoy
        fecha_hasta = hoy

    # Texto que usará {{ fecha }} en el título
    fecha = (
        f"{datetime.strptime(fecha_desde, '%Y-%m-%d').strftime('%d/%m/%Y')}"
        f" al "
        f"{datetime.strptime(fecha_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    )

    # --------- QUERY BASE ---------
    query = """
        SELECT 
            d.codigo,
            d.dni, 
            d.nombres, 
            asi.fecha, 
            asi.hora, 
            asi.observacion,
            c.nombre AS curso,
            COALESCE(h.descripcion, 'Sin horario') AS horario,
            s.nombre_sede AS sede
        FROM asistencias_docentes asi
        JOIN docentes d 
            ON asi.id_docente = d.id_docente
        LEFT JOIN asignaciones_docentes ad 
            ON ad.id_docente = d.id_docente
        LEFT JOIN cursos c 
            ON c.id_curso = ad.id_curso
        LEFT JOIN horarios h
            ON h.id_horario = ad.id_horario
        LEFT JOIN sedes s 
            ON s.id_sede = c.id_sede
        WHERE asi.fecha BETWEEN ? AND ?
    """

    params = [fecha_desde, fecha_hasta]

    # 🔍 Filtro por docente
    if buscar_docente:
        query += " AND (d.nombres LIKE ? OR d.dni LIKE ?)"
        like = f"%{buscar_docente}%"
        params.extend([like, like])

    # 📍 Filtro por sede según el usuario
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario sin sedes asignadas → que no vea nada
            query += " AND 1=0"

    query += " ORDER BY asi.fecha, d.nombres, asi.hora"

    con = sqlite3.connect("base_datos.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    filas = cur.execute(query, params).fetchall()
    con.close()

    return render_template(
        "asistencias_docentes.html",
        asistencias=filas,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        buscar_docente=buscar_docente,
        fecha=fecha,
    )


#====================================================================
# RUTA EXPORTAR ASISTENCIAS DOCENTES A EXCEL
@app.route('/exportar_asistencias_docentes')
def exportar_asistencias_docentes():
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import pandas as pd
    from datetime import datetime

    # 🔐 Solo usuario logueado
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    sedes_ids = session.get("sedes") or []
    id_sede_usuario = session.get("id_sede")

    # Fallback: si solo tienes id_sede
    if not sedes_ids and id_sede_usuario:
        sedes_ids = [id_sede_usuario]

    # Filtros recibidos desde el botón / formulario
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    buscar_docente = request.args.get('buscar_docente', '').strip()

    # Si no mandan fechas, usamos hoy (igual que en la vista normal)
    if not fecha_desde or not fecha_hasta:
        hoy = datetime.now().strftime("%Y-%m-%d")
        fecha_desde = hoy
        fecha_hasta = hoy

    # Texto que se verá en el encabezado del Excel
    fecha_texto = f"Del {fecha_desde} al {fecha_hasta}"

    # SQL base
    query = """
        SELECT 
            d.codigo,              -- código del docente
            d.dni, 
            d.nombres, 
            asi.fecha, 
            asi.hora, 
            asi.observacion,
            c.nombre AS curso,     -- curso donde dicta
            COALESCE(h.descripcion, 'Sin horario') AS horario,
            s.nombre_sede AS sede  -- sede donde dicta
        FROM asistencias_docentes asi
        JOIN docentes d 
            ON asi.id_docente = d.id_docente
        LEFT JOIN asignaciones_docentes ad
            ON ad.id_docente = d.id_docente
        LEFT JOIN cursos c 
            ON c.id_curso = ad.id_curso
        LEFT JOIN horarios h
            ON h.id_horario = ad.id_horario
        LEFT JOIN sedes s 
            ON s.id_sede = c.id_sede
        WHERE asi.fecha BETWEEN ? AND ?
    """

    params = [fecha_desde, fecha_hasta]

    # 🔍 Filtro opcional por docente (nombre o DNI)
    if buscar_docente:
        query += " AND (d.nombres LIKE ? OR d.dni LIKE ?)"
        like = f"%{buscar_docente}%"
        params.extend([like, like])

    # 🔍 Filtro por SEDE según usuario (no admin)
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario sin sedes asignadas → no debe ver nada
            query += " AND 1=0"

    query += " ORDER BY asi.fecha, d.nombres, asi.hora"

    # Ejecutar SQL
    con = get_db_connection()
    cur = con.cursor()
    cur.execute(query, params)
    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    con.close()

    # Si no hay datos, devolvemos mensaje
    if not filas:
        return "⚠️ No se encontraron asistencias de docentes en el rango seleccionado."

    # DataFrame
    df = pd.DataFrame(filas, columns=columnas)

    # Renombrar columnas para el Excel
    df.rename(columns={
        "codigo": "Código",
        "dni": "DNI",
        "nombres": "Docente",
        "sede": "Sede",
        "curso": "Curso",
        "horario": "Horario",
        "fecha": "Fecha",
        "hora": "Hora",
        "observacion": "Observación"
    }, inplace=True)

    # Reordenar columnas
    df = df[[
        "Fecha",
        "Código",
        "DNI",
        "Docente",
        "Sede",
        "Curso",
        "Horario",
        "Hora",
        "Observación"
    ]]

    # Crear Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        hoja = "Asistencias Docentes"
        df.to_excel(writer, index=False, startrow=4, sheet_name=hoja)

        sheet = writer.sheets[hoja]

        # Título
        sheet.merge_cells("A1:I1")
        sheet["A1"] = "REPORTE DE ASISTENCIAS DE DOCENTES"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Institución
        sheet.merge_cells("A2:I2")
        sheet["A2"] = "Institución: JGM - ICILT"
        sheet["A2"].font = Font(bold=True, size=12)
        sheet["A2"].alignment = Alignment(horizontal="center")

        # Rango de fechas
        sheet.merge_cells("A3:I3")
        sheet["A3"] = fecha_texto
        sheet["A3"].alignment = Alignment(horizontal="center")

        # Ajustar ancho de columnas (versión compacta)
        col_widths = {
            "A": 11,  # Fecha
            "B": 20,  # Código
            "C": 12,  # DNI
            "D": 45,  # Docente
            "E": 14,  # Sede
            "F": 45,  # Curso
            "G": 45,  # Horario
            "H": 12,  # Hora
            "I": 25,  # Observación
        }

        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width

    output.seek(0)

    nombre_archivo = f"asistencias_docentes_{fecha_desde}_a_{fecha_hasta}.xlsx"

    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#====================================================================
# RUTA CUMPLEAÑOS ALUMNOS
@app.route('/cumple_alumnos')
def cumple_alumnos():
    con = get_db_connection()
    cur = con.cursor()

    # Fecha actual
    hoy = datetime.now().strftime("%m-%d")

    # Buscar alumnos que cumplan años hoy
    cumple_hoy = cur.execute("""
        SELECT id_alumno, nombres, fecha_nacimiento, celular, correo
        FROM alumnos
        WHERE strftime('%m-%d', fecha_nacimiento) = ?
        ORDER BY nombres ASC
    """, (hoy,)).fetchall()

    # Buscar alumnos que cumplan en los próximos 7 días
    proximos = cur.execute("""
        SELECT id_alumno, nombres, fecha_nacimiento, celular, correo
        FROM alumnos
        WHERE CAST(strftime('%m%d', fecha_nacimiento) AS INTEGER)
              BETWEEN CAST(strftime('%m%d', 'now') AS INTEGER)
              AND CAST(strftime('%m%d', date('now', '+7 day')) AS INTEGER)
        ORDER BY fecha_nacimiento
    """).fetchall()

    con.close()

    return render_template("cumple_alumnos.html", cumple_hoy=cumple_hoy, proximos=proximos)

#====================================================================
# RUTAS PARA GESTIÓN DE ALUMNOS
@app.route('/agregar', methods=['GET', 'POST'])
def agregar():
    # -----------------------------
    #  POST → GUARDAR ALUMNO
    # -----------------------------
    if request.method == "POST":

        # ---------- LECTURA DE CAMPOS ----------
        codigo = (request.form.get("codigo") or "").strip()
        dni = (request.form.get("dni") or "").strip()
        nombres = (request.form.get("nombres") or "").strip()
        apellidos = (request.form.get("apellidos") or "").strip()
        fecha_nacimiento = request.form.get("fecha_nacimiento") or None
        celular = (request.form.get("celular") or "").strip()
        correo = (request.form.get("correo") or "").strip()
        institucion = request.form.get("institucion")
        grado_academico = request.form.get("grado_academico")
        observacion = (request.form.get("observacion") or "").strip()
        estado_alumno = (request.form.get("estado") or "Matriculado (Pago inicial)").strip()
        estado_matricula = (request.form.get("matricula_estado") or "Matriculado (Pago inicial)").strip()
        tipo_pago = (request.form.get("tipo_pago") or "").strip()
        id_curso = request.form.get("id_curso")
        metodo_pago = (request.form.get("metodo_pago") or "").strip()
        saldo_matricula = (request.form.get("saldo_matricula") or "0.0").strip()
        mensualidad = (request.form.get("mensualidad") or "0.0").strip()
        fecha_vencimiento = request.form.get("fecha_vencimiento") or None
        aula = (request.form.get("aula") or "").strip()
        try:
            deuda = float(request.form.get("deuda") or 0.0)
        except:
            deuda = 0.0

        try:
            monto = float(request.form.get("monto") or 0.0)
        except:
            monto = 0.0

        fecha_matricula = datetime.now().strftime("%Y-%m-%d")

        # ---------- VALIDACIONES BÁSICAS ----------
        if not dni or not nombres:
            return redirect(url_for("agregar"))

        if not id_curso or id_curso == "0":
            return redirect(url_for("agregar"))

        # nombre completo que se grabará en la columna nombres
        nombres = f"{nombres} {apellidos}".strip()

        # ---------- DETERMINAR SEDE DEL USUARIO ----------
        rol = session.get("rol")
        id_sede_session = session.get("id_sede")
        sedes_ids = session.get("sedes", [])  # lista de sedes permitidas (para usuarios multi-sede)

        # id_sede_usuario = la sede con la que se grabará alumno/matrícula
        if rol == "admin":
            # admin elige libremente desde el formulario
            id_sede_usuario = request.form.get("id_sede")
        else:
            # usuario normal
            if sedes_ids and len(sedes_ids) > 1:
                # tiene varias sedes → usa la que viene del formulario (pero validamos)
                id_sede_form = request.form.get("id_sede")
                try:
                    id_sede_int = int(id_sede_form)
                except (TypeError, ValueError):
                    return redirect(url_for("agregar"))

                if id_sede_int not in sedes_ids:
                    # intentó guardar en una sede que no le corresponde
                    return redirect(url_for("agregar"))

                id_sede_usuario = id_sede_int
            else:
                # solo una sede → se fuerza a la de la sesión
                id_sede_usuario = id_sede_session

        # ---------- INSERTAR EN BD ----------
        try:
            con = get_db_connection()
            cur = con.cursor()

            # Generar código si no existe
            if not codigo:
                import time
                base = f"ALU{int(time.time())}"
                codigo = base
                cnt = 0
                while True:
                    cur.execute("SELECT 1 FROM alumnos WHERE codigo = ?", (codigo,))
                    if not cur.fetchone():
                        break
                    cnt += 1
                    codigo = f"{base}_{cnt}"

            # Validación de DNI duplicado
            cur.execute("SELECT id_alumno FROM alumnos WHERE dni = ?", (dni,))
            if cur.fetchone():
                con.close()
                return redirect(url_for("agregar", error='dni'))

            # Insertar alumno
            cur.execute("""
                INSERT INTO alumnos (
                    codigo, dni, nombres, fecha_nacimiento, celular, correo,
                    institucion, grado_academico, estado, id_sede
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                codigo, dni, nombres, fecha_nacimiento, celular, correo,
                institucion, grado_academico, estado_alumno, id_sede_usuario
            ))

            id_alumno = cur.lastrowid

            # Insertar matrícula asociada a la sede del usuario
            cur.execute("""
                INSERT INTO matriculas (
                    id_alumno, id_curso, fecha_matricula, deuda,
                    estado, monto, tipo_pago, metodo_pago,
                    saldo_matricula, mensualidad, fecha_vencimiento,
                    observacion, id_sede, aula
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                id_alumno, int(id_curso), fecha_matricula, deuda, estado_matricula,
                monto, tipo_pago, metodo_pago, saldo_matricula, mensualidad,
                fecha_vencimiento, observacion, id_sede_usuario, aula
            ))

            con.commit()
            con.close()

            # ---------- GENERACIÓN DE QR (opcional) ----------
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                try:
                    s.connect(("8.8.8.8", 80))
                    local_ip = s.getsockname()[0]
                except:
                    local_ip = "127.0.0.1"
                finally:
                    s.close()

                qr_url = f"http://{local_ip}:5000" + url_for(
                    'redirigir_a_asistencia',
                    codigo=codigo,
                    _external=False
                )

                qr_folder = os.path.join(STATIC_DIR, "qrcodes")
                os.makedirs(qr_folder, exist_ok=True)
                qr_path = os.path.join(qr_folder, f"{codigo}.png")

                img = qrcode.make(qr_url)
                img.save(qr_path)

            except Exception as e:
                app.logger.exception("Error generando QR: %s", e)

            return redirect(url_for("agregar", exito="1"))

        except Exception as e:
            try:
                con.close()
            except:
                pass
            app.logger.exception("Error al agregar: %s", e)
            return redirect(url_for("agregar"))

    # -----------------------------
    #  GET → Cargar formulario
    # -----------------------------
    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    rol = session.get("rol")
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes", [])
    if rol == "admin":
        # Admin ve todos los cursos y todas las sedes
        cursos = cur.execute("""
            SELECT 
                c.id_curso,
                c.nombre,
                c.id_sede,
                h.id_horario,
                h.dias AS dias,   -- 👈 AHORA sí tienes curso.dias
                COALESCE(
                    h.nombre_horario || ' ' ||
                    COALESCE(h.dias || ' ', '') ||
                    '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                    c.nombre_horario,
                    ''
                ) AS horario_texto
            FROM cursos c
            LEFT JOIN horarios h ON c.id_horario = h.id_horario
            ORDER BY c.nombre
        """).fetchall()


        sedes = cur.execute("""
            SELECT id_sede, nombre_sede
            FROM sedes
            ORDER BY nombre_sede
        """).fetchall()

    else:
        # Usuario normal
        if sedes_ids and len(sedes_ids) > 1:
            # varias sedes
            placeholders = ",".join("?" * len(sedes_ids))

            cursos = cur.execute(f"""
                SELECT 
                    c.id_curso,
                    c.nombre,
                    c.id_sede,
                    h.id_horario,
                    h.dias AS dias,
                    COALESCE(
                        h.nombre_horario || ' ' ||
                        COALESCE(h.dias || ' ', '') ||
                        '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                        c.nombre_horario,
                        ''
                    ) AS horario_texto
                FROM cursos c
                LEFT JOIN horarios h ON c.id_horario = h.id_horario
                WHERE c.id_sede IN ({placeholders})
                ORDER BY c.nombre
            """, sedes_ids).fetchall()

            sedes = cur.execute(f"""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre_sede
            """, sedes_ids).fetchall()

        else:
            # solo una sede
            cursos = cur.execute("""
                SELECT 
                    c.id_curso,
                    c.nombre,
                    c.id_sede,
                    h.id_horario,
                    h.dias AS dias,
                    COALESCE(
                        h.nombre_horario || ' ' ||
                        COALESCE(h.dias || ' ', '') ||
                        '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                        c.nombre_horario,
                        ''
                    ) AS horario_texto
                FROM cursos c
                LEFT JOIN horarios h ON c.id_horario = h.id_horario
                WHERE c.id_sede = ?
                ORDER BY c.nombre
            """, (id_sede_usuario,)).fetchall()

            sedes = cur.execute("""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede = ?
            """, (id_sede_usuario,)).fetchall()

    con.close()

    return render_template(
        "agregar_alumno.html",
        cursos=cursos,
        sedes=sedes,
        id_sede_usuario=id_sede_usuario
    )

#====================================================================================
# RUTA API PARA OBTENER CURSOS POR SEDE (USADO EN AJAX)
@app.route("/api/cursos_por_sede")
def api_cursos_por_sede():
    id_sede = request.args.get("id_sede")
    con = get_db_connection()
    cur = con.cursor()

    cursos = cur.execute("""
        SELECT id_curso, nombre
        FROM cursos
        WHERE id_sede = ?
        ORDER BY nombre
    """, (id_sede,)).fetchall()

    con.close()

    return jsonify([dict(row) for row in cursos])

#====================================================================================
# RUTA GENERAR FOTOCHECK
@app.route('/generar_fotocheck', methods=['GET', 'POST'])
def generar_fotocheck():
    # GET rápido: lista inicial
    if request.method == 'GET':
        con = get_db_connection()
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # --- filtros por rol / sedes ---
        rol = (session.get("rol") or "").strip().lower()
        id_sede_usuario = session.get("id_sede")
        sedes_ids = session.get("sedes") or []

        sql = """
            SELECT a.id_alumno,
                   a.nombres AS nombre_completo,
                   a.dni,
                   a.codigo,
                   c.nombre AS curso,
                   CASE
                       WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                            THEN h.dias || ': ' || UPPER(h.nombre_horario)
                       ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
                   END AS turno,
                   s.nombre_sede AS sede
            FROM alumnos a
            LEFT JOIN matriculas m ON a.id_alumno = m.id_alumno
            LEFT JOIN cursos c     ON m.id_curso = c.id_curso
            LEFT JOIN horarios h   ON c.id_horario = h.id_horario
            LEFT JOIN sedes s      ON c.id_sede = s.id_sede
            WHERE 1=1
        """
        params = []

        # Solo restringimos sedes si NO es admin
        if rol != "admin":
            if sedes_ids:
                placeholders = ",".join("?" * len(sedes_ids))
                sql += f" AND s.id_sede IN ({placeholders})"
                params.extend(sedes_ids)
            elif id_sede_usuario:
                sql += " AND s.id_sede = ?"
                params.append(id_sede_usuario)

        sql += " ORDER BY a.id_alumno DESC LIMIT 30"

        alumnos_initial = cur.execute(sql, params).fetchall()
        cursos = cur.execute("SELECT id_curso, nombre FROM cursos ORDER BY nombre").fetchall()
        sedes = cur.execute("SELECT id_sede, nombre_sede FROM sedes ORDER BY nombre_sede").fetchall()
        con.close()

        try:
            alumnos_initial = [dict(r) for r in alumnos_initial]
        except Exception:
            pass

        return render_template(
            "generar_fotocheck.html",
            alumnos=alumnos_initial,
            cursos=cursos,
            sedes=sedes
        )


    if request.method == 'POST':
        # Aceptar tanto id_alumno como id_docente (cuando el formulario renombra el hidden)
        id_alumno = request.form.get('id_alumno')
        id_docente = request.form.get('id_docente')
        tipo = request.form.get('tipo')  # 'jgm' o 'icilt'

        if not (id_alumno or id_docente) or not tipo:
            return "⚠️ Debes seleccionar una persona (alumno o docente) y formato."

        # Si es docente, obtener datos de la tabla docentes y mapear a la estructura esperada
        is_docente = False
        persona_data = None
        # Log del form para depuración
        try:
            app.logger.debug("generar_fotocheck POST form: %s", dict(request.form))
        except Exception:
            pass
        if id_docente:
            # Normalizar y extraer dígitos por si el valor viene como '123' o ' id:123'
            id_docente_raw = (id_docente or '').strip()
            app.logger.debug("id_docente raw: %r", id_docente_raw)
            if not id_docente_raw:
                return "⚠️ ID de docente inválido."
            m = re.search(r"(\d+)", id_docente_raw)
            if not m:
                return "⚠️ ID de docente inválido."
            id_docente_int = int(m.group(1))
            con = get_db_connection()
            row = con.execute("""
                SELECT dni, nombres AS nombre, especialidad
                FROM docentes
                WHERE id_docente = ?
            """, (id_docente_int,)).fetchone()
            con.close()
            if not row:
                return "No existe docente."
            persona_data = {
                'dni': row['dni'],
                'nombre': row['nombre'],
                'codigo': row['dni'],  
                'curso': row.get('especialidad') or '',
                'turno': '',
                'sede': ''
            }
            is_docente = True
        else:
            # Alumno normal
            try:
                id_alumno_int = int(id_alumno)
            except (TypeError, ValueError):
                return "⚠️ ID de alumno inválido."
            con = get_db_connection()
            persona_data = con.execute("""
                SELECT 
                    a.dni, 
                    a.nombres AS nombre, 
                    IFNULL(a.codigo, '') AS codigo,
                    IFNULL(c.nombre, 'Sin curso') AS curso,
                    CASE
                        WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL THEN h.dias || ': ' || UPPER(h.nombre_horario)
                        ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
                    END AS turno,
                    IFNULL(s.nombre_sede, '') AS sede
                FROM alumnos a
                LEFT JOIN matriculas m ON m.id_alumno = a.id_alumno
                LEFT JOIN cursos c ON c.id_curso = m.id_curso
                LEFT JOIN horarios h ON c.id_horario = h.id_horario
                LEFT JOIN sedes s ON c.id_sede = s.id_sede
                WHERE a.id_alumno = ?
            """, (id_alumno_int,)).fetchone()
            con.close()
            if not persona_data:
                return "❌ Alumno no encontrado."

        # 🖼️ Crear fotocheck reutilizando la función existente
        fondo_path = os.path.join(STATIC_DIR, "fondos", f"{tipo}_template.jpg")
        os.makedirs(os.path.join(STATIC_DIR, "fotochecks"), exist_ok=True)
        output_path = crear_fotocheck_individual(persona_data, fondo_path, tipo)

        # Registrar uso reciente solo para alumnos (no registramos docentes aquí)
        if not is_docente:
            try:
                con = get_db_connection()
                con.execute("CREATE TABLE IF NOT EXISTS fotocheck_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, id_alumno INTEGER, fecha DATETIME DEFAULT CURRENT_TIMESTAMP)")
                con.execute("INSERT INTO fotocheck_logs (id_alumno) VALUES (?)", (id_alumno_int,))
                con.commit()
                con.close()
            except Exception as e:
                app.logger.warning("No se pudo registrar fotocheck log: %s", e)

        return send_file(output_path, as_attachment=True)
    # Fallback improbable
    return render_template("generar_fotocheck.html", alumnos=[], cursos=[], sedes=[])

#===================================================
# FUNCIÓN AUXILIAR PARA CREAR FOTOCHECK INDIVIDUAL
def crear_fotocheck_individual(alumno, fondo_path, tipo):
    # Aceptar sqlite3.Row o dict
    try:
        # sqlite3.Row has keys() -> convert to dict for easier .get usage
        if hasattr(alumno, 'keys') and not isinstance(alumno, dict):
            alumno = dict(alumno)
    except Exception:
        pass

    fondo = Image.open(fondo_path).convert("RGBA")
    draw = ImageDraw.Draw(fondo)
    w, h = fondo.size  # ancho y alto del fondo
    # 🧮 Escala dinámica (basada en resolución real 639x1016)
    base_width = 638  # Ajuste fino: ancho real confirmado 638
    esc = w / base_width  # escala base (ahora usando el ancho real del fotocheck)

    # Fuentes escaladas (ajustadas para mejor legibilidad)
    try:
        # Tamaños de fuente ajustados:
        # - nombre: más pequeño para mejor ajuste
        # - detalles: más pequeño para mejorar legibilidad
        # - sede: mantiene tamaño grande
        font_nombre = ImageFont.truetype("arialbd.ttf", int(30 * esc))  # Era 42
        font_detalle = ImageFont.truetype("arial.ttf", int(20 * esc))   # Era 28
        font_sede = ImageFont.truetype("arial.ttf", int(32 * esc))     # Más grande para sede
    except Exception:
        # Fallback a fuente por defecto si las ttf no están disponibles
        font_nombre = ImageFont.load_default()
        font_detalle = ImageFont.load_default()
        font_sede = ImageFont.load_default()

    # Helper: dividir nombre en dos líneas (solo para campo nombre)
    def split_name_two_lines(name: str):
        if not name:
            return ['']
        words = [w for w in name.strip().split() if w]
        if len(words) <= 2:
            return [' '.join(words)] if len(words) <= 1 else [' '.join(words[:-1]), words[-1]]
        # Balance por caracteres
        total = sum(len(w) for w in words)
        half = total / 2.0
        acc = 0
        best = 1
        for i, w in enumerate(words[:-1], start=1):
            acc += len(w)
            if abs(acc - half) < abs(sum(len(x) for x in words[:best]) - half):
                best = i
        return [' '.join(words[:best]), ' '.join(words[best:])]

    # Helper: wrap genérico una sola línea si cabe, si no partir por palabras
    def wrap_text(draw_obj, text, font, max_width):
        text = (text or '').strip()
        if not text:
            return ['']
        try:
            w_txt = draw_obj.textbbox((0,0), text, font=font)[2]
        except Exception:
            w_txt = font.getsize(text)[0]
        if w_txt <= max_width:
            return [text]
        words = text.split()
        lines = []
        current = ''
        for w in words:
            tentative = (current + ' ' + w).strip() if current else w
            try:
                w_len = draw_obj.textbbox((0,0), tentative, font=font)[2]
            except Exception:
                w_len = font.getsize(tentative)[0]
            if w_len <= max_width:
                current = tentative
            else:
                if current:
                    lines.append(current)
                current = w
        if current:
            lines.append(current)
        return lines

    # Helper principal: ajusta texto a caja, aplica colores y negrita
    def draw_text_fit_box(draw_obj, box, text, formato='jgm', campo='nombre', max_size=None, min_size=8, line_spacing=1.15):
        colors = {
            'jgm': {'nombre': 'black', 'curso': '#ffffff', 'turno': '#0775b0', 'sede': '#0775b0'},
            'icilt': {'nombre': 'black', 'curso': '#ffffff', 'turno': '#0775b0', 'sede': '#0775b0'}
        }
        fill = colors.get(formato, {}).get(campo, 'black')
        x_box, y_box, w_box, h_box = box
        font_path = 'arialbd.ttf'  # todo en negrita según requerimiento
        if max_size is None:
            max_size = int(30 * esc)
        font_size = max_size
        chosen_font = None
        lines = []
        while font_size >= min_size:
            try:
                font_try = ImageFont.truetype(font_path, font_size)
            except Exception:
                font_try = ImageFont.load_default()
            # Determinar líneas según campo
            upper_text = (text or '').upper()
            if campo == 'nombre':
                candidate_lines = split_name_two_lines(upper_text)
            else:
                candidate_lines = wrap_text(draw_obj, upper_text, font_try, w_box)
            # interlineado dinámico: un poco más cuando hay salto de línea
            ls = 1.3 if len(candidate_lines) > 1 else line_spacing
            # Medir bloque
            widths = []
            heights = []
            for ln in candidate_lines:
                try:
                    bbox = draw_obj.textbbox((0,0), ln, font=font_try)
                    w_ln = bbox[2]-bbox[0]
                    h_ln = bbox[3]-bbox[1]
                except Exception:
                    w_ln, h_ln = font_try.getsize(ln)
                widths.append(w_ln)
                heights.append(h_ln)
            block_h = int(sum(heights) * ls)
            if max(widths) <= w_box and block_h <= h_box:
                chosen_font = font_try
                lines = candidate_lines
                line_spacing = ls  # fijar el espaciado elegido
                break
            font_size -= 1
        if chosen_font is None:
            try:
                chosen_font = ImageFont.truetype(font_path, font_size if font_size >= min_size else min_size)
            except Exception:
                chosen_font = ImageFont.load_default()
            lines = [ (text or '').upper()[:15] ]  # fallback
            heights = [chosen_font.getsize(lines[0])[1]]
            widths = [chosen_font.getsize(lines[0])[0]]
            block_h = int(sum(heights) * line_spacing)
        # Re-medición final si no realizada
        measured = []
        for ln in lines:
            try:
                bbox = draw_obj.textbbox((0,0), ln, font=chosen_font)
                measured.append((bbox[2]-bbox[0], bbox[3]-bbox[1]))
            except Exception:
                sz = chosen_font.getsize(ln)
                measured.append(sz)
        block_h = int(sum(h for _,h in measured) * (1.3 if len(lines) > 1 else line_spacing))
        block_w = max(w for w,_ in measured)
        start_y = y_box + max(0, (h_box - block_h)//2)
        cur_y = start_y
        for (ln_w, ln_h), ln in zip(measured, lines):
            ln_x = x_box + max(0, (w_box - ln_w)//2)
            draw_obj.text((ln_x, cur_y), ln, font=chosen_font, fill=fill)
            cur_y += int(ln_h * (1.3 if len(lines) > 1 else line_spacing))

    # Intentar cargar configuración guardada (posiciones en px sobre 800x1000)
    config_path = os.path.join(STATIC_DIR, "config_fotocheck.json")
    format_cfg = None
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
                format_cfg = cfg.get(tipo)
        except Exception:
            format_cfg = None

    # Helper para escalar posiciones
    def s(v):
        return int(v * esc)

    # Si hay configuración, usar posiciones exactas
    if format_cfg:
        # FOTO
        foto_cfg = format_cfg.get('foto')
        if foto_cfg:
            foto_path = os.path.join(STATIC_DIR, "fotos_alumnos", f"{alumno['codigo']}.jpg")
            
            # ⚠️ IMPORTANTE: Los valores del JSON están en píxeles del template REAL
            # NO necesitan escalarse porque trabajamos con el template a tamaño original
            x_f = foto_cfg.get('x', 0)
            y_f = foto_cfg.get('y', 0)
            
            # Obtener dimensiones del contenedor desde configuración (SIN ESCALAR)
            if 'width' in foto_cfg and 'height' in foto_cfg:
                # Usar valores exactos del JSON (ya están en píxeles correctos)
                w_f = foto_cfg['width']
                h_f = foto_cfg['height']
            else:
                # Fallback: defaults escalados
                w_f = int(w * 0.35)
                h_f = int(w_f * 1.25)
            
            if os.path.exists(foto_path):
                try:
                    foto_original = Image.open(foto_path).convert('RGB')
                    
                    # Obtener dimensiones originales
                    foto_w, foto_h = foto_original.size
                    
                    # Log de depuración
                    app.logger.info(f"📸 Foto original: {foto_w}x{foto_h} | Contenedor: {w_f}x{h_f}")
                    
                    # Calcular la proporción de la imagen original y del contenedor
                    foto_ratio = foto_w / foto_h
                    contenedor_ratio = w_f / h_f
                    
                    app.logger.info(f"📐 Ratio foto: {foto_ratio:.3f} | Ratio contenedor: {contenedor_ratio:.3f}")
                    
                    # Redimensionar manteniendo proporción para que cubra el área completamente
                    if foto_ratio > contenedor_ratio:
                        # Foto más ancha que el contenedor: ajustar por altura y recortar ancho
                        new_h = h_f
                        new_w = int(h_f * foto_ratio)
                    else:
                        # Foto más alta que el contenedor: ajustar por ancho y recortar altura
                        new_w = w_f
                        new_h = int(w_f / foto_ratio)
                    
                    app.logger.info(f"🔧 Redimensionar a: {new_w}x{new_h} | Recortar a: {w_f}x{h_f}")
                    
                    # Redimensionar manteniendo proporción original
                    foto_resized = foto_original.resize((new_w, new_h), Image.Resampling.LANCZOS)
                    
                    # Recortar el centro para ajustar exactamente a w_f x h_f (sin distorsión)
                    left = (new_w - w_f) // 2
                    top = (new_h - h_f) // 2
                    foto_final = foto_resized.crop((left, top, left + w_f, top + h_f))
                    
                    app.logger.info(f"✅ Foto final pegada en: ({x_f}, {y_f})")
                    fondo.paste(foto_final, (x_f, y_f))
                except Exception as e:
                    app.logger.warning(f"Error procesando foto: {e}")
                    pass
            else:
                # 🖼️ PLACEHOLDER cuando no hay foto
                try:
                    # Crear imagen placeholder gris con texto
                    placeholder = Image.new('RGB', (w_f, h_f), color='#cccccc')
                    draw_ph = ImageDraw.Draw(placeholder)
                    
                    # Texto del placeholder
                    try:
                        font_ph = ImageFont.truetype("arial.ttf", int(w_f * 0.1))
                    except Exception:
                        font_ph = ImageFont.load_default()
                    
                    text = "SIN FOTO"
                    try:
                        bbox = draw_ph.textbbox((0, 0), text, font=font_ph)
                        text_w = bbox[2] - bbox[0]
                        text_h = bbox[3] - bbox[1]
                    except Exception:
                        text_w, text_h = font_ph.getsize(text)
                    
                    text_x = (w_f - text_w) // 2
                    text_y = (h_f - text_h) // 2
                    draw_ph.text((text_x, text_y), text, fill='#666666', font=font_ph)
                    
                    fondo.paste(placeholder, (x_f, y_f))
                    app.logger.info(f"📷 Placeholder colocado en ({x_f}, {y_f})")
                except Exception as e:
                    app.logger.warning(f"Error creando placeholder: {e}")
  
        # QR (misma lógica para alumnos y docentes: URL con IP + código)
        qr_cfg = format_cfg.get('qr')

        # Determinar código y rol
        codigo_qr = alumno.get('codigo') or alumno.get('dni') or ''
        rol = alumno.get('rol', 'alumno')   # por defecto, alumno

        if qr_cfg and codigo_qr:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                try:
                    sock.connect(('8.8.8.8', 80))
                    local_ip = sock.getsockname()[0]
                except Exception:
                    local_ip = '127.0.0.1'
                finally:
                    sock.close()

                # 👇 Ruta según rol
                if rol == 'docente':
                    path = f"/registrar_asistencia_docente?codigo={codigo_qr}"
                else:
                    path = f"/registrar_asistencia?codigo={codigo_qr}"

                qr_data = f"http://{local_ip}:5000{path}"
            except Exception:
                qr_data = codigo_qr  # fallback

            x_q = s(qr_cfg.get('x', 0))
            y_q = s(qr_cfg.get('y', 0))
            w_q = s(qr_cfg.get('width', int(w * 0.2)))
            h_q = s(qr_cfg.get('height', w_q))
            try:
                import qrcode
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=1
                )
                qr.add_data(qr_data)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                qr_img = qr_img.resize((w_q, h_q))
                fondo.paste(qr_img, (x_q, y_q))
            except Exception:
                pass

        # TEXTOS: nombre, turno, curso, sede
        for key, font in (('nombre', font_nombre), ('turno', font_detalle), ('curso', font_detalle), ('sede', font_sede)):
            tcfg = format_cfg.get(key)
            if not tcfg:
                continue
            x_t = s(tcfg.get('x', 0))
            y_t = s(tcfg.get('y', 0))
            w_t = s(tcfg.get('width', 100))
            h_t = s(tcfg.get('height', 30))
            text = ''
            if key == 'nombre':
                text = alumno.get('nombre','').upper()
            elif key == 'turno':
                text = alumno.get('turno','')
            elif key == 'curso':
                text = alumno.get('curso','')
            elif key == 'sede':
                text = alumno.get('sede','')
            try:
                draw_text_fit_box(draw, (x_t, y_t, w_t, h_t), text, formato=tipo, campo=key, max_size=int(30 * esc), min_size=10)
            except Exception:
                draw.text((x_t, y_t), text, fill='black', font=font)

    else:
        # Comportamiento legacy (sin configuración): posición relativa
        # Inicializar variables para posicionamiento
        y_foto = int(h * 0.15)  # 15% desde arriba
        foto_height = int(h * 0.35)  # Altura predeterminada si no hay foto

        # 📸 Foto del alumno con proporción fija y centrado
        foto_path = os.path.join(STATIC_DIR, "fotos_alumnos", f"{alumno['codigo']}.jpg")
        if os.path.exists(foto_path):
            try:
                # Cargar y preparar la foto
                foto_original = Image.open(foto_path)
                # Calcular dimensiones para mantener proporción 4:5
                foto_width = int(w * 0.35)  # 35% del ancho del fondo
                foto_height = int(foto_width * 1.25)  # Proporción 4:5
                
                # Recortar la foto original a proporción 4:5 si es necesario
                w_orig, h_orig = foto_original.size
                if w_orig/h_orig > 0.8:  # Si es más ancha que 4:5
                    new_w = int(h_orig * 0.8)
                    left = (w_orig - new_w) // 2
                    foto_original = foto_original.crop((left, 0, left + new_w, h_orig))
                else:  # Si es más alta que 4:5
                    new_h = int(w_orig * 1.25)
                    top = (h_orig - new_h) // 2
                    foto_original = foto_original.crop((0, top, w_orig, top + new_h))
                
                # Redimensionar a tamaño final
                foto = foto_original.resize((foto_width, foto_height), Image.Resampling.LANCZOS)
                
                # Centrar la foto horizontalmente
                x_foto = (w - foto_width) // 2
                fondo.paste(foto, (x_foto, y_foto))
            except Exception as e:
                print(f"Error al procesar la foto: {e}")
                # Si hay error, continuamos sin foto
                pass

        # 📝 Texto centrado y espaciado optimizado (legacy)
        nombre = alumno.get("nombre","" ).upper()
        # Nombre
        box_x = int(w * 0.1)
        box_w = int(w * 0.8)
        box_y = y_foto + foto_height + int(h * 0.03)
        box_h = int(h * 0.12)
        draw_text_fit_box(draw, (box_x, box_y, box_w, box_h), nombre, formato=tipo, campo='nombre', max_size=int(30 * esc), min_size=10)
        # Turno
        turno_box_y = box_y + box_h + int(h * 0.005)
        draw_text_fit_box(draw, (box_x, turno_box_y, box_w, int(h * 0.06)), alumno.get("turno",""), formato=tipo, campo='turno', max_size=int(20 * esc), min_size=8)
        # Curso
        curso_box_y = turno_box_y + int(h * 0.06) + int(h * 0.005)
        draw_text_fit_box(draw, (box_x, curso_box_y, box_w, int(h * 0.06)), alumno.get("curso",""), formato=tipo, campo='curso', max_size=int(20 * esc), min_size=8)
        # Sede
        sede_box_y = curso_box_y + int(h * 0.06) + int(h * 0.005)
        draw_text_fit_box(draw, (box_x, sede_box_y, box_w, int(h * 0.06)), alumno.get("sede",""), formato=tipo, campo='sede', max_size=int(32 * esc), min_size=10)

#== QR centrado en la parte inferior (legacy)
###codigo bueno y mejor para crear qr con el codigo 
        # QR legacy (solo URL de redirección)
        codigo_qr = alumno.get('codigo') or ''
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            try:
                sock.connect(('8.8.8.8', 80))
                local_ip = sock.getsockname()[0]
            except Exception:
                local_ip = '127.0.0.1'
            finally:
                sock.close()
            # Solo URL sin DNI
            qr_payload = f"http://{local_ip}:5000/r/{codigo_qr}" if codigo_qr else codigo_qr
        except Exception:
            # Fallback: solo código
            qr_payload = codigo_qr
        qr_tamaño = int(w * 0.25)
        qr_img = qrcode.make(qr_payload).resize((qr_tamaño, qr_tamaño))
        x_qr = (w - qr_tamaño) // 2
        y_qr = int(h * 0.75)
        fondo.paste(qr_img, (x_qr, y_qr))

    # 💾 Guardar imagen con mejor calidad
    os.makedirs(os.path.join(STATIC_DIR, "fotochecks"), exist_ok=True)
    output_path = os.path.join(STATIC_DIR, "fotochecks", f"{alumno.get('dni','')}_{tipo}.png")
    fondo.save(output_path, quality=95, optimize=True)
    return output_path

#====================================================================
# RUTAS ASISTENCIA DOCENTES
@app.route('/registrar_asistencia_docente')
def registrar_asistencia_docente():
    codigo = request.args.get("codigo")
    if not codigo:
        return "Código no recibido", 400

    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    hora_actual = datetime.now().strftime("%H:%M:%S")

    with get_db_connection() as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # 1) Buscar DOCENTE por codigo
        docente = cur.execute("""
            SELECT 
                d.id_docente,
                d.codigo,
                d.nombres,
                d.profesion,
                d.especialidad,
                COALESCE(s.nombre_sede, '') AS sede,
                COALESCE(h.nombre_horario, '') AS nombre_horario,
                COALESCE(h.dias, '') AS dias,
                COALESCE(h.hora_inicio, '') AS h_inicio,
                COALESCE(h.hora_fin, '') AS h_fin,
                COALESCE(h.tolerancia_inicio, 0) AS tolerancia,
                COALESCE(h.entrada_temprana, 0) AS entrada_temprana,
                COALESCE(h.salida_temprana, 0) AS salida_temprana
            FROM docentes d
            LEFT JOIN asignaciones_docentes ad
                ON ad.id_docente = d.id_docente          -- 👈 SIN ad.estado
            LEFT JOIN cursos c
                ON c.id_curso = ad.id_curso
            LEFT JOIN horarios h
                ON h.id_horario = ad.id_horario
            LEFT JOIN sedes s
                ON c.id_sede = s.id_sede
            WHERE d.codigo = ?
            LIMIT 1
        """, (codigo,)).fetchone()


        if not docente:
            return "❌ Docente no registrado en el sistema."

        # ====== LÓGICA HORARIO / ENTRADA / SALIDA ======
        hora_inicio = (docente["h_inicio"] or "").strip()
        hora_fin    = (docente["h_fin"] or "").strip()
        tolerancia  = docente["tolerancia"]
        entrada_temprana = int(docente["entrada_temprana"] or 0)
        salida_temprana  = int(docente["salida_temprana"] or 0)

        cumple_hoy = False   # si luego quieres cumpleaños para docentes, aquí se calcula

        def _fmt(hhmmss: str) -> str:
            return "%H:%M:%S" if len((hhmmss or "").strip()) > 5 else "%H:%M"

        def to_dt_today(hhmmss: str) -> datetime:
            fmt = _fmt(hhmmss)
            t = datetime.strptime(hhmmss, fmt)
            today = datetime.now()
            return today.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0), microsecond=0)

        # Registros del día
        registros = cur.execute("""
            SELECT id_asistencia_doc, fecha, hora, observacion
            FROM asistencias_docentes
            WHERE id_docente = ? AND fecha = ?
            ORDER BY hora ASC
        """, (docente["id_docente"], fecha_hoy)).fetchall()

        ya_tiene_entrada = any((r["observacion"] or "").startswith("Entrada") for r in registros)
        ya_tiene_salida  = any("Salida" in (r["observacion"] or "") for r in registros)

        observacion = None

        if hora_inicio and hora_fin:
            h_actual_full  = to_dt_today(hora_actual)
            h_inicio_full  = to_dt_today(hora_inicio)
            h_fin_full     = to_dt_today(hora_fin)

            h_inicio_temprano   = h_inicio_full - timedelta(minutes=entrada_temprana)
            h_salida_permitida  = h_fin_full - timedelta(minutes=salida_temprana)

            # 1) Primera marca
            if not registros:
                if h_actual_full < h_inicio_temprano:
                    datos = {
                        "codigo": docente["codigo"],
                        "nombres": docente["nombres"],
                        "curso": docente["especialidad"] or '',
                        "turno": f"{hora_inicio or '—'} - {hora_fin or '—'}",
                        "fecha": fecha_hoy,
                        "hora": hora_actual,
                        "estado_asistencia": "⛔ Fuera del horario permitido",
                        "cumple_hoy": cumple_hoy,
                        "sede": docente["sede"] or "No asignada",
                    }
                    return render_template("registro_exitoso_docente.html", datos=datos)

                if h_actual_full <= h_fin_full:
                    # usa tu misma función para evaluar entrada
                    observacion = evaluar_entrada(
                        hora_actual,
                        hora_inicio,
                        tolerancia_min=tolerancia,
                        entrada_temprana_min=entrada_temprana
                    )
                else:
                    observacion = "Salida ✅"

            # 2) Ya tiene entrada, no salida
            elif ya_tiene_entrada and not ya_tiene_salida:
                if h_actual_full >= h_salida_permitida:
                    observacion = (
                        "Salida ✅ (Temprana)" if h_actual_full < h_fin_full
                        else "Salida ✅"
                    )
                else:
                    datos = {
                        "codigo": docente["codigo"],
                        "nombres": docente["nombres"],
                        "curso": docente["especialidad"] or '',
                        "turno": f"{hora_inicio or '—'} - {hora_fin or '—'}",
                        "fecha": fecha_hoy,
                        "hora": hora_actual,
                        "estado_asistencia": "⛔ Aún no puedes registrar salida",
                        "cumple_hoy": cumple_hoy,
                        "sede": docente["sede"] or "No asignada",
                    }
                    return render_template("registro_exitoso_docente.html", datos=datos)

            # 3) Ya tiene entrada y salida
            elif ya_tiene_entrada and ya_tiene_salida:
                datos = {
                    "codigo": docente["codigo"],
                    "nombres": docente["nombres"],
                    "curso": docente["especialidad"] or '',
                    "turno": f"{hora_inicio or '—'} - {hora_fin or '—'}",
                    "fecha": fecha_hoy,
                    "hora": hora_actual,
                    "estado_asistencia": "Salida ya registrada",
                    "cumple_hoy": cumple_hoy,
                    "sede": docente["sede"] or "No asignada",
                }
                return render_template("registro_exitoso_docente.html", datos=datos)

            else:
                # salvaguarda
                observacion = evaluar_entrada(
                    hora_actual,
                    hora_inicio,
                    tolerancia_min=tolerancia,
                    entrada_temprana_min=entrada_temprana
                )
        else:
            # sin horario configurado → marcar simple
            observacion = "Entrada ✅"

        # INSERT en asistencias_docentes
        cur.execute("""
            INSERT INTO asistencias_docentes (id_docente, fecha, hora, observacion)
            VALUES (?, ?, ?, ?)
        """, (docente["id_docente"], fecha_hoy, hora_actual, observacion))
        con.commit()

        datos = {
            "codigo": docente["codigo"],
            "nombres": docente["nombres"],
            "curso": docente["especialidad"] or '',
            "turno": f"{hora_inicio or '—'} - {hora_fin or '—'}",
            "fecha": fecha_hoy,
            "hora": hora_actual,
            "estado_asistencia": observacion,
            "cumple_hoy": cumple_hoy,
            "sede": docente["sede"] or "No asignada",
        }

    # 👇 AQUÍ YA NO HAY REDIRECT AL LISTADO
    return render_template("registro_exitoso_docente.html", datos=datos)

@app.route('/api/alumnos_search')
def api_alumnos_search():
    q        = (request.args.get('q') or '').strip()
    curso_id = (request.args.get('curso') or '').strip()
    sede_id_filtro = (request.args.get('sede') or '').strip()

    try:
        limit = max(1, min(int(request.args.get('limit', 30)), 200))
    except ValueError:
        limit = 30
    try:
        offset = max(0, int(request.args.get('offset', 0)))
    except ValueError:
        offset = 0

    # ---- CONEXIÓN + CONFIG BÁSICA ----
    con = get_db_connection()
    con.row_factory = sqlite3.Row

    # Normalización para búsquedas sin tildes
    import unicodedata
    def remove_accents(s):
        if s is None:
            return ''
        return ''.join(
            c for c in unicodedata.normalize('NFD', s)
            if unicodedata.category(c) != 'Mn'
        )

    def normalize_sqlite(text):
        return remove_accents(text.lower() if text else '')

    con.create_function("NORMALIZE", 1, normalize_sqlite)

    # ---- FILTRO POR USUARIO / SEDES ----
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []

    where = []
    params = []

    # Solo restringimos sedes si NO es admin
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            where.append(f"s.id_sede IN ({placeholders})")
            params.extend(sedes_ids)
        elif id_sede_usuario:
            where.append("s.id_sede = ?")
            params.append(id_sede_usuario)

    # ---- BÚSQUEDA POR TEXTO ----
    if q:
        q_normalized = remove_accents(q.lower())
        like = f"%{q_normalized}%"
        where.append("""
            (NORMALIZE(a.nombres) LIKE ? 
             OR NORMALIZE(a.dni) LIKE ? 
             OR NORMALIZE(a.codigo) LIKE ?)
        """)
        params.extend([like, like, like])

    # ---- FILTRO POR CURSO (opcional) ----
    if curso_id and curso_id.isdigit():
        where.append("c.id_curso = ?")
        params.append(int(curso_id))

    # ---- FILTRO POR SEDE EN LA QUERYSTRING (opcional) ----
    # Esto se cruza con las sedes permitidas por arriba.
    if sede_id_filtro and sede_id_filtro.isdigit():
        where.append("s.id_sede = ?")
        params.append(int(sede_id_filtro))

    # ---- SQL FINAL ----
    sql = """
        SELECT 
            a.id_alumno,
            a.nombres AS nombre_completo,
            a.dni,
            a.codigo,
            c.nombre AS curso,
            CASE
                WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                    THEN h.dias || ': ' || UPPER(h.nombre_horario)
                ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
            END AS turno,
            s.nombre_sede AS sede
        FROM alumnos a
        LEFT JOIN matriculas m ON a.id_alumno = m.id_alumno
        LEFT JOIN cursos c      ON m.id_curso  = c.id_curso
        LEFT JOIN horarios h    ON c.id_horario = h.id_horario
        LEFT JOIN sedes s       ON c.id_sede   = s.id_sede
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY a.nombres LIMIT ? OFFSET ?"
    params.extend([limit, offset])

    filas = con.execute(sql, params).fetchall()
    con.close()

    try:
        data = [dict(r) for r in filas]
    except Exception:
        data = []

    return jsonify({
        "results": data,
        "count": len(data),
        "limit": limit,
        "offset": offset
    })

@app.route('/api/alumnos_recientes')
def api_alumnos_recientes():
    try:
        con = get_db_connection()
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # Asegurar tabla de logs
        cur.execute("""
            CREATE TABLE IF NOT EXISTS fotocheck_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_alumno INTEGER,
                fecha DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # ---- FILTRO POR USUARIO / SEDES ----
        rol = (session.get("rol") or "").strip().lower()
        id_sede_usuario = session.get("id_sede")
        sedes_ids = session.get("sedes") or []

        sql = """
            SELECT 
                l.id_alumno, 
                l.fecha,
                a.nombres AS nombre_completo,
                a.dni,
                a.codigo,
                c.nombre AS curso,
                CASE
                    WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                        THEN h.dias || ': ' || UPPER(h.nombre_horario)
                    ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
                END AS turno,
                s.nombre_sede AS sede
            FROM fotocheck_logs l
            JOIN alumnos a          ON l.id_alumno = a.id_alumno
            LEFT JOIN matriculas m  ON a.id_alumno = m.id_alumno
            LEFT JOIN cursos c      ON m.id_curso  = c.id_curso
            LEFT JOIN horarios h    ON c.id_horario = h.id_horario
            LEFT JOIN sedes s       ON c.id_sede   = s.id_sede
            WHERE 1=1
        """

        params = []

        if rol != "admin":
            # Usuario normal → limitar a sus sedes
            if sedes_ids:
                placeholders = ",".join("?" * len(sedes_ids))
                sql += f" AND c.id_sede IN ({placeholders})"
                params.extend(sedes_ids)
            elif id_sede_usuario:
                sql += " AND c.id_sede = ?"
                params.append(id_sede_usuario)

        sql += " ORDER BY l.fecha DESC LIMIT 3"

        filas = cur.execute(sql, params).fetchall()
        con.close()

        recientes = [dict(r) for r in filas]
    except Exception as e:
        app.logger.warning("Error obteniendo recientes: %s", e)
        recientes = []

    return jsonify({"recientes": recientes})


#=================================
# Página: Editor de fotocheck
@app.route('/editor_fotocheck')
def editor_fotocheck():
    return render_template('editor_fotocheck.html')

#=================================
# API: Obtener configuración fotocheck
@app.route('/get_fotocheck_config')
def get_fotocheck_config():
    config_path = os.path.join(STATIC_DIR, "config_fotocheck.json")
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            return jsonify(json.load(f))
    return jsonify({})

#=================================
# API: Guardar configuración fotocheck
@app.route('/save_fotocheck_config', methods=['POST'])
def save_fotocheck_config():
    config_path = os.path.join(STATIC_DIR, "config_fotocheck.json")
    config = request.get_json()
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=4)
    return jsonify({"success": True})

#=================================
# API: Subir foto de alumno
@app.route('/upload_student_photo/<codigo>', methods=['POST'])
def upload_student_photo(codigo):
    if 'foto' not in request.files:
        return jsonify({"success": False, "error": "No se envió ninguna foto"})
    
    foto = request.files['foto']
    if foto.filename == '':
        return jsonify({"success": False, "error": "No se seleccionó ningún archivo"})
    
    if foto:
        # Asegurar que existe el directorio
        os.makedirs(os.path.join(STATIC_DIR, "fotos_alumnos"), exist_ok=True)
        
        # Guardar la foto con el código del alumno
        filename = f"{codigo}.jpg"
        filepath = os.path.join(STATIC_DIR, "fotos_alumnos", filename)
        
        # Procesar y guardar la imagen
        try:
            # Abrir la imagen con Pillow
            img = Image.open(foto.stream)
            
            # Convertir a RGB si es necesario
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # 🔲 GUARDAR COMO CUADRADO (800x800) para que coincida con contenedores cuadrados
            target_size = 800
            
            # Calcular dimensiones para el recorte centrado
            if img.width > img.height:
                # Imagen horizontal: recortar ancho
                new_width = img.height
                left = (img.width - new_width) // 2
                img = img.crop((left, 0, left + new_width, img.height))
            elif img.height > img.width:
                # Imagen vertical: recortar altura
                new_height = img.width
                top = (img.height - new_height) // 2
                img = img.crop((0, top, img.width, top + new_height))
            # Si es cuadrada, no recortar
            
            # Redimensionar a tamaño final cuadrado
            img = img.resize((target_size, target_size), Image.Resampling.LANCZOS)
            
            # Guardar con calidad optimizada
            img.save(filepath, 'JPEG', quality=85, optimize=True)
            
            return jsonify({
                "success": True,
                "message": "Foto actualizada correctamente",
                "url": f"/static/fotos_alumnos/{filename}?v={int(time.time())}"
            })
            
        except Exception as e:
            return jsonify({
                "success": False,
                "error": f"Error al procesar la imagen: {str(e)}"
            })
    
    return jsonify({"success": False, "error": "Error al guardar la foto"})

#=================================
# API: Búsqueda dinámica de docentes con filtros y paginado
@app.route("/api/docentes_search")
def api_docentes_search():
    q      = request.args.get("q", "").strip()
    id_cur = request.args.get("curso") or None
    id_sed = request.args.get("sede") or None
    limit  = int(request.args.get("limit", 30))
    offset = int(request.args.get("offset", 0))

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # ---- FILTRO POR USUARIO / SEDES ----
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []

    sql = """
        SELECT
            d.id_docente,
            d.nombres AS nombre_completo,
            d.dni,
            d.codigo,
            COALESCE(c.nombre, '') AS curso,
            CASE
                WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                    THEN h.dias || ': ' || UPPER(h.nombre_horario)
                ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
            END AS turno,
            COALESCE(s.nombre_sede, '') AS sede
        FROM docentes d
        LEFT JOIN asignaciones_docentes ad ON ad.id_docente = d.id_docente
        LEFT JOIN cursos   c ON c.id_curso  = ad.id_curso
        LEFT JOIN horarios h ON h.id_horario = ad.id_horario
        LEFT JOIN sedes    s ON s.id_sede   = c.id_sede
        WHERE 1=1
    """

    params = []

    # 🔒 Restringir por sedes asignadas al usuario (si NO es admin)
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            sql += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        elif id_sede_usuario:
            sql += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    # 🔍 Búsqueda por texto
    if q:
        sql += " AND (d.nombres LIKE ? OR d.dni LIKE ? OR d.codigo LIKE ?)"
        like = f"%{q}%"
        params.extend([like, like, like])

    # 🎯 Filtro por curso (opcional)
    if id_cur:
        sql += " AND c.id_curso = ?"
        params.append(id_cur)

    # 🎯 Filtro por sede desde el front (opcional, se cruza con las sedes permitidas)
    if id_sed:
        sql += " AND c.id_sede = ?"
        params.append(id_sed)

    sql += " ORDER BY d.nombres LIMIT ? OFFSET ?"
    params.extend([limit, offset])

    rows = cur.execute(sql, params).fetchall()
    con.close()

    return jsonify(
        results=[dict(r) for r in rows],
        count=len(rows)
    )

#=================================
# GENERAR FOTOCHECK DOCENTE
@app.route('/generar_fotocheck_docente', methods=['POST'])
def generar_fotocheck_docente():
    # Aceptar id_docente desde el form
    id_docente = request.form.get('id_docente')
    tipo = request.form.get('tipo')  # 'jgm' o 'icilt'

    if not id_docente or not tipo:
        return "⚠️ Debes seleccionar un docente y formato."

    # NORMALIZAR ID (por si viene con texto alrededor)
    id_docente_raw = str(id_docente).strip()
    id_docente_digits = re.sub(r"\D", "", id_docente_raw)

    if not id_docente_digits:
        return "⚠️ ID de docente inválido."

    id_docente_int = int(id_docente_digits)

    # 🔍 CONSULTAR DOCENTE + CURSO + HORARIO + SEDE
    con = get_db_connection()
    row = con.execute("""
        SELECT 
            d.dni,
            d.nombres AS nombre,
            d.especialidad,
            CASE
                WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                    THEN h.dias || ': ' || UPPER(h.nombre_horario)
                ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
            END AS turno,
            IFNULL(s.nombre_sede, '') AS sede
        FROM docentes d
        LEFT JOIN horarios h ON h.id_horario = d.id_horario   -- 👈 OJO AQUÍ
        LEFT JOIN sedes s    ON s.id_sede = d.id_sede         -- si aplica
        WHERE d.id_docente = ?
        LIMIT 1
    """, (id_docente_int,)).fetchone()

    con.close()

    if not row:
        return "❌ Docente no encontrado."

    # Mapear a lo que espera crear_fotocheck_individual
    docente = {
        "dni": row["dni"],
        "nombre": row["nombre"],
        "codigo": row["dni"],                 # aquí usas el DNI como código para foto/QR
        "curso": row["especialidad"] or "",   # línea de profesión/especialidad
        "turno": row["turno"] or "",          # 🔹 ahora con días + nombre_horario
        "sede": row["sede"] or "",            # 🔹 sede del curso
        "rol": "docente"                      # para que el QR use la ruta de docentes
    }

    fondo_path = os.path.join(STATIC_DIR, "fondos", f"{tipo}_template.jpg")
    output = crear_fotocheck_individual(docente, fondo_path, tipo)
    return send_file(output, as_attachment=True)

#=================================
# GENERAR QR DOCENTE
@app.route('/generar_qr_docente', methods=['POST'])
def generar_qr_docente():
    id_docente = request.form.get('id_docente')
    tipo = request.form.get('tipo')  # jgm o icilt

    if not id_docente or not tipo:
        return "⚠️ Debes seleccionar un docente y un formato."

    # Normalizar ID
    id_docente_raw = str(id_docente).strip()
    id_docente_digits = re.sub(r"\D", "", id_docente_raw)
    if not id_docente_digits:
        return "⚠️ ID de docente inválido."

    id_docente_int = int(id_docente_digits)

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    row = con.execute("""
        SELECT 
            d.id_docente,
            d.dni,
            d.codigo,
            d.nombres AS nombre,
            d.especialidad,

            -- TURNO (días + nombre_horario del horario asignado)
            CASE
                WHEN h.dias IS NOT NULL AND h.nombre_horario IS NOT NULL 
                    THEN h.dias || ': ' || UPPER(h.nombre_horario)
                ELSE COALESCE(h.descripcion, h.nombre_horario, 'Sin horario')
            END AS turno,

            -- SEDE DEL CURSO AL QUE ESTÁ ASIGNADO
            COALESCE(s.nombre_sede, '') AS sede

        FROM docentes d
        LEFT JOIN asignaciones_docentes ad
               ON ad.id_docente = d.id_docente           -- 🔗 relación real
        LEFT JOIN cursos c
               ON c.id_curso = ad.id_curso               -- para llegar a la sede
        LEFT JOIN horarios h
               ON h.id_horario = ad.id_horario           -- horario asignado
        LEFT JOIN sedes s
               ON s.id_sede = c.id_sede                  -- sede del curso
        WHERE d.id_docente = ?
        LIMIT 1
    """, (id_docente_int,)).fetchone()
    con.close()

    if not row:
        return "❌ Docente no encontrado."

    docente = {
        "dni": row["dni"],
        "nombre": row["nombre"],
        "codigo": row["codigo"],          # para QR/foto
        "curso": row["especialidad"] or "",
        "turno": row["turno"] or "",      # ✅ ahora con el horario real
        "sede": row["sede"] or "",        # ✅ sede real
        "rol":  "docente"
    }

    fondo_path = os.path.join(STATIC_DIR, "fondos", f"{tipo}_template.jpg")
    os.makedirs(os.path.join(STATIC_DIR, "fotochecks"), exist_ok=True)
    output_path = crear_fotocheck_individual(docente, fondo_path, tipo)
    return send_file(output_path, as_attachment=True)

# ---------------------------
# Nuevas rutas para cursos
# ---------------------------
# Use default endpoint name "listar_cursos" so url_for('listar_cursos') works in the code/templates
# Cambiar decorator para exponer endpoint 'cursos' (alternativa si no quieres tocar las plantillas)

#=================================
# LISTAR CURSOS 

@app.route('/cursos', endpoint='cursos')
def listar_cursos():
    con = get_db_connection()
    cursos = con.execute("SELECT * FROM cursos ORDER BY nombre").fetchall()
    con.close()
    return render_template("cursos.html", cursos=cursos)

#=================================
# AGREGAR CURSO
@app.route('/agregar_curso', methods=['GET', 'POST'])
def agregar_curso():
    # ---------------- POST: GUARDAR CURSO ----------------
    if request.method == 'POST':

        nombre       = (request.form.get('nombre') or "").strip()
        descripcion  = (request.form.get('descripcion') or "").strip()
        id_docente   = request.form.get('id_docente') or None
        id_sede_form = request.form.get('id_sede') or None
        aula         = (request.form.get('aula') or "").strip()
        id_horario   = request.form.get('id_horario') or None
        estado       = (request.form.get('estado') or "Activo").strip()
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin    = request.form.get('fecha_fin')

        # Validación mínima
        if not nombre or not fecha_inicio or not fecha_fin:
            return redirect(url_for('agregar_curso'))

        # --------- LÓGICA DE SEDE SEGÚN USUARIO ----------
        rol          = session.get("rol")
        sedes_ids    = session.get("sedes", [])   # lista de sedes permitidas
        id_sede_sess = session.get("id_sede")

        # Determinamos la sede final con la que se guardará el curso
        id_sede_final = None

        if rol == "admin":
            # admin elige libremente
            id_sede_final = id_sede_form
        else:
            if sedes_ids and len(sedes_ids) > 1:
                # usuario con varias sedes → debe venir una sede válida del form
                try:
                    id_sede_int = int(id_sede_form)
                except (TypeError, ValueError):
                    return redirect(url_for('agregar_curso'))

                if id_sede_int not in sedes_ids:
                    # está intentando grabar en una sede que no le corresponde
                    return redirect(url_for('agregar_curso'))

                id_sede_final = id_sede_int
            else:
                # usuario con solo una sede → forzamos la sede de sesión
                id_sede_final = id_sede_sess

        # si por alguna razón no tenemos sede válida, salimos
        if not id_sede_final:
            return redirect(url_for('agregar_curso'))

        # Calcular duración
        from datetime import datetime
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d")
        duracion_curso = (ff - fi).days

        try:
            con = get_db_connection()
            cur = con.cursor()

            # 1️⃣ Insertar curso
            cur.execute("""
                INSERT INTO cursos (
                    nombre, descripcion, id_docente, id_sede, aula, id_horario, estado,
                    fecha_inicio, fecha_fin, duracion_curso
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                nombre,
                descripcion,
                int(id_docente) if id_docente and id_docente != "0" else None,
                int(id_sede_final),  # 👈 aquí usamos la sede ya validada
                aula,
                int(id_horario) if id_horario else None,
                estado,
                fecha_inicio,
                fecha_fin,
                duracion_curso
            ))

            # 🔹 ID del curso recién creado
            id_curso_nuevo = cur.lastrowid

            # 2️⃣ RELLENAR SOLO id_curso EN asignaciones_docentes (si corresponde)
            if id_docente and id_docente != "0" and id_horario:
                cur.execute("""
                    UPDATE asignaciones_docentes
                    SET id_curso = ?
                    WHERE id_docente = ?
                      AND id_horario = ?
                      AND (id_curso IS NULL OR id_curso = 0)
                """, (
                    int(id_curso_nuevo),
                    int(id_docente),
                    int(id_horario)
                ))

            con.commit()
            con.close()

            return redirect(url_for('administrar_cursos'))

        except Exception as e:
            app.logger.exception("Error agregando curso: %s", e)
            try:
                con.close()
            except:
                pass
            return redirect(url_for('agregar_curso'))

    # ---------------- GET: CARGAR FORMULARIO ----------------
    try:
        con = get_db_connection()
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        rol          = session.get("rol")
        sedes_ids    = session.get("sedes", [])
        id_sede_sess = session.get("id_sede")

        # Docentes (todos)
        docentes = cur.execute(
            "SELECT id_docente, nombres FROM docentes ORDER BY nombres"
        ).fetchall()

        # Sedes según rol
        if rol == "admin":
            # todas las sedes
            sedes = cur.execute("""
                SELECT id_sede, nombre_sede
                FROM sedes
                ORDER BY nombre_sede
            """).fetchall()
        else:
            if sedes_ids and len(sedes_ids) > 1:
                # varias sedes → solo las suyas
                placeholders = ",".join("?" * len(sedes_ids))
                sedes = cur.execute(f"""
                    SELECT id_sede, nombre_sede
                    FROM sedes
                    WHERE id_sede IN ({placeholders})
                    ORDER BY nombre_sede
                """, sedes_ids).fetchall()
            else:
                # una sola sede → solo esa
                sedes = cur.execute("""
                    SELECT id_sede, nombre_sede
                    FROM sedes
                    WHERE id_sede = ?
                """, (id_sede_sess,)).fetchall()

        # Horarios
        horarios = cur.execute("""
            SELECT id_horario, nombre_horario, hora_inicio, hora_fin, dias
            FROM horarios
            ORDER BY nombre_horario ASC
        """).fetchall()

        con.close()
    except Exception as e:
        app.logger.exception("Error cargando datos para agregar_curso: %s", e)
        docentes, sedes, horarios = [], [], []

    return render_template(
        'agregar_curso.html',
        docentes=docentes,
        sedes=sedes,
        horarios=horarios,
        id_sede_usuario=session.get("id_sede")
    )

#=================================
# ADMINISTRAR CURSOS
@app.route('/administrar_cursos')
def administrar_cursos():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    id_sede_usuario = session.get("id_sede")         # sede principal (por compatibilidad)
    sedes_usuario = session.get("sedes") or []       # lista de sedes asignadas (multi-sede)

    try:
        con = get_db_connection()
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # ----------- CONSULTA BASE -----------
        query = """
            SELECT 
                c.id_curso, 
                c.nombre, 
                c.descripcion, 
                d.nombres AS docente, 
                s.nombre_sede AS sede, 
                c.aula,
                h.descripcion AS horario,
                c.estado,
                c.fecha_inicio,
                c.fecha_fin,
                c.duracion_curso
            FROM cursos c
            LEFT JOIN docentes d ON c.id_docente = d.id_docente
            LEFT JOIN sedes s    ON c.id_sede   = s.id_sede
            LEFT JOIN horarios h ON c.id_horario = h.id_horario
            WHERE 1=1
        """

        params = []

        # ----------- FILTRO POR SEDE -----------
        if rol != "admin":
            if sedes_usuario:   # tiene varias sedes asignadas
                placeholders = ",".join(["?"] * len(sedes_usuario))
                query += f" AND IFNULL(c.id_sede, 0) IN ({placeholders})"
                params.extend(sedes_usuario)
            else:
                # fallback: solo su sede principal
                query += " AND IFNULL(c.id_sede, 0) = ?"
                params.append(id_sede_usuario)

        query += " ORDER BY s.nombre_sede, c.nombre"

        cur.execute(query, params)
        cursos = cur.fetchall()

        con.close()

    except Exception as e:
        app.logger.exception("Error cargando cursos: %s", e)
        cursos = []

    return render_template('administrar_cursos.html', cursos=cursos)

#=================================
# EDITAR CURSO
@app.route('/editar_curso/<int:id_curso>', methods=['GET', 'POST'])
def editar_curso(id_curso):
    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    rol             = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids       = session.get("sedes", []) or []

    # ===================== POST: GUARDAR =====================
    if request.method == 'POST':
        nombre      = (request.form.get('nombre') or '').strip()
        descripcion = (request.form.get('descripcion') or '').strip()
        id_docente  = request.form.get('id_docente') or None
        id_sede     = request.form.get('id_sede') or None
        aula        = (request.form.get('aula') or '').strip()
        id_horario  = request.form.get('id_horario') or None
        estado      = (request.form.get('estado') or 'Activo').strip()

        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin    = request.form.get('fecha_fin')

        # 🛡️ Normalizar / validar sede según rol
        if rol != "admin":
            # Si el usuario tiene varias sedes en sesión (lista)
            if sedes_ids and len(sedes_ids) > 1:
                try:
                    id_sede_int = int(id_sede) if id_sede else None
                except ValueError:
                    con.close()
                    return redirect(url_for('administrar_cursos'))

                if id_sede_int not in sedes_ids:
                    # Intento de usar una sede que no le pertenece
                    con.close()
                    return redirect(url_for('administrar_cursos'))

                id_sede = id_sede_int
            else:
                # Solo una sede → forzamos a esa sede, ignorando lo que venga del form
                id_sede = id_sede_usuario

        # 1️⃣ Actualizar curso
        cur.execute("""
            UPDATE cursos
            SET nombre       = ?,
                descripcion  = ?,
                id_docente   = ?,
                id_sede      = ?,
                aula         = ?,
                id_horario   = ?,
                fecha_inicio = ?,
                fecha_fin    = ?,
                estado       = ?
            WHERE id_curso   = ?
        """, (
            nombre,
            descripcion,
            int(id_docente) if id_docente else None,
            int(id_sede)    if id_sede else None,
            aula,
            int(id_horario) if id_horario else None,
            fecha_inicio,
            fecha_fin,
            estado,
            id_curso
        ))

        # 2️⃣ Rellenar id_curso en asignaciones_docentes
        if id_docente and id_horario:
            cur.execute("""
                UPDATE asignaciones_docentes
                SET id_curso = ?
                WHERE id_docente = ? AND id_horario = ?
            """, (
                int(id_curso),
                int(id_docente),
                int(id_horario)
            ))

        con.commit()
        con.close()
        return redirect(url_for('administrar_cursos'))

    # ===================== GET: CARGAR FORM =====================
    curso = cur.execute(
        "SELECT * FROM cursos WHERE id_curso = ?",
        (id_curso,)
    ).fetchone()

    if not curso:
        con.close()
        return redirect(url_for('administrar_cursos'))

    # Docentes (los dejamos globales)
    docentes = cur.execute(
        "SELECT id_docente, nombres FROM docentes ORDER BY nombres"
    ).fetchall()

    # Sedes filtradas por rol
    if rol == "admin":
        sedes = cur.execute(
            "SELECT id_sede, nombre_sede FROM sedes ORDER BY nombre_sede"
        ).fetchall()
    else:
        if sedes_ids and len(sedes_ids) > 1:
            # Varias sedes asignadas al usuario → solo esas
            placeholders = ",".join("?" * len(sedes_ids))
            sedes = cur.execute(f"""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre_sede
            """, sedes_ids).fetchall()
        else:
            # Solo una sede
            sedes = cur.execute("""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede = ?
                ORDER BY nombre_sede
            """, (id_sede_usuario,)).fetchall()

    # Horarios (globales)
    horarios = cur.execute("""
        SELECT id_horario, nombre_horario, hora_inicio, hora_fin 
        FROM horarios
        ORDER BY nombre_horario
    """).fetchall()

    con.close()
    return render_template(
        'editar_curso.html',
        curso=curso,
        docentes=docentes,
        sedes=sedes,
        horarios=horarios
    )

@app.route('/eliminar_curso/<int:id_curso>', methods=['POST', 'GET'])
def eliminar_curso(id_curso):
    try:
        con = get_db_connection()
        cur = con.cursor()

        # 0️⃣ Limpiar id_sede de ALUMNOS vinculados a este curso
        cur.execute("""
            UPDATE alumnos
            SET id_sede = NULL
            WHERE id_alumno IN (
                SELECT DISTINCT id_alumno
                FROM matriculas
                WHERE id_curso = ?
            )
        """, (id_curso,))

        # 1️⃣ Desvincular a los alumnos en MATRÍCULAS (sin curso y sin sede)
        cur.execute("""
            UPDATE matriculas
            SET id_curso = NULL,
                id_sede  = NULL
            WHERE id_curso = ?
        """, (id_curso,))

        # 2️⃣ Desvincular asignaciones de docentes a este curso (si usas esa tabla)
        cur.execute("""
            UPDATE asignaciones_docentes
            SET id_curso = NULL
            WHERE id_curso = ?
        """, (id_curso,))

        # 3️⃣ Ahora sí, eliminar el curso
        cur.execute("DELETE FROM cursos WHERE id_curso = ?", (id_curso,))

        con.commit()
        con.close()

        return redirect(url_for('administrar_cursos', ok=1))

    except Exception as e:
        app.logger.exception("Error al eliminar curso: %s", e)
        try:
            con.close()
        except:
            pass
        return redirect(url_for('administrar_cursos', error=1))


#=================================
# RUTAS PARA HORARIOS
HHMM = re.compile(r"^\d{2}:\d{2}$")

def _valid_hhmm(s: str) -> bool:
    return bool(s and HHMM.match(s))

def _join_dias(dias_list):
    # acepta lista de checkboxes name="dias"  ->  'LU-MA-MI'
    if not dias_list:
        return ""
    if isinstance(dias_list, str):
        return dias_list.strip()
    return "-".join([d for d in dias_list if d]).strip("-")

# =========================
# LISTAR
# =========================
@app.route('/horarios')
def horarios():
    try:
        con = get_db_connection()
        filas = con.execute("""
            SELECT
                id_horario,
                COALESCE(nombre_horario,'') AS nombre_horario,
                COALESCE(descripcion,'') AS descripcion,
                COALESCE(dias,'') AS dias,
                COALESCE(hora_inicio,'') AS hora_inicio,
                COALESCE(hora_fin,'') AS hora_fin,
                COALESCE(tolerancia_inicio, COALESCE(tolerancia,0)) AS tolerancia_inicio,
                COALESCE(tolerancia_fin, 0) AS tolerancia_fin,
                COALESCE(entrada_temprana, 0) AS entrada_temprana,   -- ✅ AQUI IMPORTA entrada temprana
                COALESCE(salida_temprana, 0) AS salida_temprana,     -- ✅ AQUI IMPORTA salida temprana
                COALESCE(activo,1) AS activo,
                COALESCE(nota,'') AS nota
            FROM horarios
            ORDER BY id_horario ASC
        """).fetchall()
        con.close()
    except Exception as e:
        app.logger.exception("Error cargando horarios: %s", e)
        filas = []

    return render_template('horarios.html', horarios=filas)

# =========================
# CREAR (nuevo esquema)
# =========================
@app.route('/añadir_horario', methods=['GET', 'POST'])
def añadir_horario():
    con = get_db_connection()
    cur = con.cursor()

    if request.method == 'POST':
        dias_sel = request.form.getlist('dias')
        dias = "-".join(dias_sel) if dias_sel else ""
        turno = (request.form.get('turno') or "").strip()
        hora_inicio = (request.form.get('hora_inicio') or "").strip()
        hora_fin    = (request.form.get('hora_fin') or "").strip()

        try:
            tol_inicio = int(request.form.get('tolerancia_inicio') or 0)
            tol_fin    = int(request.form.get('tolerancia_fin') or 0)
            tolerancia = int(request.form.get('tolerancia') or 0)
            entrada_temprana = int(request.form.get('entrada_temprana') or 0)  # ✅ NUEVO
            salida_temprana = int(request.form.get('salida_temprana') or 0)
        except ValueError:
            tol_inicio = tol_fin = tolerancia = entrada_temprana = salida_temprana = 0

        nombre_horario = (request.form.get('nombre_horario') or "").strip()
        nota   = (request.form.get('nota') or "").strip()

        descripcion_form = (request.form.get('descripcion') or "").strip()
        if not descripcion_form:
            base_turno = f" {turno}" if turno else ""
            if hora_inicio and hora_fin:
                descripcion = f"{dias}{base_turno} ({hora_inicio} - {hora_fin})".strip()
            else:
                descripcion = f"{dias}{base_turno}".strip()
        else:
            descripcion = descripcion_form

        if not dias or not hora_inicio or not hora_fin:
            con.close()
            return render_template('añadir_horario.html',
                                   mensaje="⚠️ Selecciona días y horas de inicio/fin.")

        existe = cur.execute("SELECT COUNT(*) AS cnt FROM horarios WHERE descripcion = ?",
                             (descripcion,)).fetchone()["cnt"]
        if existe > 0:
            con.close()
            return render_template('añadir_horario.html',
                                   mensaje=f"⚠️ El horario '{descripcion}' ya existe.")

                # ✅ INSERT ACTUALIZADO
        cur.execute("""
            INSERT INTO horarios
                (descripcion, dias, hora_inicio, hora_fin,
                tolerancia, tolerancia_inicio, tolerancia_fin,
                entrada_temprana, salida_temprana, activo, nombre_horario, nota)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (descripcion, dias, hora_inicio, hora_fin,
            tolerancia, tol_inicio, tol_fin,
            entrada_temprana, salida_temprana, 1, nombre_horario, nota))


        con.commit()
        con.close()
        return render_template('añadir_horario.html',
                               mensaje=f"✅ Horario '{descripcion}' añadido correctamente.")

    horarios = cur.execute("""
        SELECT id_horario, descripcion, dias, hora_inicio, hora_fin, tolerancia
        FROM horarios
        ORDER BY id_horario ASC
    """).fetchall()
    con.close()
    return render_template('añadir_horario.html', horarios=horarios)

#=========================================
# EDITAR
@app.route('/editar_horario/<int:id_horario>', methods=['GET', 'POST'])
def editar_horario(id_horario):
    con = get_db_connection()
    cur = con.cursor()

    # Cargar el horario existente
    horario = cur.execute("""
        SELECT *
        FROM horarios
        WHERE id_horario = ?
    """, (id_horario,)).fetchone()

    if not horario:
        con.close()
        return "⚠️ Error: Horario no encontrado."

    if request.method == 'POST':
        dias_sel = request.form.getlist('dias')
        dias = "-".join(dias_sel) if dias_sel else ""
       
        hora_inicio = (request.form.get('hora_inicio') or "").strip()
        hora_fin = (request.form.get('hora_fin') or "").strip()
        
        try:
            tol_inicio = int(request.form.get('tolerancia_inicio') or 0)
            tol_fin = int(request.form.get('tolerancia_fin') or 0)
            tolerancia = int(request.form.get('tolerancia') or 0)
            entrada_temprana = int(request.form.get('entrada_temprana') or 0)  # ✅ NUEVO
            salida_temprana = int(request.form.get('salida_temprana') or 0)
        except ValueError:
            tol_inicio = tol_fin = tolerancia = entrada_temprana = salida_temprana = 0

        nombre_horario = (request.form.get('nombre_horario') or "").strip()
        nota = (request.form.get('nota') or "").strip()

        # Generar descripción
        if hora_inicio and hora_fin:
            descripcion = f"{dias} {nombre_horario} ({hora_inicio} - {hora_fin})".strip()
        else:
            descripcion = f"{dias} {nombre_horario}".strip()

        # Validar campos esenciales
        if not dias or not hora_inicio or not hora_fin:
            con.close()
            return render_template('editar_horario.html', horario=horario,
                                   mensaje="⚠️ Seleccione días y horas válidos.")

        # Evitar duplicado (pero permitir el mismo registro actual)
        existe = cur.execute("""
            SELECT COUNT(*) AS cnt
            FROM horarios
            WHERE descripcion = ? AND id_horario <> ?
        """, (descripcion, id_horario)).fetchone()["cnt"]

        if existe > 0:
            con.close()
            return render_template('editar_horario.html', horario=horario,
                                   mensaje=f"⚠️ Ya existe otro horario igual: {descripcion}")

        # Actualizar
        cur.execute("""
            UPDATE horarios
                SET descripcion=?, dias=?, hora_inicio=?, hora_fin=?,
                tolerancia=?, tolerancia_inicio=?, tolerancia_fin=?,
                entrada_temprana=?, nombre_horario=?, nota=?,
                salida_temprana=?
            
            WHERE id_horario=?
        """, (descripcion, dias, hora_inicio, hora_fin,
              tolerancia, tol_inicio, tol_fin, entrada_temprana, nombre_horario, nota, salida_temprana, id_horario))

        con.commit()
        con.close()
        return redirect(url_for('horarios'))

    # GET → mostrar formulario
    con.close()
    return render_template('editar_horario.html', horario=horario)

# =========================
# SOFT DELETE (inactivar)
# =========================
@app.route('/eliminar_horario/<int:id_horario>')
def eliminar_horario(id_horario):
    try:
        con = get_db_connection()
        con.execute("PRAGMA busy_timeout = 5000")

        # 1) Liberar cursos que usen este horario:
        con.execute("""
            UPDATE cursos SET id_horario = NULL
            WHERE id_horario = ?
        """, (id_horario,))

        # 2) Borrar el horario definitivamente:
        con.execute("""
            DELETE FROM horarios
            WHERE id_horario = ?
        """, (id_horario,))

        con.commit()
        con.close()

    except Exception as e:
        print("Error al eliminar horario:", e)
        try: con.close()
        except: pass

    return redirect(url_for('horarios'))

#===================================================
# REDIRECCIÓN CORTA A ASISTENCIA

@app.route('/r/<codigo>')
def redirigir_a_asistencia(codigo):
    destino = url_for('registrar_asistencia', codigo=codigo, _external=True)
    return redirect(destino)

# ---------------------------
# RUTA: Asistencias de alumnos
# ---------------------------
@app.route('/asistencias')
def asistencias():
    # 🔐 Solo usuarios logueados
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    sedes_ids = session.get("sedes") or []
    id_sede_usuario = session.get("id_sede")

    # Fallback: si solo tienes id_sede
    if not sedes_ids and id_sede_usuario:
        sedes_ids = [id_sede_usuario]

    fecha_desde   = request.args.get('fecha_desde')
    fecha_hasta   = request.args.get('fecha_hasta')
    buscar_alumno = (request.args.get('buscar_alumno') or '').strip()
    buscar_curso  = (request.args.get('buscar_curso')  or '').strip()

    # Si no envían fechas → usar HOY
    if not fecha_desde and not fecha_hasta:
        hoy = datetime.now().strftime("%Y-%m-%d")
        fecha_desde = hoy
        fecha_hasta = hoy

    # -------------------- SQL BASE --------------------
    query = """
        SELECT 
            a.codigo, 
            a.dni, 
            a.nombres,
            c.nombre AS curso,
            COALESCE(
                h.nombre_horario || ' (' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                c.nombre_horario
            ) AS horario_texto,
            s.nombre_sede AS sede,   -- 👈 SEDE
            m.estado AS matricula_estado,
            m.aula AS aula_matricula,
            asi.fecha,
            asi.hora,
            asi.observacion
        FROM asistencias asi
        JOIN matriculas m      ON asi.id_matricula = m.id_matricula
        JOIN alumnos   a       ON m.id_alumno = a.id_alumno
        LEFT JOIN cursos   c   ON m.id_curso = c.id_curso
        LEFT JOIN horarios h   ON c.id_horario = h.id_horario
        LEFT JOIN sedes    s   ON s.id_sede = c.id_sede   -- 👈 importante: sede del curso
        WHERE 1=1
    """

    params = []

    # 🔎 Filtro por fechas
    if fecha_desde and fecha_hasta:
        query += " AND asi.fecha BETWEEN ? AND ?"
        params.extend([fecha_desde, fecha_hasta])

    # 🔎 Filtro por alumno
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.codigo LIKE ? OR a.dni LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # 🔎 Filtro por curso
    if buscar_curso:
        query += " AND c.nombre = ?"
        params.append(buscar_curso)

    # 📍 Filtro por sede según usuario
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario sin sedes asignadas → no ve nada
            query += " AND 1=0"

    query += " ORDER BY asi.fecha, a.nombres, asi.hora"

    # -------------------- EJECUTAR --------------------
    with get_db_connection() as con:
        cur = con.cursor()
        filas = cur.execute(query, params).fetchall()
        cursos = cur.execute("SELECT nombre FROM cursos ORDER BY nombre").fetchall()

    # Texto para el título
    if fecha_desde == fecha_hasta:
        fecha_mostrar = fecha_desde
    else:
        fecha_mostrar = f"{fecha_desde} → {fecha_hasta}"

    print(">>> BD QUE LEE FLASK:", os.path.abspath("base_datos.db"))

    return render_template(
        "asistencias.html",
        asistencias=filas,
        fecha=fecha_mostrar,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        buscar_alumno=buscar_alumno,
        buscar_curso=buscar_curso,
        cursos=cursos,
    )

#===================================================================
# LIMPIEZA DE FECHAS EN ASISTENCIAS

def limpiar_fechas():
    con = sqlite3.connect("base_datos.db")
    cur = con.cursor()

    print("🔧 Corrigiendo fechas en asistencias...")

    # 1. TRIM general
    cur.execute("""
        UPDATE asistencias
        SET fecha = TRIM(fecha)
        WHERE fecha IS NOT NULL
    """)

    # 2. Valores basura
    cur.execute("""
        UPDATE asistencias
        SET fecha = NULL
        WHERE fecha IS NULL
           OR fecha = ''
           OR fecha = ' '
           OR fecha LIKE '%--%'
           OR fecha LIKE '--%'
           OR fecha LIKE '%--'
           OR LOWER(fecha) = 'null'
    """)

    # 3. Convertir DD/MM/YYYY → YYYY-MM-DD
    cur.execute("""
        UPDATE asistencias
        SET fecha = substr(fecha,7,4) || '-' || substr(fecha,4,2) || '-' || substr(fecha,1,2)
        WHERE fecha LIKE '__/__/____';
    """)

    # ⚠ ADVERTENCIA: eliminar paso 4 porque daña las fechas

    # 5. TRIM final
    cur.execute("""
        UPDATE asistencias
        SET fecha = TRIM(fecha)
        WHERE fecha IS NOT NULL;
    """)
    
    con.commit()
    con.close()

    print("✅ Fechas corregidas correctamente.")


# --- Auxiliares -------------------------------------------------------------

def dentro_de_horario(hora_actual, hora_inicio, hora_fin):
    """Devuelve True si h_actual está entre inicio y fin."""
    fmt = "%H:%M:%S" if len(hora_actual) > 5 else "%H:%M"
    h_actual = datetime.strptime(hora_actual, fmt).time()
    
    fmt2 = "%H:%M:%S" if len(hora_inicio) > 5 else "%H:%M"
    h_inicio = datetime.strptime(hora_inicio, fmt2).time()

    fmt3 = "%H:%M:%S" if len(hora_fin) > 5 else "%H:%M"
    h_fin = datetime.strptime(hora_fin, fmt3).time()

    return h_inicio <= h_actual <= h_fin

#===========================================================================
# EXTRACTORES Y EVALUADORES DE HORARIOS
_TIME_IN = ("%H:%M:%S", "%H:%M")  # acepta HH:MM:SS o HH:MM


def _parse_hora(hora_str: str) -> datetime:
    """Devuelve datetime hoy con la hora dada. Acepta HH:MM o HH:MM:SS."""
    for fmt in _TIME_IN:
        try:
            t = datetime.strptime(hora_str.strip(), fmt)
            # normalizamos a hoy con esa hora
            now = datetime.now()
            return now.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0), microsecond=0)
        except ValueError:
            continue
    raise ValueError(f"Formato de hora inválido: {hora_str!r}")

def _norm_hhmm(hora_str: str) -> str:
    """Normaliza a HH:MM (sin segundos)."""
    dt = _parse_hora(hora_str)
    return dt.strftime("%H:%M")

# ---------------------------------
# Extractores (uno solo y flexible)
# ---------------------------------

def extraer_horario(texto: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """
    Extrae 'HH:MM - HH:MM' (con o sin paréntesis, con -, – o — y opcional :SS).
    Devuelve (inicio, fin) normalizados a HH:MM. Si no encuentra, (None, None).
    """
    if not texto:
        return None, None
    # unificamos guiones
    txt = texto.replace("–", "-").replace("—", "-")
    # tolerante con espacios y paréntesis
    m = re.search(r'\(?\s*(\d{1,2}:\d{2}(?::\d{2})?)\s*-\s*(\d{1,2}:\d{2}(?::\d{2})?)\s*\)?', txt)
    if not m:
        return None, None
    ini, fin = m.group(1), m.group(2)
    try:
        return _norm_hhmm(ini), _norm_hhmm(fin)
    except ValueError:
        return None, None

# Alias si quieres mantener nombres antiguos
extraer_hora_inicio_fin = extraer_horario

# ---------------------------------
# Evaluador de asistencia (único)
# ---------------------------------

def evaluar_entrada(hora_actual_str, hora_inicio_str, tolerancia_min=0, entrada_temprana_min=0):
    if not hora_inicio_str:
        return "Entrada ✅"

    fmt_actual  = "%H:%M:%S" if len(hora_actual_str.strip())  > 5 else "%H:%M"
    fmt_inicio  = "%H:%M:%S" if len(hora_inicio_str.strip())  > 5 else "%H:%M"

    h_actual = datetime.strptime(hora_actual_str, fmt_actual)
    h_inicio = datetime.strptime(hora_inicio_str, fmt_inicio)

    entrada_temprana_min = int(entrada_temprana_min or 0)
    tolerancia_min       = int(tolerancia_min or 0)

    diff_min = (h_actual - h_inicio).total_seconds() / 60

    # Dentro de la ventana de entrada temprana (antes del inicio pero no más que el rango)
    if diff_min < 0 and abs(diff_min) <= entrada_temprana_min:
        return "Entrada 🌅 Temprana"

    # Puntual (≤ a la hora de inicio)
    if diff_min <= 0:
        return "Entrada ✅ Puntual"

    # Dentro de tolerancia positiva
    if diff_min <= tolerancia_min:
        return "Entrada ⏱️ Justificada"

    # Tardanza
    return "Entrada ❌ Tardanza grave"

#===========================================================
# RUTA PRINCIPAL DE REGISTRO DE ASISTENCIA 
@app.route('/registrar_asistencia')
def registrar_asistencia():
    codigo = request.args.get("codigo")
    if not codigo:
        return "Código no recibido", 400

    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    hora_actual = datetime.now().strftime("%H:%M:%S")

    # ------------------  CONEXIÓN CONTROLADA  ------------------
    with get_db_connection() as con:
        cur = con.cursor()

        # TODO: aquí va tu código completo sin NINGÚN con.close()
    # Alumno
    alumno = cur.execute("""
        SELECT a.id_alumno, a.codigo, a.nombres, a.fecha_nacimiento,
               a.institucion, s.nombre_sede AS sede
        FROM alumnos a
        LEFT JOIN sedes s ON a.id_sede = s.id_sede
        WHERE a.codigo = ?
    """, (codigo,)).fetchone()

    if not alumno:
  
                return redirect(url_for('registrar_asistencia_docente', codigo=codigo))

    # matrícula + horario
    matricula = cur.execute("""
        SELECT 
            m.id_matricula,
            m.deuda,
            c.nombre AS curso,
            COALESCE(s.nombre_sede, '') AS sede,
            COALESCE(h.nombre_horario, '') AS nombre_horario,
            COALESCE(h.dias, '') AS dias,
            COALESCE(h.hora_inicio, '') AS h_inicio,
            COALESCE(h.hora_fin, '') AS h_fin,
            COALESCE(h.tolerancia_inicio, 0) AS tolerancia,
            COALESCE(h.entrada_temprana, 0) AS entrada_temprana,
            COALESCE(h.salida_temprana, 0) AS salida_temprana
        FROM matriculas m
        JOIN cursos c ON m.id_curso = c.id_curso
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        LEFT JOIN horarios h ON c.id_horario = h.id_horario
        WHERE m.id_alumno = ?
        ORDER BY m.id_matricula DESC
        LIMIT 1
    """, (alumno["id_alumno"],)).fetchone()

    if not matricula:
      
        return f"⚠️ El alumno {alumno['nombres']} no tiene matrícula activa."

    hora_inicio = (matricula["h_inicio"] or "").strip()
    hora_fin = (matricula["h_fin"] or "").strip()
    tolerancia = matricula["tolerancia"]

    # Cumpleaños (antes de usar cumple_hoy)
    cumple_hoy = False
    if alumno["fecha_nacimiento"]:
        try:
            cumple_hoy = datetime.strptime(alumno["fecha_nacimiento"], "%Y-%m-%d").strftime("%m-%d") == datetime.now().strftime("%m-%d")
        except:
            pass

    # Si no hay horario → no controlar entrada/salida
    if hora_inicio and hora_fin:
        try:
            if not dentro_de_horario(hora_actual, hora_inicio, hora_fin):
              
                datos = {
                    "codigo": alumno["codigo"],
                    "nombres": alumno["nombres"],
                    "curso": matricula["curso"],
                    "turno": f"{hora_inicio} - {hora_fin}",
                    "fecha": fecha_hoy,
                    "hora": hora_actual,
                    "estado_asistencia": "⛔ Fuera del horario permitido",
                    "cumple_hoy": cumple_hoy,
                    "institucion": alumno["institucion"] or "No registrada",
                    "sede": matricula["sede"] or "No asignada",
                }
                return render_template("registro_fuera_horario.html", datos=datos)
        except:
            pass

    # Bloqueo por deuda
    if matricula["deuda"] and float(matricula["deuda"]) > 0:
        datos = {
            "codigo": alumno["codigo"],
            "nombres": alumno["nombres"],
            "curso": matricula["curso"],
            "turno": f"{hora_inicio or '—'} - {hora_fin or '—'}",
            "fecha": fecha_hoy,
            "hora": hora_actual,
            "estado_asistencia": "⛔ Bloqueado por deuda",
            "institucion": alumno["institucion"] or "No registrada",
            "sede": matricula["sede"] or "No asignada",
            "deuda": matricula["deuda"],
        }
        
        return render_template("registro_bloqueado.html", datos=datos)


    def _fmt(hhmmss: str) -> str:
        return "%H:%M:%S" if len((hhmmss or "").strip()) > 5 else "%H:%M"

    def to_dt_today(hhmmss: str) -> datetime:
        fmt = _fmt(hhmmss)
        t = datetime.strptime(hhmmss, fmt)
        today = datetime.now()
        return today.replace(hour=t.hour, minute=t.minute, second=getattr(t, "second", 0), microsecond=0)
    
    registros = cur.execute("""
        SELECT id_asistencia, fecha, hora, observacion
        FROM asistencias
        WHERE id_matricula = ? AND fecha = ?
        ORDER BY hora ASC
    """, (matricula["id_matricula"], fecha_hoy)).fetchall()

    # Flags de estado
    ya_tiene_entrada = any((r["observacion"] or "").startswith("Entrada") for r in registros)
    ya_tiene_salida  = any("Salida" in (r["observacion"] or "") for r in registros)

    # Normalizar horas a datetime de hoy
    h_actual_full  = to_dt_today(hora_actual)
    h_inicio_full  = to_dt_today(hora_inicio)
    h_fin_full     = to_dt_today(hora_fin)
    entrada_temprana = int(matricula["entrada_temprana"] or 0)
    salida_temprana = int (matricula ["salida_temprana"] or 0)
    
    # Límite mínimo permitido (inicio - entrada_temprana)
    h_inicio_temprano = h_inicio_full - timedelta(minutes=entrada_temprana)

    # ------------- DECISION TREE -------------
    # 1) PRIMERA MARCA DEL DÍA
    if not registros:
        # ⛔ Cualquier primera marca FUERA de la ventana permitida
        #    (antes de h_inicio_temprano o después de h_fin_full)
        if h_actual_full < h_inicio_temprano or h_actual_full > h_fin_full:
            return render_template("fuera_horario.html", datos={
                "codigo": alumno["codigo"],
                "nombres": alumno["nombres"],
                "curso": matricula["curso"],
                "turno": f"{hora_inicio} - {hora_fin}",
                "fecha": fecha_hoy,
                "hora": hora_actual,
                "estado_asistencia": "⛔ Fuera del horario permitido",
                "cumple_hoy": cumple_hoy,
                "institucion": alumno["institucion"] or "No registrada",
                "sede": matricula["sede"] or "No asignada",
            })

        # ✅ Primera marca DENTRO de la ventana → ENTRADA
        observacion = evaluar_entrada(
            hora_actual,
            hora_inicio,
            tolerancia_min=tolerancia,
            entrada_temprana_min=entrada_temprana
    )
        
        
    # 2) YA HAY ENTRADA y NO HAY SALIDA: no permitir más marcas hasta hora_fin; solo a partir de hora_fin se registra salida
    elif ya_tiene_entrada and not ya_tiene_salida:

        # Permitir salida temprana
        h_salida_permitida = h_fin_full - timedelta(minutes=salida_temprana)

        if h_actual_full >= h_salida_permitida:
            observacion = "Salida ✅ (Temprana)" if h_actual_full < h_fin_full else "Salida ✅"
        else:
            return render_template("registro_duplicado.html", datos={
                "codigo": alumno["codigo"],
                "nombres": alumno["nombres"],
                "curso": matricula["curso"],
                "turno": f"{hora_inicio} - {hora_fin}",
                "fecha": fecha_hoy,
                "hora": hora_actual,
                "estado_asistencia": "⛔ Aún no puedes registrar salida",
                "cumple_hoy": cumple_hoy,
                "institucion": alumno["institucion"] or "No registrada",
                "sede": matricula["sede"] or "No asignada",
            })

    # 3) YA HAY ENTRADA y SALIDA: no permitir más marcas
    elif ya_tiene_entrada and ya_tiene_salida:
        return render_template("registro_salida_duplicado.html", datos={
            "codigo": alumno["codigo"],
            "nombres": alumno["nombres"],
            "curso": matricula["curso"],
            "turno": f"{hora_inicio} - {hora_fin}",
            "fecha": fecha_hoy,
            "hora": hora_actual,
            "estado_asistencia": "Salida ya registrada",
            "institucion": alumno["institucion"] or "No registrada",
            "sede": matricula["sede"] or "No asignada",
        })

    # 4) Salvaguarda (no debería llegar aquí, pero por si acaso)
    else:
        observacion = evaluar_entrada(
            hora_actual,
            hora_inicio,
            tolerancia_min=tolerancia,
            entrada_temprana_min=entrada_temprana
        )

    # --- Insert final (solo si llegamos aquí con 'observacion') ---
    cur.execute("""
        INSERT INTO asistencias (id_matricula, fecha, hora, observacion)
        VALUES (?, ?, ?, ?)
    """, (matricula["id_matricula"], fecha_hoy, hora_actual, observacion))
    con.commit()

    return render_template("registro_exitoso.html", datos={
        "codigo": alumno["codigo"],
        "nombres": alumno["nombres"],
        "curso": matricula["curso"],
        "turno": f"{hora_inicio} - {hora_fin}",
        "fecha": fecha_hoy,
        "hora": hora_actual,
        "estado_asistencia": observacion,
        "cumple_hoy": cumple_hoy,
        "institucion": alumno["institucion"] or "No registrada",
        "sede": matricula["sede"] or "No asignada",
    })

#===========================================
# EXPORTAR ASISTENCIAS A EXCEL
@app.route('/exportar_asistencias')
def exportar_asistencias():
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import pandas as pd
    from datetime import datetime, date

    # 🔐 Solo usuario logueado
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    sedes_ids = session.get("sedes") or []
    id_sede_usuario = session.get("id_sede")

    # Fallback: si solo tienes id_sede
    if not sedes_ids and id_sede_usuario:
        sedes_ids = [id_sede_usuario]

    # 🔹 Capturar filtros desde la URL
    fecha_filtro = request.args.get('fecha')
    fecha_desde  = request.args.get('fecha_desde')
    fecha_hasta  = request.args.get('fecha_hasta')
    mes_filtro   = request.args.get('mes')

    con = get_db_connection()
    cur = con.cursor()

    # ==============================
    # 1) Armar el SELECT base
    # ==============================
    query = """
        SELECT 
            a.id_alumno,
            a.codigo,
            a.dni,
            a.nombres,
            a.correo,
            a.celular,
            a.institucion,
            c.nombre AS curso,
            m.aula AS aula_matricula,
            COALESCE(h.descripcion, c.id_horario) AS turno,
            s.nombre_sede AS sede,
            asi.fecha,
            asi.hora,
            asi.observacion AS tipo_registro,
            m.estado AS estado_pago
        FROM asistencias asi
        JOIN matriculas m ON asi.id_matricula = m.id_matricula
        JOIN alumnos   a  ON m.id_alumno = a.id_alumno
        LEFT JOIN cursos   c ON m.id_curso = c.id_curso
        LEFT JOIN horarios h ON c.id_horario = h.id_horario
        LEFT JOIN sedes    s ON c.id_sede = s.id_sede
        WHERE 1=1
    """

    params = []

    # ==============================
    # 2) Filtros de FECHA
    # ==============================
    if fecha_desde and fecha_hasta:
        # 🗓 Rango personalizado
        query += " AND asi.fecha BETWEEN ? AND ?"
        params.extend([fecha_desde, fecha_hasta])
        fecha_texto    = f"Del {fecha_desde} al {fecha_hasta}"
        nombre_archivo = f"Asistencias_{fecha_desde}_a_{fecha_hasta}.xlsx"

    elif mes_filtro == 'actual':
        hoy = date.today()
        primer_dia_mes = hoy.replace(day=1)
        fecha_inicio   = primer_dia_mes.strftime("%Y-%m-%d")
        fecha_fin      = hoy.strftime("%Y-%m-%d")

        query += " AND asi.fecha BETWEEN ? AND ?"
        params.extend([fecha_inicio, fecha_fin])

        fecha_texto    = f"Del {fecha_inicio} al {fecha_fin}"
        nombre_archivo = f"Asistencias_{fecha_inicio}_a_{fecha_fin}.xlsx"

    elif fecha_filtro:
        # 📅 Fecha única
        query += " AND asi.fecha = ?"
        params.append(fecha_filtro)

        fecha_texto    = f"Fecha: {fecha_filtro}"
        nombre_archivo = f"Asistencias_{fecha_filtro}.xlsx"

    else:
        # 📅 Día actual
        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        query += " AND asi.fecha = ?"
        params.append(fecha_hoy)

        fecha_texto    = f"Fecha: {fecha_hoy}"
        nombre_archivo = f"Asistencias_{fecha_hoy}.xlsx"

    # ==============================
    # 3) Filtro por SEDE según usuario
    # ==============================
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario sin sedes → no debe ver nada
            query += " AND 1=0"

    # ==============================
    # 4) Orden
    # ==============================
    query += " ORDER BY asi.fecha DESC, a.nombres, asi.hora"

    # Ejecutar
    cur.execute(query, params)
    filas = cur.fetchall()

    # 🧾 Si no hay datos
    if not filas:
        con.close()
        return "⚠️ No se encontraron registros en el rango seleccionado."

    columnas = [desc[0] for desc in cur.description]
    con.close()

    # 🧾 Crear DataFrame
    df = pd.DataFrame(filas, columns=columnas)

    # 🔖 Renombrar columnas
    df.rename(columns={
        "id_alumno": "ID Alumno",
        "codigo": "Código",
        "dni": "DNI",
        "nombres": "Nombres y Apellidos",
        "correo": "Correo Electrónico",
        "celular": "Celular",
        "institucion": "Institución",
        "curso": "Curso",
        "aula_matricula": "Aula",
        "turno": "Horario/Turno",
        "sede": "Sede",
        "fecha": "Fecha",
        "hora": "Hora",
        "tipo_registro": "Tipo de Registro",
        "estado_pago": "Estado de Pago"
    }, inplace=True)

    # 📤 Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=4, sheet_name='Asistencias')

        sheet = writer.sheets['Asistencias']

        # 🏷 Encabezado
        last_col_letter = get_column_letter(df.shape[1])

        sheet.merge_cells(f"A1:{last_col_letter}1")
        sheet["A1"] = "REPORTE DE ASISTENCIAS"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(f"A2:{last_col_letter}2")
        sheet["A2"] = "Institución: JGM - ICILT"
        sheet["A2"].font = Font(bold=True, size=12)
        sheet["A2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(f"A3:{last_col_letter}3")
        sheet["A3"] = fecha_texto
        sheet["A3"].alignment = Alignment(horizontal="center")

        # 📏 Ajustar anchos (ajusta a tu gusto)
        col_widths = {
            "A": 11,  # ID alumno
            "B": 16,  # Código
            "C": 12,  # DNI
            "D": 42,  # Nombre
            "E": 35,  # Correo
            "F": 30,  # Celular
            "G": 18,  # Institución
            "H": 45,  # Curso
            "I": 12,  # Aula
            "J": 50,  # Horario/Turno
            "K": 16,  # Sede
            "L": 14,  # Fecha
            "M": 22,  # Hora
            "N": 35,  # Tipo de registro
            "O": 30,  # Estado de pago
        }
        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width

    output.seek(0)

    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#=====================================================
# 🔹 RUTA: Debug alumnos (mostrar filas y keys)
@app.route('/debug_alumnos')
def debug_alumnos():
    con = get_db_connection()
    filas = con.execute("SELECT * FROM alumnos LIMIT 10").fetchall()
    con.close()
    out = ["Rows and keys:"]
    for r in filas:
        # mostrar las keys de cada Row y su contenido
        out.append(str(list(r.keys())))
        out.append(str(dict(r)))
    return "<pre>\n" + "\n".join(out) + "\n</pre>"

#=====================================================
# 🔹 RUTA: Editar alumno
@app.route("/editar/<int:id>", methods=["GET", "POST"])
@app.route("/editar/<int:id_alumno>", methods=["GET", "POST"])
def editar(id=None, id_alumno=None):
    target_id = id if id is not None else id_alumno
    if target_id is None:
        return redirect(url_for("index"))

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    rol_usuario = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")

    if rol_usuario == "admin":
        cursos = cur.execute("""
            SELECT id_curso, nombre
            FROM cursos
            ORDER BY nombre ASC
        """).fetchall()
    else:
        cursos = cur.execute("""
            SELECT id_curso, nombre
            FROM cursos
            WHERE id_sede = ?
            ORDER BY nombre ASC
        """, (id_sede_usuario,)).fetchall()

    matricula = cur.execute("""
        SELECT id_curso
        FROM matriculas
        WHERE id_alumno = ?
        ORDER BY id_matricula DESC
        LIMIT 1
    """, (target_id,)).fetchone()

    id_curso_actual = matricula["id_curso"] if matricula else None

    if request.method == "POST":
        nombres   = (request.form.get("nombres")   or "").strip()
        apellidos = (request.form.get("apellidos") or "").strip()
        dni = (request.form.get("dni") or "").strip()
        fecha_nacimiento = request.form.get("fecha_nacimiento") or None
        celular = (request.form.get("celular") or "").strip()
        correo = (request.form.get("correo") or "").strip()
        estado = (request.form.get("estado") or "Matriculado").strip()
        id_curso_nuevo = request.form.get("id_curso")
        institucion = request.form.get('institucion')
        grado_academico = request.form.get("grado_academico")

        nombres_guardar = f"{nombres} {apellidos}".strip()
        print("DEBUG EDITAR:", repr(nombres), repr(apellidos), "=>", repr(nombres_guardar))

        if rol_usuario != "admin":
            fila = cur.execute("""
                SELECT 1
                FROM cursos
                WHERE id_curso = ? AND id_sede = ?
            """, (id_curso_nuevo, id_sede_usuario)).fetchone()
            if not fila:
                con.close()
                return redirect(url_for("editar", id=target_id))

        try:
            cur.execute("""
                UPDATE alumnos
                SET nombres = ?, dni = ?, fecha_nacimiento = ?, celular = ?, correo = ?, 
                    institucion = ?, grado_academico = ?, estado = ?
                WHERE id_alumno = ?
            """, (
                nombres_guardar,
                dni,
                fecha_nacimiento,
                celular,
                correo,
                institucion,
                grado_academico,
                estado,
                target_id
            ))

            if matricula:
                cur.execute("""
                    UPDATE matriculas
                    SET id_curso = ?
                    WHERE id_alumno = ?
                """, (id_curso_nuevo, target_id))
            else:
                cur.execute("""
                    INSERT INTO matriculas (id_alumno, id_curso, fecha_matricula, deuda, estado)
                    VALUES (?, ?, DATE('now'), 0, ?)
                """, (target_id, id_curso_nuevo, estado))

            con.commit()
            con.close()
            return redirect(url_for("control_alumnos"))

        except Exception as e:
            print("⚠️ Error al actualizar alumno:", e)
            con.close()
            return redirect(url_for("editar", id=target_id))

    alumno = cur.execute("SELECT * FROM alumnos WHERE id_alumno = ?", (target_id,)).fetchone()
    con.close()

    if not alumno:
        return redirect(url_for("index"))

    full_name = (alumno["nombres"] or "").strip()
    last_space = full_name.rfind(" ")

    if last_space == -1:
        nombres_val = full_name
        apellidos_val = ""
    else:
        nombres_val = full_name[:last_space]
        apellidos_val = full_name[last_space + 1:]

    grado_academico = [
        "Sin secundaria",
        "Secundaria completa",
        "Técnico incompleto",
        "Técnico completo",
        "Universitario incompleto",
        "Universitario completo"
    ]

    return render_template(
        "editar_alumno.html",
        alumno=alumno,
        nombres_val=nombres_val,
        apellidos_val=apellidos_val,
        cursos=cursos,
        id_curso_actual=id_curso_actual,
        grados=grado_academico
    )

#=====================================================
# 🔹 RUTA: Ver QR
@app.route('/qrcodes/<codigo>')
def ver_qr(codigo):
    filename = f"{codigo}.png"
    full_path = os.path.join(QR_FOLDER, filename)
    app.logger.debug("Ver QR solicitado: %s -> %s", codigo, full_path)

    if not os.path.isfile(full_path):
        app.logger.warning("QR no encontrado: %s", full_path)
        abort(404)

    try:
        return send_from_directory(QR_FOLDER, filename, mimetype='image/png')
    except Exception as e:
        app.logger.exception("Error sirviendo QR %s: %s", full_path, e)
        abort(500)

#=====================================================
# 🔹 RUTA: Eliminar alumno
@app.route('/eliminar/<int:id_alumno>', methods=['GET'])
def eliminar(id_alumno):
    con = get_db_connection()
    cur = con.cursor()
    try:
        # Borra asistencias y matriculas explícitamente por seguridad si la BD no tiene cascada
        cur.execute("""
            DELETE FROM asistencias
            WHERE id_matricula IN (SELECT id_matricula FROM matriculas WHERE id_alumno=?)
        """, (id_alumno,))
        cur.execute("DELETE FROM matriculas WHERE id_alumno=?", (id_alumno,))
        cur.execute("DELETE FROM alumnos WHERE id_alumno=?", (id_alumno,))
        con.commit()
    except Exception as e:
        app.logger.exception("Error eliminando alumno: %s", e)
    finally:
        con.close()
    return redirect(url_for('index'))

#=====================================================
# 🔹 RUTA: Administrar docentes (por asignaciones)
@app.route('/administrar_docentes')
def administrar_docentes():
    if "usuario" not in session:
        return redirect(url_for("login"))

    rol          = session.get("rol")
    sedes_ids    = session.get("sedes", [])   # lista de sedes que puede ver
    id_sede_user = session.get("id_sede")     # sede principal (si solo tiene una)

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # 👇 Ahora la consulta parte de asignaciones_docentes
    query = """
        SELECT
            ad.id_asignacion,

            d.id_docente,
            d.nombres,
            d.dni,
            d.celular,
            d.correo,
            d.especialidad,
            d.profesion,
            d.codigo,

            c.id_curso,
            c.nombre       AS curso_nombre,

            s.id_sede,
            s.nombre_sede  AS sede_nombre,

            h.id_horario,
            h.descripcion  AS horario_nombre,
            h.hora_inicio,
            h.hora_fin

        FROM asignaciones_docentes ad
        JOIN docentes d ON d.id_docente = ad.id_docente
        LEFT JOIN cursos   c ON c.id_curso   = ad.id_curso
        LEFT JOIN horarios h ON h.id_horario = ad.id_horario
        LEFT JOIN sedes    s ON s.id_sede    = c.id_sede
        WHERE 1=1
    """

    params = []

    # 🔒 Filtro por sede según el usuario
    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes → ve solo sus sedes
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        elif id_sede_user:
            # Usuario con una sola sede
            query += " AND c.id_sede = ?"
            params.append(id_sede_user)

    query += """
        ORDER BY 
            d.nombres ASC,
            curso_nombre ASC,
            horario_nombre ASC
    """

    docentes = cur.execute(query, params).fetchall()
    con.close()

    return render_template("administrar_docentes.html", docentes=docentes)

#=====================================================
# 🔹 RUTA: Agregar docente
@app.route('/agregar_docente', methods=['GET', 'POST'])
def agregar_docente():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol          = session.get("rol")
    sedes_ids    = session.get("sedes", [])   # lista de sedes del usuario (si tiene varias)
    id_sede_user = session.get("id_sede")     # sede principal

    # ---------- POST: guardar ----------
    if request.method == 'POST':
        nombres      = (request.form.get('nombres') or "").strip()
        dni          = (request.form.get('dni') or "").strip()
        celular      = (request.form.get('celular') or "").strip()
        correo       = (request.form.get('correo') or "").strip()
        especialidad = (request.form.get('especialidad') or "").strip()
        profesion    = (request.form.get('profesion') or "").strip()

        id_curso   = request.form.get('id_curso')   or None
        id_horario = request.form.get('id_horario') or None
        id_sede_form = request.form.get('id_sede')  or None
        codigo     = (request.form.get('codigo') or "").strip()

        # Validación básica
        if not nombres or not dni or not id_horario:
            return redirect(url_for('agregar_docente'))

        con = get_db_connection()
        con.row_factory = sqlite3.Row
        cur = con.cursor()

        # 🔒 Validar que el curso pertenece a una sede permitida
        if id_curso:
            if rol == "admin":
                fila = cur.execute(
                    "SELECT id_sede FROM cursos WHERE id_curso = ?",
                    (id_curso,)
                ).fetchone()
            else:
                # usuario normal → solo sus sedes
                if sedes_ids:
                    placeholders = ",".join("?" * len(sedes_ids))
                    params = [id_curso] + sedes_ids
                    fila = cur.execute(
                        f"""
                        SELECT id_sede
                        FROM cursos
                        WHERE id_curso = ?
                          AND id_sede IN ({placeholders})
                        """,
                        params
                    ).fetchone()
                else:
                    fila = cur.execute(
                        """
                        SELECT id_sede
                        FROM cursos
                        WHERE id_curso = ?
                          AND id_sede = ?
                        """,
                        (id_curso, id_sede_user)
                    ).fetchone()

            if not fila:
                con.close()
                # curso no permitido para este usuario
                return redirect(url_for('agregar_docente'))

        # Generar código si no viene
        if not codigo:
            import time
            base = f"DOC{int(time.time())}"
            codigo = base
            cnt = 0
            while True:
                cur.execute("SELECT 1 FROM docentes WHERE codigo = ?", (codigo,))
                if not cur.fetchone():
                    break
                cnt += 1
                codigo = f"{base}_{cnt}"

        # Insertar docente
        cur.execute("""
            INSERT INTO docentes (nombres, dni, celular, correo, especialidad, profesion, codigo)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (nombres, dni, celular, correo, especialidad, profesion, codigo))

        id_docente = cur.lastrowid

        # Insertar asignación (docente-curso-horario)
        cur.execute("""
            INSERT INTO asignaciones_docentes (id_docente, id_curso, id_horario)
            VALUES (?, ?, ?)
        """, (id_docente, id_curso, id_horario))

        con.commit()
        con.close()

        return redirect(url_for('administrar_docentes'))

    # ---------- GET: cargar combos ----------
    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # Horarios (no dependen de sede)
    horarios = cur.execute("""
        SELECT id_horario, nombre_horario, hora_inicio, hora_fin
        FROM horarios
        ORDER BY nombre_horario
    """).fetchall()

    # Sedes y cursos según rol
    if rol == "admin":
        sedes = cur.execute("""
            SELECT id_sede, nombre_sede
            FROM sedes
            ORDER BY nombre_sede
        """).fetchall()

        cursos = cur.execute("""
            SELECT id_curso, nombre, id_sede
            FROM cursos
            ORDER BY nombre
        """).fetchall()
    else:
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            sedes = cur.execute(f"""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre_sede
            """, sedes_ids).fetchall()

            cursos = cur.execute(f"""
                SELECT id_curso, nombre, id_sede
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre
            """, sedes_ids).fetchall()
        else:
            sedes = cur.execute("""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede = ?
            """, (id_sede_user,)).fetchall()

            cursos = cur.execute("""
                SELECT id_curso, nombre, id_sede
                FROM cursos
                WHERE id_sede = ?
                ORDER BY nombre
            """, (id_sede_user,)).fetchall()

    con.close()

    return render_template(
        'agregar_docente.html',
        cursos=cursos,
        horarios=horarios,
        sedes=sedes
    )

#=====================================================
# 🔹 RUTA: Editar docente
@app.route('/editar_docente/<int:id_docente>', methods=['GET', 'POST'])
def editar_docente(id_docente):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol          = session.get("rol")
    sedes_ids    = session.get("sedes", [])
    id_sede_user = session.get("id_sede")

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # 👉 Datos del docente
    docente = cur.execute(
        "SELECT * FROM docentes WHERE id_docente = ?",
        (id_docente,)
    ).fetchone()

    if not docente:
        con.close()
        return redirect(url_for('administrar_docentes'))

    # 👉 Asignación actual (curso + horario + sede del curso)
    asignacion = cur.execute("""
        SELECT 
            ad.id_curso,
            ad.id_horario,
            c.id_sede
        FROM asignaciones_docentes ad
        LEFT JOIN cursos c ON c.id_curso = ad.id_curso
        WHERE ad.id_docente = ?
        LIMIT 1
    """, (id_docente,)).fetchone()

    id_curso_actual   = asignacion["id_curso"]   if asignacion else None
    id_horario_actual = asignacion["id_horario"] if asignacion else None
    id_sede_actual    = asignacion["id_sede"]    if asignacion else None

    # 👉 Horarios (no dependen de sede)
    horarios = cur.execute("""
        SELECT id_horario, nombre_horario, hora_inicio, hora_fin
        FROM horarios
        ORDER BY nombre_horario
    """).fetchall()

    # 👉 Sedes + cursos, filtrados por rol
    if rol == "admin":
        sedes = cur.execute("""
            SELECT id_sede, nombre_sede
            FROM sedes
            ORDER BY nombre_sede
        """).fetchall()

        cursos = cur.execute("""
            SELECT id_curso, nombre, id_sede
            FROM cursos
            ORDER BY nombre
        """).fetchall()
    else:
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            sedes = cur.execute(f"""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre_sede
            """, sedes_ids).fetchall()

            cursos = cur.execute(f"""
                SELECT id_curso, nombre, id_sede
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre
            """, sedes_ids).fetchall()
        else:
            sedes = cur.execute("""
                SELECT id_sede, nombre_sede
                FROM sedes
                WHERE id_sede = ?
            """, (id_sede_user,)).fetchall()

            cursos = cur.execute("""
                SELECT id_curso, nombre, id_sede
                FROM cursos
                WHERE id_sede = ?
                ORDER BY nombre
            """, (id_sede_user,)).fetchall()

    # ---------- POST: guardar cambios ----------
    if request.method == 'POST':
        nombres      = (request.form.get('nombres') or "").strip()
        dni          = (request.form.get('dni') or "").strip()
        celular      = (request.form.get('celular') or "").strip()
        correo       = (request.form.get('correo') or "").strip()
        especialidad = (request.form.get('especialidad') or "").strip()
        profesion    = (request.form.get('profesion') or "").strip()

        id_curso_new   = request.form.get('id_curso')   or None
        id_horario_new = request.form.get('id_horario') or None
        id_sede_form   = request.form.get('id_sede')    or None

        if not nombres or not dni or not id_horario_new:
            con.close()
            return redirect(url_for('editar_docente', id_docente=id_docente))

        # 🔒 Validar que el curso pertenece a una sede permitida
        if id_curso_new:
            if rol == "admin":
                fila = cur.execute(
                    "SELECT id_sede FROM cursos WHERE id_curso = ?",
                    (id_curso_new,)
                ).fetchone()
            else:
                if sedes_ids:
                    placeholders = ",".join("?" * len(sedes_ids))
                    params = [id_curso_new] + sedes_ids
                    fila = cur.execute(
                        f"""
                        SELECT id_sede
                        FROM cursos
                        WHERE id_curso = ?
                          AND id_sede IN ({placeholders})
                        """,
                        params
                    ).fetchone()
                else:
                    fila = cur.execute(
                        """
                        SELECT id_sede
                        FROM cursos
                        WHERE id_curso = ?
                          AND id_sede = ?
                        """,
                        (id_curso_new, id_sede_user)
                    ).fetchone()

            if not fila:
                con.close()
                return redirect(url_for('editar_docente', id_docente=id_docente))

        try:
            # Actualizar docente
            cur.execute("""
                UPDATE docentes
                SET nombres=?, dni=?, celular=?, correo=?, especialidad=?, profesion=?
                WHERE id_docente=?
            """, (nombres, dni, celular, correo, especialidad, profesion, id_docente))

            # Actualizar / crear asignación
            fila_asig = cur.execute("""
                SELECT id_asignacion
                FROM asignaciones_docentes
                WHERE id_docente = ?
                LIMIT 1
            """, (id_docente,)).fetchone()

            if fila_asig:
                cur.execute("""
                    UPDATE asignaciones_docentes
                    SET id_curso = ?, id_horario = ?
                    WHERE id_docente = ?
                """, (id_curso_new, id_horario_new, id_docente))
            else:
                cur.execute("""
                    INSERT INTO asignaciones_docentes (id_docente, id_curso, id_horario)
                    VALUES (?, ?, ?)
                """, (id_docente, id_curso_new, id_horario_new))

            con.commit()
        except Exception as e:
            app.logger.exception("Error al editar docente: %s", e)
        finally:
            con.close()

        return redirect(url_for('administrar_docentes'))

    # ---------- GET: mostrar formulario ----------
    con.close()
    return render_template(
        'editar_docente.html',
        docente=docente,
        cursos=cursos,
        horarios=horarios,
        sedes=sedes,
        id_curso_actual=id_curso_actual,
        id_horario_actual=id_horario_actual,
        id_sede_actual=id_sede_actual
    )

#=====================================================
# 🔹 RUTA: Eliminar docente
@app.route('/eliminar_docente/<int:id_docente>')
def eliminar_docente(id_docente):
    try:
        con = get_db_connection()

        # 1. Quitar asignación en cursos
        con.execute("UPDATE cursos SET id_docente=NULL WHERE id_docente=?", (id_docente,))

        # 2. Borrar asistencias del docente (pierdes historial)
        con.execute("DELETE FROM asistencias_docentes WHERE id_docente=?", (id_docente,))

        # 3. Desvincular asignaciones
        con.execute("UPDATE asignaciones_docentes SET id_docente=NULL WHERE id_docente=?", (id_docente,))

        # 4. Ahora sí borrar el docente
        con.execute("DELETE FROM docentes WHERE id_docente=?", (id_docente,))

        con.commit()
        con.close()
        return redirect(url_for('administrar_docentes'))

    except Exception as e:
        app.logger.exception("Error al eliminar docente: %s", e)
        return redirect(url_for('administrar_docentes'))

 
#=====================================================
# 🔹 RUTA: Listar responsables
@app.route('/responsables')
def responsables():
    
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # 🚨 PROTECCIÓN POR ROL
    if session.get("rol") != "admin":
        return redirect(url_for('dashboard'))  # O una página 403

    
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    # Consulta todos los responsables
    cur.execute("""
        SELECT id_responsable,
               nombres,
               dni,
               celular,
               correo,
               cargo
        FROM responsables
        ORDER BY id_responsable DESC
    """)

    responsables = cur.fetchall()
    conexion.close()

    return render_template('responsables.html', responsables=responsables, rol=session.get("rol") )
# ======================================================
# 🔹 RUTA: Agregar nuevo responsable
@app.route('/agregar_responsable', methods=['GET', 'POST'])
def agregar_responsable():
    if request.method == 'POST':
        nombres = request.form.get('nombres')
        dni = request.form.get('dni')
        celular = request.form.get('celular')
        correo = request.form.get('correo')
        cargo = request.form.get('cargo')

        conexion = sqlite3.connect('base_datos.db')
        cur = conexion.cursor()
        cur.execute("""
            INSERT INTO responsables (nombres, dni, celular, correo, cargo)
            VALUES (?, ?, ?, ?, ?)
        """, (nombres, dni, celular, correo, cargo))
        conexion.commit()
        conexion.close()

        return redirect(url_for('responsables'))

    return render_template('agregar_responsable.html')
# ======================================================
# 🔹 RUTA: Editar responsable existente        
@app.route('/editar_responsable/<int:id_responsable>', methods=['GET', 'POST'])
def editar_responsable(id_responsable):
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    if request.method == 'POST':
        nombres = request.form['nombres']
        dni = request.form['dni']
        celular = request.form['celular']
        correo = request.form['correo']
        cargo = request.form['cargo']

        cur.execute("""
            UPDATE responsables
            SET nombres = ?, dni = ?, celular = ?, correo = ?, cargo = ?
            WHERE id_responsable = ?
        """, (nombres, dni, celular, correo, cargo, id_responsable))

        conexion.commit()
        conexion.close()
        return redirect(url_for('responsables'))

    # Obtener los datos actuales del responsable
    cur.execute("SELECT * FROM responsables WHERE id_responsable = ?", (id_responsable,))
    responsable = cur.fetchone()
    conexion.close()

    if responsable is None:
        return redirect(url_for('responsables'))

    return render_template('editar_responsable.html', responsable=responsable)

# ======================================================
# 🔹 RUTA: Eliminar responsable
@app.route('/eliminar_responsable/<int:id_responsable>', methods=['GET'])
def eliminar_responsable(id_responsable):
    conexion = sqlite3.connect('base_datos.db')
    cur = conexion.cursor()
    cur.execute("DELETE FROM responsables WHERE id_responsable = ?", (id_responsable,))
    conexion.commit()
    conexion.close()
    return redirect(url_for('responsables'))

#=====================================================
# 🔹 RUTA: Administrar matrículas (multi-sede)
@app.route('/administrar_matriculas')
def administrar_matriculas():

    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    id_sede_usuario = session.get("id_sede")      # primera sede (compatibilidad)
    sedes_ids = session.get("sedes") or []        # lista de sedes permitidas

    # fallback por si todavía no tienes multi-sede completo
    if not sedes_ids and id_sede_usuario:
        sedes_ids = [id_sede_usuario]

    # 🧩 Filtros por GET
    curso_filtro   = request.args.get("curso") or ""
    buscar_alumno  = (request.args.get("buscar_alumno") or "").strip()

    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    # ============================
    # 1) LISTA DE CURSOS
    # ============================
    if rol == "admin":
        cur.execute("SELECT DISTINCT nombre FROM cursos ORDER BY nombre ASC")
        cursos = [r[0] for r in cur.fetchall()]
    else:
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            cur.execute(f"""
                SELECT DISTINCT nombre
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre ASC
            """, sedes_ids)
            cursos = [r[0] for r in cur.fetchall()]
        else:
            # usuario sin sedes → no ve nada
            cursos = []

    # ============================
    # 2) CONSULTA BASE DE MATRÍCULAS
    # ============================
    query = """
        SELECT 
            m.id_matricula,
            a.codigo AS codigo_alumno,
            a.nombres AS nombre_alumno,
            a.institucion AS institucion_alumno,
            COALESCE(c.nombre, 'SIN CURSO') AS nombre_curso,
            c.fecha_inicio AS fecha_inicio_curso,
            c.fecha_fin AS fecha_fin_curso,
            h.nombre_horario AS turno,
            m.fecha_matricula,
            m.deuda,
            m.monto,
            m.observacion, 
            m.estado,
            COALESCE(s.nombre_sede, 'SIN SEDE') AS nombre_sede,
            m.tipo_pago,
            m.metodo_pago,
            m.saldo_matricula,
            m.fecha_vencimiento,
            m.mensualidad,
            m.aula
        FROM matriculas m
        JOIN alumnos a ON m.id_alumno = a.id_alumno
        LEFT JOIN cursos c ON m.id_curso = c.id_curso     -- 👈 AQUÍ EL CAMBIO CLAVE
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        LEFT JOIN horarios h ON c.id_horario = h.id_horario
        WHERE 1=1
    """


    params = []

    # 🎯 Filtro por curso
    if curso_filtro:
        query += " AND IFNULL(c.nombre, '') = ?"
        params.append(curso_filtro)

    # 🎯 Filtro por alumno (nombre, DNI o código)
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # 🎯 Filtro por sede si NO es admin (multi-sede)
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Sin sedes asignadas → no devuelve nada
            query += " AND 1=0"

    query += " ORDER BY m.id_matricula DESC"

    cur.execute(query, params)
    matriculas = cur.fetchall()
    conexion.close()

    return render_template(
        'administrar_matriculas.html',
        matriculas=matriculas,
        cursos=cursos,
        curso_filtro=curso_filtro,
        buscar_alumno=buscar_alumno
    )

#===========================================
# EXPORTAR MATRÍCULAS A EXCEL
@app.route('/exportar_matriculas_excel')
def exportar_matriculas_excel():
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    import pandas as pd

    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = session.get("rol")
    id_sede_usuario = session.get("id_sede")

    # 🔹 Filtros por GET (mismos que en administrar_matriculas / exportar_matriculas_pdf)
    curso_filtro  = request.args.get("curso") or ""
    buscar_alumno = (request.args.get("buscar_alumno") or "").strip()

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    query = """
    SELECT 
        m.id_matricula,
        a.codigo        AS codigo_alumno,
        a.nombres       AS alumno,
        a.dni,
        a.correo,
        a.celular,
        a.institucion,
        s.nombre_sede   AS sede,
        c.nombre        AS curso,
        m.aula          AS aula,
        c.fecha_inicio  AS fecha_inicio_curso,
        c.fecha_fin     AS fecha_fin_curso,
        h.descripcion   AS horario,
        m.fecha_matricula,
        m.fecha_vencimiento,
        m.monto,
        m.mensualidad,
        m.saldo_matricula,
        m.deuda,
        m.tipo_pago,
        m.metodo_pago,
        m.observacion,
        m.estado
    FROM matriculas m
    JOIN alumnos  a ON m.id_alumno = a.id_alumno
    JOIN cursos   c ON m.id_curso  = c.id_curso
    LEFT JOIN sedes    s ON c.id_sede    = s.id_sede
    LEFT JOIN horarios h ON c.id_horario = h.id_horario
    WHERE 1=1
    """

    params = []

    # 🎯 Filtro por curso
    if curso_filtro:
        query += " AND IFNULL(c.nombre, '') = ?"
        params.append(curso_filtro)

    # 🎯 Filtro por alumno (nombre, DNI o código)
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # 🎯 Filtro por sede si NO es admin
    if rol != "admin":
        query += " AND IFNULL(c.id_sede, 0) = ?"
        params.append(id_sede_usuario)

    query += " ORDER BY m.id_matricula DESC"

    cur.execute(query, params)
    filas = cur.fetchall()

    columnas = [desc[0] for desc in cur.description]
    con.close()

    # 🧾 Si no hay datos
    if not filas:
        return "⚠️ No se encontraron matrículas con los filtros seleccionados."

    # 🧾 DataFrame
    df = pd.DataFrame(filas, columns=columnas)

    # 🔹 Renombrar columnas a algo más amigable
    df.rename(columns={
        "codigo_alumno":      "Código Alumno",
        "alumno":             "Alumno",
        "institucion":        "Institución",
        "curso":              "Curso",
        "aula":               "Aula",
        "fecha_inicio_curso": "Inicio de Curso",
        "fecha_fin_curso":    "Final de Curso",
        "sede":               "Sede",
        "fecha_matricula":    "Fecha Matrícula",
        "monto":              "Monto Matrícula",
        "saldo_matricula":    "Saldo Matrícula",
        "mensualidad":        "Mensualidad",
        "fecha_vencimiento":  "Fecha Vencimiento de Mensualidad",
        "horario":            "Horario/Turno",
        "deuda":              "Deuda",
        "observacion":        "Observación",
    }, inplace=True)

    # 🔹 Dejar SOLO las columnas que vas a ver en el Excel
    df = df[
        [
            "Código Alumno",
            "Alumno",
            "Institución",
            "Curso",
            "Aula",
            "Inicio de Curso",
            "Final de Curso",
            "Sede",
            "Fecha Matrícula",
            "Monto Matrícula",
            "Saldo Matrícula",
            "Mensualidad",
            "Fecha Vencimiento de Mensualidad",
            "Horario/Turno",
            "Deuda",
            "Observación",
        ]
    ]

    # 📤 Crear Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=4, sheet_name='Matrículas')

        sheet = writer.sheets['Matrículas']

        # 🏷 Encabezados superiores (A hasta O = 15 columnas)
        last_col_letter = "O"

        sheet.merge_cells(f"A1:{last_col_letter}1")
        sheet["A1"] = "REPORTE DE MATRÍCULAS"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(f"A2:{last_col_letter}2")
        sheet["A2"] = f"Curso: {curso_filtro or 'Todos'}"
        sheet["A2"].alignment = Alignment(horizontal="center")
        sheet["A2"].font = Font(bold=True, size=12)

        sheet.merge_cells(f"A3:{last_col_letter}3")
        sheet["A3"] = f"Filtro alumno: {buscar_alumno or 'Todos'}"
        sheet["A3"].alignment = Alignment(horizontal="center")

        # 📏 Anchos de columnas (A–O)
        col_widths = {
            "A": 15,  # COD ALUMNO  
            "B": 40,  # ALUMNO
            "C": 10,  # ISNTITUCION
            "D": 40,  # CURSO
            "E": 10,  # AULA
            "F": 20,  # FECHA INICIAL DE CURSO
            "G": 20,  # FECHA FINAL DE CURSO
            "H": 16,  # SEDE
            "I": 16,  # FECHA DE MATRICUSLA 
            "J": 20,  # MONTO DE MATRICULA 
            "K": 15,  # SALDO DE MATRICULA
            "L": 15,  # MENSUALIDAD
            "M": 32,  # VENCIMIENTO MENSUALIDAD
            "N": 45,  # HORARIO
            "O": 15,  # DEUDA
            "P": 20,  # OBSERVACIÓN

        }

        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width

    output.seek(0)

    nombre_archivo = "Matrículas.xlsx"
    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#=====================================================
# 🔹 RUTA: Exportar reporte general de matriculados a EXCEL
@app.route('/exportar_reporte_matriculados')
def exportar_reporte_matriculados():
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import pandas as pd

    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # 👈 lista de sedes para usuarios multi-sede

    con = get_db_connection()
    cur = con.cursor()

    # Filtro base por sede (como en reporte_total, pero soportando varias sedes)
    where = "WHERE 1=1"
    params = []

    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes asignadas
            placeholders = ",".join("?" * len(sedes_ids))
            where += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario con una sola sede
            where += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    # 🔍 Traemos EL DETALLE de todos los matriculados
    cur.execute(f"""
        SELECT 
            a.nombres                         AS alumno,
            a.dni                             AS dni,
            a.celular                         AS celular,
            a.correo                          AS correo,
            c.nombre                          AS curso,
            m.aula                            AS aula,
            IFNULL(s.nombre_sede,'SIN SEDE')  AS sede,
            m.fecha_matricula                 AS fecha_matricula,
            m.monto                           AS monto,
            m.mensualidad                     AS mensualidad
        FROM matriculas m
        JOIN alumnos  a ON m.id_alumno = a.id_alumno
        JOIN cursos   c ON m.id_curso  = c.id_curso
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        {where}
        ORDER BY sede ASC, curso ASC, alumno ASC
    """, params)

    filas = cur.fetchall()
    columnas = [desc[0] for desc in cur.description]
    con.close()

    if not filas:
        return "⚠️ No hay matrículas para exportar."

    # dataframe base
    df = pd.DataFrame(filas, columns=columnas)

    # Aseguramos numéricos
    for col in ["monto", "mensualidad"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Total general
    total_matriculas = len(df)

    # Resumen por curso
    df_cursos = (
        df.groupby("curso")
        .size()
        .reset_index(name="Total Matriculados")
        .sort_values("curso")
    )

    # Resumen por sede
    df_sedes = (
        df.groupby("sede")
        .size()
        .reset_index(name="Total Alumnos")
        .sort_values("sede")
    )

    # Copia con nombres amigables para la hoja de detalle
    df_detalle = df.rename(columns={
        "alumno":          "Alumno",
        "dni":             "DNI",
        "celular":         "Celular",
        "correo":          "Correo",
        "curso":           "Curso",
        "aula":            "Aula",
        "sede":            "Sede",
        "fecha_matricula": "Fecha Matrícula",
        "monto":           "Monto Matrícula",
        "mensualidad":     "Mensualidad",
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 🧾 Hoja 1: Detalle de alumnos matriculados
        df_detalle.to_excel(
            writer,
            index=False,
            startrow=4,
            sheet_name="Detalle Matriculados"
        )
        sheet_det = writer.sheets["Detalle Matriculados"]

        last_col_letter = get_column_letter(df_detalle.shape[1])

        # Título
        sheet_det.merge_cells(f"A1:{last_col_letter}1")
        sheet_det["A1"] = "REPORTE DE MATRÍCULAS"
        sheet_det["A1"].font = Font(bold=True, size=14)
        sheet_det["A1"].alignment = Alignment(horizontal="center")

        # Subtítulo con total
        sheet_det.merge_cells(f"A2:{last_col_letter}2")
        sheet_det["A2"] = f"Total matriculados: {total_matriculas}"
        sheet_det["A2"].font = Font(bold=True, size=12)
        sheet_det["A2"].alignment = Alignment(horizontal="center")

        sheet_det["A3"] = "Detalle de alumnos matriculados"

        # Ancho de columnas
        col_widths = {
            "A": 45,  # Alumno
            "B": 12,  # DNI
            "C": 30,  # Celular
            "D": 40,  # Correo
            "E": 45,  # Curso
            "F": 10,  # Aula
            "G": 18,  # Sede
            "H": 16,  # Fecha Matrícula
            "I": 16,  # Monto Matrícula
            "J": 16,  # Mensualidad
        }
        for col_letter, width in col_widths.items():
            sheet_det.column_dimensions[col_letter].width = width

        # Hojas de resumen
        df_cursos.to_excel(writer, index=False, sheet_name="Resumen por curso")
        sheet_c = writer.sheets["Resumen por curso"]
        sheet_c["A1"].font = Font(bold=True)
        sheet_c["B1"].font = Font(bold=True)

        df_sedes.to_excel(writer, index=False, sheet_name="Resumen por sede")
        sheet_s = writer.sheets["Resumen por sede"]
        sheet_s["A1"].font = Font(bold=True)
        sheet_s["B1"].font = Font(bold=True)

    output.seek(0)

    return send_file(
        output,
        download_name="reporte_matriculas.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


#=====================================================
# 🔹 RUTA: Exportar alumnos morosos a EXCEL
@app.route('/exportar_alumnos_morosos')
def exportar_alumnos_morosos():
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import pandas as pd

    if 'usuario' not in session:
        return redirect(url_for('login'))

    # 🔍 Normalizar rol y sedes igual que en reporte_total
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes para usuarios multi-sede

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # Filtro base por sede (igual que en reporte_total, versión multi-sede)
    where_base = "WHERE 1=1"
    params_base = []

    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes asignadas
            placeholders = ",".join("?" * len(sedes_ids))
            where_base += f" AND c.id_sede IN ({placeholders})"
            params_base.extend(sedes_ids)
        elif id_sede_usuario:
            # Usuario con una sola sede
            where_base += " AND c.id_sede = ?"
            params_base.append(id_sede_usuario)
        # Si no hay nada en sesión, no filtro (peor es no ver nada)

    # Criterio de moroso (mismo que en reporte_total)
    estados_morosos = (
        "Con Deuda (Deuda Vencida)",
        "Moroso (Debe mas de 2 meses)"
    )

    where_morosos = (
        where_base +
        " AND ( m.estado IN (?, ?) OR IFNULL(m.deuda,0) > 0 )"
    )
    params_morosos = params_base + list(estados_morosos)

    # 🔍 Traer SOLO morosos con todos los datos del alumno
    cur.execute(f"""
        SELECT 
            a.nombres                         AS alumno,
            a.dni                             AS dni,
            a.celular                         AS celular,
            a.correo                          AS correo,
            c.nombre                          AS curso,
            IFNULL(s.nombre_sede,'SIN SEDE')  AS sede,
            m.fecha_matricula                 AS fecha_matricula,
            m.fecha_vencimiento               AS fecha_vencimiento,
            m.monto                           AS monto_matricula,
            m.saldo_matricula                 AS saldo_matricula,
            m.mensualidad                     AS mensualidad,
            m.deuda                           AS deuda,
            m.tipo_pago                       AS tipo_pago,
            m.metodo_pago                     AS metodo_pago,
            m.estado                          AS estado
        FROM matriculas m
        JOIN alumnos  a ON m.id_alumno = a.id_alumno
        JOIN cursos   c ON m.id_curso  = c.id_curso
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        {where_morosos}
        ORDER BY sede ASC, curso ASC, alumno ASC
    """, params_morosos)

    filas = cur.fetchall()
    columnas = [d[0] for d in cur.description]
    con.close()

    if not filas:
        return "⚠️ No hay alumnos morosos para exportar."

    df = pd.DataFrame(filas, columns=columnas)

    # Asegurar numéricos
    for col in ["monto_matricula", "saldo_matricula", "mensualidad", "deuda"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    total_morosos = len(df)

    # Resumen por curso (solo morosos)
    df_curso = (
        df.groupby("curso")
          .size()
          .reset_index(name="Total Morosos")
          .sort_values("curso")
    )

    # Resumen por sede (solo morosos)
    df_sede = (
        df.groupby("sede")
          .size()
          .reset_index(name="Total Morosos")
          .sort_values("sede")
    )

    # Detalle morosos con nombres de columnas amigables
    df_detalle = df.rename(columns={
        "alumno":            "Alumno",
        "dni":               "DNI",
        "celular":           "Celular",
        "correo":            "Correo",
        "curso":             "Curso",
        "sede":              "Sede",
        "fecha_matricula":   "Fecha Matrícula",
        "fecha_vencimiento": "Fecha Vencimiento",
        "monto_matricula":   "Monto Matrícula",
        "saldo_matricula":   "Saldo Matrícula",
        "mensualidad":       "Mensualidad",
        "deuda":             "Deuda",
        "tipo_pago":         "Tipo Pago",
        "metodo_pago":       "Método Pago",
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 🧾 Hoja 1: Detalle de morosos
        df_detalle.to_excel(
            writer,
            index=False,
            startrow=4,
            sheet_name="Morosos Detalle"
        )
        sh = writer.sheets["Morosos Detalle"]
        last_col = get_column_letter(df_detalle.shape[1])

        sh.merge_cells(f"A1:{last_col}1")
        sh["A1"] = "REPORTE DE ALUMNOS MOROSOS"
        sh["A1"].font = Font(bold=True, size=14)
        sh["A1"].alignment = Alignment(horizontal="center")

        sh.merge_cells(f"A2:{last_col}2")
        sh["A2"] = f"Total morosos: {total_morosos}"
        sh["A2"].font = Font(bold=True, size=12)
        sh["A2"].alignment = Alignment(horizontal="center")

        sh["A3"] = "Detalle de alumnos morosos"

        # Ancho de columnas personalizado
        col_widths = {
            "A": 45,  # Alumno
            "B": 12,  # DNI
            "C": 30,  # Celular
            "D": 40,  # Correo
            "E": 45,  # Curso
            "F": 18,  # Sede
            "G": 16,  # Fecha Matrícula
            "H": 25,  # Fecha Vencimiento
            "I": 18,  # Monto Matrícula
            "J": 16,  # Saldo Matrícula
            "K": 16,  # Mensualidad
            "L": 16,  # Deuda
            "M": 20,  # Tipo Pago
            "N": 20,  # Método Pago
        }

        for col_letter, width in col_widths.items():
            sh.column_dimensions[col_letter].width = width

    output.seek(0)

    return send_file(
        output,
        download_name="reporte_morosos.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

#=====================================================
# 🔹 RUTA: Reporte total de matrículas y morosos
@app.route('/reporte_total')
def reporte_total():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # Normalizar rol y sedes
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes (para usuarios multi-sede)

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # === Helper: filtro base para TODO lo que pasa por cursos (alias c) ===
    where_base = "WHERE 1=1"
    params_base = []

    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes asignadas
            placeholders = ",".join("?" * len(sedes_ids))
            where_base += f" AND c.id_sede IN ({placeholders})"
            params_base.extend(sedes_ids)
        elif id_sede_usuario:
            # Usuario con una sola sede
            where_base += " AND c.id_sede = ?"
            params_base.append(id_sede_usuario)
        # Si no hay ni sedes_ids ni id_sede_usuario → no agrego filtro (mejor ver todo que ver nada)

    # === 2) Total matrículas ===
    cur.execute(f"""
        SELECT COUNT(*) AS total_matriculas
        FROM matriculas m
        JOIN cursos c ON m.id_curso = c.id_curso
        {where_base}
    """, params_base)
    total_matriculas = cur.fetchone()["total_matriculas"]

    # === 3) Matrículas por curso ===
    cur.execute(f"""
        SELECT 
            c.nombre AS curso,
            COUNT(*) AS total
        FROM matriculas m
        JOIN cursos c ON m.id_curso = c.id_curso
        {where_base}
        GROUP BY c.id_curso, c.nombre
        ORDER BY c.nombre ASC
    """, params_base)
    matriculas_por_curso = cur.fetchall()

    # === 4) Alumnos por curso (detalle) ===
    cur.execute(f"""
        SELECT 
            c.nombre  AS curso,
            a.nombres AS alumno,
            a.celular AS celular
        FROM matriculas m
        JOIN alumnos a ON m.id_alumno = a.id_alumno
        JOIN cursos  c ON m.id_curso  = c.id_curso
        {where_base}
        ORDER BY c.nombre ASC, a.nombres ASC
    """, params_base)
    alumnos_por_curso = cur.fetchall()

    # === 5) Morosos (según tu lógica) ===
    estados_morosos = (
        "Con Deuda (Deuda Vencida)",
        "Moroso (Debe mas de 2 meses)"
    )

    where_morosos = (
        where_base +
        " AND ( m.estado IN (?, ?) OR IFNULL(m.deuda,0) > 0 )"
    )
    params_morosos = params_base + list(estados_morosos)

    # 5.1 Morosos por curso
    cur.execute(f"""
        SELECT 
            c.nombre AS curso,
            COUNT(*) AS total_morosos
        FROM matriculas m
        JOIN cursos c ON m.id_curso = c.id_curso
        {where_morosos}
        GROUP BY c.id_curso, c.nombre
        ORDER BY c.nombre ASC
    """, params_morosos)
    morosos_por_curso = cur.fetchall()

    # 5.2 Morosos por sede (vía cursos → sedes)
    cur.execute(f"""
        SELECT 
            IFNULL(s.nombre_sede, 'SIN SEDE') AS sede,
            COUNT(*) AS total_morosos
        FROM matriculas m
        JOIN cursos  c ON m.id_curso = c.id_curso
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        {where_morosos}
        GROUP BY s.id_sede, s.nombre_sede
        ORDER BY sede ASC
    """, params_morosos)
    morosos_por_sede = cur.fetchall()

    # 5.3 Detalle de morosos
    cur.execute(f"""
        SELECT 
            a.nombres              AS alumno,
            a.celular              AS celular,
            a.dni                  AS dni,
            c.nombre               AS curso,
            IFNULL(s.nombre_sede, 'SIN SEDE') AS sede,
            m.monto                AS monto_matricula,
            m.saldo_matricula      AS saldo_matricula,
            m.mensualidad          AS mensualidad,
            m.deuda                AS deuda,
            m.tipo_pago            AS tipo_pago,
            m.metodo_pago          AS metodo_pago,
            m.estado               AS estado,
            m.fecha_matricula      AS fecha_matricula,
            m.fecha_vencimiento    AS fecha_vencimiento
        FROM matriculas m
        JOIN alumnos a ON m.id_alumno = a.id_alumno
        JOIN cursos  c ON m.id_curso  = c.id_curso
        LEFT JOIN sedes s ON c.id_sede = s.id_sede
        {where_morosos}
        ORDER BY sede ASC, curso ASC, alumno ASC
    """, params_morosos)
    morosos_detalle = cur.fetchall()

    # 6) Alumnos por sede (aquí el filtro es por alumnos.id_sede)
    where_alumnos = "WHERE 1=1"
    params_alumnos = []

    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            where_alumnos += f" AND a.id_sede IN ({placeholders})"
            params_alumnos.extend(sedes_ids)
        elif id_sede_usuario:
            where_alumnos += " AND a.id_sede = ?"
            params_alumnos.append(id_sede_usuario)

    cur.execute(f"""
        SELECT 
            IFNULL(s.nombre_sede, 'SIN SEDE') AS sede,
            COUNT(*) AS total_alumnos
        FROM alumnos a
        LEFT JOIN sedes s ON a.id_sede = s.id_sede
        {where_alumnos}
        GROUP BY a.id_sede, s.nombre_sede
        ORDER BY sede ASC
    """, params_alumnos)
    alumnos_por_sede = cur.fetchall()

    con.close()

    # 👇 convertir todo a lista de diccionarios para |tojson
    def rows_to_dict_list(rows):
        return [dict(r) for r in rows]

    return render_template(
        "reportes.html",
        total_matriculas=total_matriculas,
        matriculas_por_curso=rows_to_dict_list(matriculas_por_curso),
        alumnos_por_curso=rows_to_dict_list(alumnos_por_curso),
        morosos_por_curso=rows_to_dict_list(morosos_por_curso),
        morosos_por_sede=rows_to_dict_list(morosos_por_sede),
        alumnos_por_sede=rows_to_dict_list(alumnos_por_sede),
        morosos_detalle=rows_to_dict_list(morosos_detalle),
    )


#=====================================================
# 🔹 RUTA: Exportar matrículas a PDF
@app.route('/exportar_matriculas_pdf')
def exportar_matriculas_pdf():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # 👈 lista de sedes para usuarios multi-sede

    curso = request.args.get("curso") or ""
    buscar_alumno = (request.args.get("buscar_alumno") or "").strip()

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # CONSULTA BASE
    query = """
    SELECT 
        m.fecha_matricula,
        a.nombres        AS alumno,
        a.dni,
        a.celular,
        a.institucion,
        s.nombre_sede    AS sede,
        c.nombre         AS curso,
        d.nombres        AS docente,
        h.descripcion    AS horario,
        m.monto,
        m.mensualidad,
        m.saldo_matricula,
        m.deuda,
        m.tipo_pago,
        m.metodo_pago,
        m.fecha_vencimiento,
        m.estado
    FROM matriculas m
    JOIN alumnos  a ON m.id_alumno = a.id_alumno
    JOIN cursos   c ON m.id_curso  = c.id_curso
    LEFT JOIN sedes    s ON c.id_sede    = s.id_sede
    LEFT JOIN horarios h ON c.id_horario = h.id_horario
    LEFT JOIN docentes d ON c.id_docente = d.id_docente
    WHERE 1=1
    """

    params = []

    # Filtro por curso
    if curso:
        query += " AND IFNULL(c.nombre, '') = ?"
        params.append(curso)

    # Filtro por alumno (nombre, DNI o código)
    if buscar_alumno:
        query += " AND (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)"
        like = f"%{buscar_alumno}%"
        params.extend([like, like, like])

    # 🔥 Filtro por sede según rol
    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes
            placeholders = ",".join("?" * len(sedes_ids))
            query += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario con una sola sede
            query += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    query += " ORDER BY m.id_matricula DESC"

    cur.execute(query, params)
    matriculas = cur.fetchall()
    con.close()

    # ====== GENERAR PDF EN MEMORIA (LANDSCAPE) ======
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=20,
        rightMargin=20
    )

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        "header_style",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.white
    )

    cell_style = styles["BodyText"]
    cell_style.fontSize = 8
    cell_style.leading  = 8

    elements = []

    titulo = Paragraph("<b>Reporte de Matrículas</b>", styles["Title"])
    elements.append(titulo)

    data = [[
        Paragraph("Fecha Matrícula", header_style),
        Paragraph("Alumno", header_style),
        Paragraph("Institución", header_style),
        Paragraph("Sede", header_style),
        Paragraph("Curso", header_style),
        Paragraph("Monto Matrícula", header_style),
        Paragraph("Fecha Vencimiento<br/>de Mensualidad", header_style),
        Paragraph("Mensualidad", header_style),
    ]]

    for m in matriculas:
        def to_float(v):
            try:
                return float(v) if v not in (None, "", " ") else 0.0
            except:
                return 0.0

        monto       = to_float(m["monto"])
        mensualidad = to_float(m["mensualidad"])

        data.append([
            Paragraph(m["fecha_matricula"] or "-", cell_style),
            Paragraph(m["alumno"] or "-", cell_style),
            Paragraph(m["institucion"] or "-", cell_style),
            Paragraph(m["sede"] or "-", cell_style),
            Paragraph(m["curso"] or "-", cell_style),
            Paragraph(f"S/ {monto:.2f}", cell_style),
            Paragraph(m["fecha_vencimiento"] or "-", cell_style),
            Paragraph(f"S/ {mensualidad:.2f}", cell_style),
        ])

    table = Table(data, colWidths=[
        60, 90, 70, 75, 120, 70, 110, 70
    ])

    table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor("#1d3557")),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 7),
        ('GRID',          (0,0), (-1,-1), 0.25, colors.grey),
        ('ROWBACKGROUNDS',(0,1), (-1,-1),
                          [colors.whitesmoke, colors.lightgrey]),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING',    (0,1), (-1,-1), 3),
    ]))

    elements.append(table)
    pdf.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="reporte_matriculas.pdf",
        mimetype="application/pdf"
    )

#=====================================================
# 🔹 RUTA: Agregar nueva matrícula
@app.route('/agregar_matricula', methods=['GET', 'POST'])
def agregar_matricula():
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    if request.method == 'POST':
        id_alumno = request.form['id_alumno']
        id_curso = request.form['id_curso']
        fecha_matricula = request.form['fecha_matricula']
        deuda = request.form['deuda'] or 0.0
        estado = request.form['estado']
        monto = float(request.form.get('monto') or 0.0)

        cur.execute("""
            INSERT INTO matriculas (id_alumno, id_curso, fecha_matricula, deuda, estado, monto)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (id_alumno, id_curso, fecha_matricula, deuda, estado, monto))

        conexion.commit()
        conexion.close()
        return redirect(url_for('administrar_matriculas'))

    # Si es GET: obtener listas de alumnos y cursos
    cur.execute("SELECT id_alumno, nombres, codigo FROM alumnos ORDER BY nombres ASC")
    alumnos = cur.fetchall()

    cur.execute("SELECT id_curso, nombre FROM cursos ORDER BY nombre ASC")
    cursos = cur.fetchall()

    conexion.close()

    return render_template('agregar_matricula.html', alumnos=alumnos, cursos=cursos)
        
@app.route('/editar_matricula/<int:id_matricula>', methods=['GET', 'POST'])
def editar_matricula(id_matricula):
    # 🔐 Solo usuarios logueados
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # Datos de sesión para filtro por sedes
    rol = session.get("rol")
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes", [])

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # ================== POST: GUARDAR CAMBIOS ==================
    if request.method == 'POST':
        id_alumno         = request.form['id_alumno']
        id_curso_raw      = request.form.get('id_curso')  # puede venir vacío
        fecha_matricula   = request.form['fecha_matricula']
        deuda             = float(request.form.get('deuda') or 0.0)
        estado            = request.form['estado']
        monto             = float(request.form.get('monto') or 0.0)
        tipo_pago         = request.form['tipo_pago']
        metodo_pago       = request.form.get('metodo_pago', '')
        saldo_matricula   = float(request.form.get('saldo_matricula') or 0.0)
        mensualidad       = float(request.form.get('mensualidad') or 0.0)
        observacion       = request.form.get('observacion', '')
        fecha_vencimiento = request.form.get('fecha_vencimiento') or None
        institucion       = (request.form.get('institucion') or '').strip()
        aula              = (request.form.get('aula') or '').strip()

        # ▶ Normalizar id_curso (None si viene vacío → SIN CURSO)
        id_curso = id_curso_raw if id_curso_raw not in (None, '', '0') else None

        # ⭐ OBTENER id_sede DESDE EL CURSO (para rellenar matriculas/alumnos)
        id_sede_curso = None
        if id_curso is not None:
            cur.execute("SELECT id_sede FROM cursos WHERE id_curso = ?", (id_curso,))
            row_sede = cur.fetchone()
            if row_sede:
                # row_factory = sqlite3.Row, así que se puede por nombre o índice
                id_sede_curso = row_sede["id_sede"]

        # ✅ Actualizar matrícula (incluyendo id_sede)
        cur.execute("""
            UPDATE matriculas
            SET id_alumno         = ?,
                id_curso          = ?,
                id_sede           = ?,         -- ⭐ nuevo
                fecha_matricula   = ?,
                deuda             = ?,
                estado            = ?,
                monto             = ?,
                tipo_pago         = ?,
                metodo_pago       = ?,
                saldo_matricula   = ?,
                mensualidad       = ?,
                fecha_vencimiento = ?,
                aula              = ?,
                observacion       = ?
            WHERE id_matricula    = ?
        """, (
            id_alumno,
            id_curso,
            id_sede_curso,
            fecha_matricula,
            deuda,
            estado,
            monto,
            tipo_pago,
            metodo_pago,
            saldo_matricula,
            mensualidad,
            fecha_vencimiento,
            aula,
            observacion,
            id_matricula
        ))

        # ✅ Actualizar institución **y sede** del alumno
        cur.execute("""
            UPDATE alumnos
            SET institucion = ?,
                id_sede     = ?     -- ⭐ nuevo
            WHERE id_alumno = ?
        """, (institucion, id_sede_curso, id_alumno))

        con.commit()
        con.close()
        return redirect(url_for('administrar_matriculas'))

    # ================== GET: CARGAR DATOS ==================
    # Traer matrícula + institución del alumno
    cur.execute("""
        SELECT 
            m.*,
            a.institucion
        FROM matriculas m
        JOIN alumnos a ON m.id_alumno = a.id_alumno
        WHERE m.id_matricula = ?
    """, (id_matricula,))
    matricula = cur.fetchone()

    if not matricula:
        con.close()
        return redirect(url_for('administrar_matriculas'))

    # ---------- Alumnos (los mantienes como los tienes) ----------
    cur.execute("SELECT id_alumno, nombres, codigo FROM alumnos ORDER BY nombres ASC")
    alumnos = cur.fetchall()

    # ---------- Cursos (MISMA LÓGICA QUE EN agregar) ----------
    if rol == "admin":
        # Admin ve todos los cursos y todas las sedes
        cursos = cur.execute("""
            SELECT 
                c.id_curso,
                c.nombre,
                c.id_sede,
                h.id_horario,
                h.dias AS dias,
                COALESCE(
                    h.nombre_horario || ' ' ||
                    COALESCE(h.dias || ' ', '') ||
                    '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                    c.nombre_horario,
                    ''
                ) AS horario_texto
            FROM cursos c
            LEFT JOIN horarios h ON c.id_horario = h.id_horario
            ORDER BY c.nombre
        """).fetchall()

    else:
        # Usuario normal
        if sedes_ids and len(sedes_ids) > 1:
            # varias sedes
            placeholders = ",".join("?" * len(sedes_ids))
            cursos = cur.execute(f"""
                SELECT 
                    c.id_curso,
                    c.nombre,
                    c.id_sede,
                    h.id_horario,
                    h.dias AS dias,
                    COALESCE(
                        h.nombre_horario || ' ' ||
                        COALESCE(h.dias || ' ', '') ||
                        '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                        c.nombre_horario,
                        ''
                    ) AS horario_texto
                FROM cursos c
                LEFT JOIN horarios h ON c.id_horario = h.id_horario
                WHERE c.id_sede IN ({placeholders})
                ORDER BY c.nombre
            """, sedes_ids).fetchall()
        else:
            # solo una sede
            cursos = cur.execute("""
                SELECT 
                    c.id_curso,
                    c.nombre,
                    c.id_sede,
                    h.id_horario,
                    h.dias AS dias,
                    COALESCE(
                        h.nombre_horario || ' ' ||
                        COALESCE(h.dias || ' ', '') ||
                        '(' || h.hora_inicio || ' - ' || h.hora_fin || ')',
                        c.nombre_horario,
                        ''
                    ) AS horario_texto
                FROM cursos c
                LEFT JOIN horarios h ON c.id_horario = h.id_horario
                WHERE c.id_sede = ?
                ORDER BY c.nombre
            """, (id_sede_usuario,)).fetchall()

    con.close()

    return render_template(
        'editar_matricula.html',
        matricula=matricula,
        alumnos=alumnos,
        cursos=cursos
    )


#=====================================================
# 🔹 RUTA: Eliminar matrícula (y alumno si ya no tiene más)
@app.route('/eliminar_matricula/<int:id_matricula>')
def eliminar_matricula(id_matricula):
    conexion = None
    try:
        conexion = sqlite3.connect('base_datos.db')
        conexion.row_factory = sqlite3.Row
        cur = conexion.cursor()

        # 1) Buscar el id_alumno de esa matrícula
        cur.execute("""
            SELECT id_alumno
            FROM matriculas
            WHERE id_matricula = ?
        """, (id_matricula,))
        fila = cur.fetchone()

        # Si no existe la matrícula, salimos normal
        if not fila:
            conexion.close()
            return redirect(url_for('administrar_matriculas'))

        id_alumno = fila["id_alumno"]

        # 2) Borrar la matrícula
        cur.execute("DELETE FROM matriculas WHERE id_matricula = ?", (id_matricula,))

        # 3) Ver si ese alumno tiene otras matrículas
        cur.execute("""
            SELECT COUNT(*) 
            FROM matriculas 
            WHERE id_alumno = ?
        """, (id_alumno,))
        restantes = cur.fetchone()[0]

        # Si ya no tiene ninguna matrícula → borramos al alumno
        if restantes == 0:
            cur.execute("DELETE FROM alumnos WHERE id_alumno = ?", (id_alumno,))

        conexion.commit()

    except Exception as e:
        app.logger.exception("Error al eliminar matrícula/alumno: %s", e)
        if conexion:
            conexion.rollback()
    finally:
        if conexion:
            conexion.close()

    return redirect(url_for('administrar_matriculas'))


#=====================================================
# 🔹 RUTA: Listar pagos pendientes
@app.route('/pagos_pendientes')
def pagos_pendientes():
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    cur.execute("""
        SELECT 
            m.id_matricula,
            a.codigo AS codigo_alumno,
            a.nombres AS nombre_alumno,
            c.nombre AS nombre_curso,
            s.nombre_sede AS sede,
            m.fecha_matricula,
            m.deuda,
            m.estado
        FROM matriculas m
        JOIN alumnos a ON m.id_alumno = a.id_alumno
        JOIN cursos c ON m.id_curso = c.id_curso
        JOIN sedes s ON c.id_sede = s.id_sede
        WHERE m.deuda > 0
        ORDER BY c.nombre, a.nombres;
    """)

    pagos_pendientes = cur.fetchall()
    conexion.close()

  
    return render_template('pagos_pendientes.html', pagos_pendientes=pagos_pendientes)

#=====================================================
# 🔹 RUTA: Administrar sedes
@app.route('/administrar_sedes')
def administrar_sedes():
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    cur.execute("""
        SELECT s.id_sede,
               s.nombre_sede,
               s.direccion,
               s.telefono,
               r1.nombres AS responsable_principal,
               r2.nombres AS responsable_secundario
        FROM sedes s
        LEFT JOIN responsables r1 ON s.id_responsable = r1.id_responsable
        LEFT JOIN responsables r2 ON s.id_responsable_secundario = r2.id_responsable
        ORDER BY s.id_sede DESC
    """)
    sedes = cur.fetchall()
    conexion.close()

    return render_template('administrar_sedes.html', sedes=sedes)

#=====================================================
# 🔹 RUTA: Agregar sede
@app.route('/agregar_sede', methods=['GET', 'POST'])
def agregar_sede():
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    if request.method == 'POST':
        nombre_sede = request.form['nombre_sede']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        id_responsable_1 = request.form['id_responsable_1']
        id_responsable_2 = request.form.get('id_responsable_2') or None  # opcional

        cur.execute("""
            INSERT INTO sedes (nombre_sede, direccion, telefono, id_responsable, id_responsable_secundario)
            VALUES (?, ?, ?, ?, ?)
        """, (nombre_sede, direccion, telefono, id_responsable_1, id_responsable_2))

        conexion.commit()
        conexion.close()
        return redirect(url_for('administrar_sedes'))

    # Obtener los responsables 
    cur.execute("SELECT id_responsable, nombres FROM responsables ORDER BY nombres ASC")
    responsables = cur.fetchall()
    conexion.close()

    return render_template('agregar_sede.html', responsables=responsables)

#=====================================================
# 🔹 RUTA: Editar sede
@app.route('/editar_sede/<int:id_sede>', methods=['GET', 'POST'])
def editar_sede(id_sede):
    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    if request.method == 'POST':
        nombre_sede = request.form['nombre_sede']
        direccion = request.form['direccion']
        telefono = request.form['telefono']
        id_responsable_1 = request.form['id_responsable_1']
        id_responsable_2 = request.form.get('id_responsable_2') or None

        cur.execute("""
            UPDATE sedes
            SET nombre_sede = ?, direccion = ?, telefono = ?, 
                id_responsable = ?, id_responsable_secundario = ?
            WHERE id_sede = ?
        """, (nombre_sede, direccion, telefono, id_responsable_1, id_responsable_2, id_sede))

        conexion.commit()
        conexion.close()
        return redirect(url_for('administrar_sedes'))

    # Si es GET, obtener la información actual de la sede
    cur.execute("SELECT * FROM sedes WHERE id_sede = ?", (id_sede,))
    sede = cur.fetchone()

    # Obtener los responsables disponibles para los combos
    cur.execute("SELECT id_responsable, nombres FROM responsables ORDER BY nombres ASC")
    responsables = cur.fetchall()
    conexion.close()

    if sede is None:
        return redirect(url_for('administrar_sedes'))

    return render_template('editar_sede.html', sede=sede, responsables=responsables)

#=====================================================
# 🔹 RUTA: Eliminar sede
@app.route('/eliminar_sede/<int:id_sede>')
def eliminar_sede(id_sede):
    conexion = sqlite3.connect('base_datos.db')
    cur = conexion.cursor()

    # Verificar si existe antes de eliminar
    cur.execute("SELECT * FROM sedes WHERE id_sede = ?", (id_sede,))
    sede = cur.fetchone()

    if not sede:
        conexion.close()
        return redirect(url_for('administrar_sedes'))

    # Eliminar la sede
    cur.execute("DELETE FROM sedes WHERE id_sede = ?", (id_sede,))
    conexion.commit()
    conexion.close()

    return redirect(url_for('administrar_sedes'))

#=====================================================
# 🔹 INYECTAR ROL EN TODAS LAS PLANTILLAS
@app.context_processor
def inject_user_role():
    return {
        "rol": session.get("rol")
    }

#=====================================================
# 🔹 RUTA: Administrar usuarios
@app.route('/administrar_usuarios')
def administrar_usuarios():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # 🚨 Solo admin puede entrar
    if session.get("rol") != "admin":
        return redirect(url_for('dashboard'))

    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    cur.execute("""
        SELECT 
            u.id_usuario,
            u.nombre_usuario,
            u.rol,
            -- 👇 TODAS las sedes del usuario, separadas por ' / '
            COALESCE(GROUP_CONCAT(s.nombre_sede, ' / '), 'Sin asignar') AS sedes
        FROM usuarios u
        LEFT JOIN usuario_sedes us ON u.id_usuario = us.id_usuario
        LEFT JOIN sedes s          ON us.id_sede   = s.id_sede
        GROUP BY u.id_usuario, u.nombre_usuario, u.rol
        ORDER BY u.id_usuario DESC
    """)

    usuarios = cur.fetchall()
    conexion.close()

    return render_template(
        'administrar_usuarios.html',
        usuarios=usuarios,
        rol=session.get("rol")
    )


@app.route('/agregar_usuario', methods=['GET', 'POST'])
def agregar_usuario():
    if 'usuario' not in session:
        return redirect(url_for('login'))

    conexion = sqlite3.connect('base_datos.db')
    conexion.row_factory = sqlite3.Row
    cur = conexion.cursor()

    if request.method == 'POST':
        nombre_usuario = request.form['nombre_usuario']
        clave = request.form['clave']
        rol = request.form['rol']

        # ✅ ahora leemos una LISTA de sedes
        sedes_seleccionadas = request.form.getlist('sedes')  # ['1', '3', ...]

        cur.execute("""
            INSERT INTO usuarios (nombre_usuario, clave, rol)
            VALUES (?, ?, ?)
        """, (nombre_usuario, clave, rol))
        id_usuario = cur.lastrowid

        # Insertar en tabla puente usuario_sedes
        for id_sede in sedes_seleccionadas:
            cur.execute("""
                INSERT INTO usuario_sedes (id_usuario, id_sede)
                VALUES (?, ?)
            """, (id_usuario, id_sede))

        conexion.commit()
        conexion.close()
        return redirect(url_for('administrar_usuarios'))

    cur.execute("SELECT id_sede, nombre_sede FROM sedes")
    sedes = cur.fetchall()
    conexion.close()

    return render_template('agregar_usuario.html', sedes=sedes)

#=====================================================
# 🔹 RUTA: Editar usuario (con varias sedes)
@app.route('/editar_usuario/<int:id_usuario>', methods=['GET', 'POST'])
def editar_usuario(id_usuario):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    # Solo admin edita usuarios (ajusta si quieres otra lógica)
    if session.get("rol") != "admin":
        return redirect(url_for('dashboard'))

    con = sqlite3.connect('base_datos.db')
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    if request.method == 'POST':
        nombre_usuario = (request.form.get('nombre_usuario') or '').strip()
        clave          = (request.form.get('clave') or '').strip()
        rol            = request.form.get('rol') or 'Usuario'
        sedes_ids      = request.form.getlist('sedes')  # lista de strings

        # Actualizar datos básicos del usuario
        cur.execute("""
            UPDATE usuarios
            SET nombre_usuario = ?, clave = ?, rol = ?
            WHERE id_usuario = ?
        """, (nombre_usuario, clave, rol, id_usuario))

        # Limpiamos relaciones anteriores de sedes
        cur.execute("DELETE FROM usuario_sedes WHERE id_usuario = ?", (id_usuario,))

        if rol != 'admin':
            # Para usuarios normales: grabamos las sedes seleccionadas
            for sid in sedes_ids:
                cur.execute("""
                    INSERT INTO usuario_sedes (id_usuario, id_sede)
                    VALUES (?, ?)
                """, (id_usuario, sid))

            # Opcional: compatibilidad con columna usuarios.id_sede
            # -> dejamos la primera sede como "principal"
            id_sede_principal = sedes_ids[0] if sedes_ids else None
            cur.execute("""
                UPDATE usuarios
                SET id_sede = ?
                WHERE id_usuario = ?
            """, (id_sede_principal, id_usuario))
        else:
            # Admin: ve todas las sedes, no se guardan sedes específicas
            cur.execute("""
                UPDATE usuarios
                SET id_sede = NULL
                WHERE id_usuario = ?
            """, (id_usuario,))

        con.commit()
        con.close()
        return redirect(url_for('administrar_usuarios'))

    # ---------- GET: cargar datos para el formulario ----------
    # Usuario
    cur.execute("SELECT * FROM usuarios WHERE id_usuario = ?", (id_usuario,))
    usuario = cur.fetchone()
    if not usuario:
        con.close()
        return "Usuario no encontrado", 404

    # Todas las sedes
    cur.execute("SELECT id_sede, nombre_sede FROM sedes ORDER BY nombre_sede")
    sedes = cur.fetchall()

    # Sedes asignadas actualmente a este usuario
    cur.execute("""
        SELECT id_sede
        FROM usuario_sedes
        WHERE id_usuario = ?
    """, (id_usuario,))
    sedes_usuario = [fila["id_sede"] for fila in cur.fetchall()]

    con.close()

    return render_template(
        'editar_usuario.html',
        usuario=usuario,
        sedes=sedes,
        sedes_usuario=sedes_usuario
    )

#=====================================================
# 🔹 RUTA: Eliminar usuario
@app.route('/eliminar_usuario/<int:id_usuario>', methods=['GET'])
def eliminar_usuario(id_usuario):
    try:
        conexion = sqlite3.connect('base_datos.db')
        conexion.row_factory = sqlite3.Row
        cur = conexion.cursor()

        # Verifica si existe
        cur.execute("SELECT rol FROM usuarios WHERE id_usuario = ?", (id_usuario,))
        usuario = cur.fetchone()

        if not usuario:
            conexion.close()
            return redirect(url_for('administrar_usuarios') + "?msg=Usuario+no+existe")

        # Evitar eliminar el último admin
        if usuario['rol'] == 'admin':
            cur.execute("SELECT COUNT(*) AS total FROM usuarios WHERE rol = 'admin'")
            admins = cur.fetchone()['total']
            if admins <= 1:
                conexion.close()
                return redirect(url_for('administrar_usuarios') + "?msg=No+puedes+eliminar+el+último+administrador")

        # Eliminar usuario
        cur.execute("DELETE FROM usuarios WHERE id_usuario = ?", (id_usuario,))
        conexion.commit()
        conexion.close()

        return redirect(url_for('administrar_usuarios') + "?msg=Usuario+eliminado+correctamente")

    except Exception as e:
        print("⚠️ Error eliminando usuario:", e)
        return redirect(url_for('administrar_usuarios') + "?msg=Error+al+eliminar")

#=====================================================
# 🔹 RUTA: Listar pagos (con filtro por rol/sede)
@app.route("/pagos")
def listar_pagos():
    # --- Seguridad básica: si no hay sesión, al login ---
    if "usuario" not in session:
        return redirect(url_for("login"))

    # Datos de sesión
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes para usuarios multi-sede

    # Filtros desde el formulario
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    curso = request.args.get("curso")

    con = get_db_connection()
    cur = con.cursor()

    # =====================================================
    # 1) LISTA DE CURSOS SEGÚN ROL / SEDES (igual patrón que control_alumnos)
    # =====================================================
    if rol == "admin":
        cur.execute("""
            SELECT id_curso, nombre
            FROM cursos
            ORDER BY nombre
        """)
        cursos = cur.fetchall()
    else:
        # Usuario normal → solo cursos de sus sedes
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            cur.execute(f"""
                SELECT id_curso, nombre
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre
            """, sedes_ids)
        else:
            cur.execute("""
                SELECT id_curso, nombre
                FROM cursos
                WHERE id_sede = ?
                ORDER BY nombre
            """, (id_sede_usuario,))
        cursos = cur.fetchall()

    # =====================================================
    # 2) CONSULTA PRINCIPAL DE PAGOS
    # =====================================================
    sql = """
        SELECT 
            p.id_pago,
            p.mes,
            p.monto,
            p.tipo_pago,
            p.metodo_pago,
            p.proximo_pago,
            p.fecha_pago,
            p.observacion,
            p.fecha_registro,
            a.nombres AS alumno,
            a.celular,
            a.correo,
            c.nombre AS curso
        FROM pagos p
        JOIN alumnos a ON p.id_alumno = a.id_alumno

        -- ⭐ MATRÍCULA ÚNICA DEL ALUMNO
        LEFT JOIN (
            SELECT id_alumno, id_curso 
            FROM matriculas 
            GROUP BY id_alumno
        ) m ON m.id_alumno = a.id_alumno

        LEFT JOIN cursos c ON c.id_curso = m.id_curso
        WHERE 1=1
    """

    params = []

    # Filtro por rango de fechas
    if desde and hasta:
        sql += " AND substr(p.fecha_pago, 1, 10) BETWEEN ? AND ?"
        params.extend([desde, hasta])

    # Filtro por curso
    if curso:
        sql += " AND c.id_curso = ?"
        params.append(curso)

    # =====================================================
    # 3) FILTRO POR SEDE SEGÚN ROL (ADMIN ve todo)
    # =====================================================
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            sql += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            sql += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    sql += " ORDER BY p.fecha_pago DESC"

    cur.execute(sql, params)
    pagos = cur.fetchall()

    con.close()

    return render_template(
        "pagos.html",
        pagos=pagos,
        cursos=cursos,
        desde=desde,
        hasta=hasta,
        curso=curso
    )

@app.route("/exportar_pagos")
def exportar_pagos():
    from io import BytesIO
    import pandas as pd
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from flask import send_file

    # 🔐 Solo usuarios logueados
    if "usuario" not in session:
        return redirect(url_for("login"))

    # Datos de sesión (mismo patrón que en /pagos)
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []

    fecha_desde = request.args.get("desde")
    fecha_hasta = request.args.get("hasta")
    curso = request.args.get("curso")   # id_curso

    con = get_db_connection()
    cur = con.cursor()

    # ------- BASE SQL (igual que listar_pagos, pero para Excel) -------
    sql = """
        SELECT 
            p.id_pago,
            a.codigo,
            a.nombres AS alumno,
            c.nombre AS curso,
            p.mes,
            p.monto,
            p.metodo_pago,
            p.observacion,
            p.fecha_pago
        FROM pagos p
        JOIN alumnos a ON p.id_alumno = a.id_alumno

        -- ⭐ MATRÍCULA ÚNICA DEL ALUMNO
        LEFT JOIN (
            SELECT id_alumno, id_curso 
            FROM matriculas 
            GROUP BY id_alumno
        ) m ON m.id_alumno = a.id_alumno

        LEFT JOIN cursos c ON c.id_curso = m.id_curso
        WHERE 1 = 1
    """

    params = []

    # 🔽 Filtro por fechas (solo si vienen ambas)
    if fecha_desde and fecha_hasta:
        sql += " AND date(p.fecha_pago) BETWEEN ? AND ?"
        params.extend([fecha_desde, fecha_hasta])

    # 🔽 Filtro por curso (id_curso)
    if curso:
        sql += " AND c.id_curso = ?"
        params.append(curso)

    # 🔽 Filtro por sede(s) según rol
    if rol != "admin":
        if sedes_ids:
            placeholders = ",".join("?" * len(sedes_ids))
            sql += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            sql += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    sql += " ORDER BY p.fecha_pago DESC"

    cur.execute(sql, params)
    filas = cur.fetchall()

    # Si NO hay registros que exportar → volvemos a la pantalla de pagos
    if not filas:
        # Recargamos lista de cursos para que el template no se rompa
        if rol == "admin":
            cursos_filtro = cur.execute("""
                SELECT id_curso, nombre 
                FROM cursos
                ORDER BY nombre
            """).fetchall()
        else:
            if sedes_ids:
                placeholders = ",".join("?" * len(sedes_ids))
                cursos_filtro = cur.execute(f"""
                    SELECT id_curso, nombre
                    FROM cursos
                    WHERE id_sede IN ({placeholders})
                    ORDER BY nombre
                """, sedes_ids).fetchall()
            else:
                cursos_filtro = cur.execute("""
                    SELECT id_curso, nombre
                    FROM cursos
                    WHERE id_sede = ?
                    ORDER BY nombre
                """, (id_sede_usuario,)).fetchall()

        con.close()
        return render_template(
            "pagos.html",
            pagos=[],
            cursos=cursos_filtro,
            mensaje="⚠️ No hay registros para exportar con esos filtros."
        )

    # ----- Título según filtros -----
    if fecha_desde and fecha_hasta:
        titulo = f"Pagos del {fecha_desde} al {fecha_hasta}"
        if curso:
            titulo += " - Filtrado por curso"
        nombre_archivo = f"Pagos_{fecha_desde}_a_{fecha_hasta}.xlsx"
    else:
        # 👉 Si no se selecciona rango de fechas, exporta TODO lo que cumpla filtros
        titulo = "Pagos - Completo"
        nombre_archivo = "Pagos_completo.xlsx"

    # Crear DataFrame
    columnas = [desc[0] for desc in cur.description]
    df = pd.DataFrame(filas, columns=columnas)

    # Renombrar columnas para Excel
    df.rename(columns={
        "id_pago": "ID",
        "codigo": "Código",
        "alumno": "Alumno",
        "curso": "Curso",
        "mes": "Mes Pagado",
        "monto": "Monto (S/.)",
        "metodo_pago": "Método",
        "observacion": "Observación",
        "fecha_pago": "Fecha Pago"
    }, inplace=True)

    output = BytesIO()

    # Exportar a Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, sheet_name='Pagos')
        sheet = writer.sheets['Pagos']

        # Encabezado principal
        sheet.merge_cells("A1:I1")
        sheet["A1"] = "REPORTE DE PAGOS"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # Subtítulo
        sheet.merge_cells("A2:I2")
        sheet["A2"] = titulo
        sheet["A2"].alignment = Alignment(horizontal="center")

        # Ajustar anchos de columnas
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_length + 2

    con.close()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


#=====================================================
# 🔹 RUTA: Agregar nuevo pago (con filtro por rol / sedes)
@app.route("/pagos/agregar", methods=["GET", "POST"])
def agregar_pago():
    # 🔐 Seguridad: solo usuarios logueados
    if "usuario" not in session:
        return redirect(url_for("login"))

    # Datos de sesión
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes si es multi-sede

    con = get_db_connection()
    cur = con.cursor()

    # ============================================
    # 1) Listas de ALUMNOS y CURSOS según rol/sedes
    # ============================================
    if rol == "admin":
        # Admin ve todo
        alumnos = cur.execute("""
            SELECT id_alumno, nombres 
            FROM alumnos 
            ORDER BY nombres ASC
        """).fetchall()

        cursos = cur.execute("""
            SELECT DISTINCT nombre 
            FROM cursos 
            ORDER BY nombre ASC
        """).fetchall()

    else:
        # Usuario normal → solo alumnos y cursos de sus sede(s)
        if sedes_ids:
            # Multi-sede
            placeholders = ",".join("?" * len(sedes_ids))

            alumnos = cur.execute(f"""
                SELECT DISTINCT a.id_alumno, a.nombres
                FROM alumnos a
                JOIN matriculas m ON m.id_alumno = a.id_alumno
                JOIN cursos c ON c.id_curso = m.id_curso
                WHERE c.id_sede IN ({placeholders})
                ORDER BY a.nombres ASC
            """, sedes_ids).fetchall()

            cursos = cur.execute(f"""
                SELECT DISTINCT nombre
                FROM cursos
                WHERE id_sede IN ({placeholders})
                ORDER BY nombre ASC
            """, sedes_ids).fetchall()
        else:
            # Una sola sede
            alumnos = cur.execute("""
                SELECT DISTINCT a.id_alumno, a.nombres
                FROM alumnos a
                JOIN matriculas m ON m.id_alumno = a.id_alumno
                JOIN cursos c ON c.id_curso = m.id_curso
                WHERE c.id_sede = ?
                ORDER BY a.nombres ASC
            """, (id_sede_usuario,)).fetchall()

            cursos = cur.execute("""
                SELECT DISTINCT nombre
                FROM cursos
                WHERE id_sede = ?
                ORDER BY nombre ASC
            """, (id_sede_usuario,)).fetchall()

    # ============================================
    # 2) POST: guardar el pago
    # ============================================
    if request.method == "POST":
        id_alumno    = request.form.get("id_alumno")
        curso        = request.form.get("curso")       # texto del curso, como ya lo usas
        mes          = request.form.get("mes")
        tipo_pago    = request.form.get("tipo_pago")
        monto        = float(request.form.get("monto") or 0)
        metodo_pago  = request.form.get("metodo_pago")
        proximo_pago = request.form.get("proximo_pago")
        observacion  = request.form.get("observacion")

        cur.execute("""
            INSERT INTO pagos 
            (id_alumno, curso, mes, tipo_pago, monto, metodo_pago, proximo_pago, observacion, fecha_pago, fecha_registro)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, DATE('now'), DATETIME('now','localtime'))
        """, (
            id_alumno,
            curso,
            mes,
            tipo_pago,
            monto,
            metodo_pago,
            proximo_pago,
            observacion
        ))

        con.commit()
        con.close()
        return redirect(url_for("listar_pagos"))

    # GET → mostrar formulario
    con.close()
    return render_template("agregar_pago.html", alumnos=alumnos, cursos=cursos)

#=====================================================
# 🔹 RUTA: Buscar alumno (para autocompletar, filtrado por sede)
@app.route("/buscar_alumno")
def buscar_alumno():
    q = (request.args.get("q") or "").strip()

    # --- Datos de sesión para filtrar por sede ---
    rol = (session.get("rol") or "").strip().lower()
    id_sede_usuario = session.get("id_sede")
    sedes_ids = session.get("sedes") or []   # lista de sedes para usuarios multi-sede

    con = get_db_connection()
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # Base: alumnos + matriculas + cursos (para poder filtrar por sede)
    sql = """
        SELECT DISTINCT
            a.id_alumno,
            a.nombres
        FROM alumnos a
        LEFT JOIN matriculas m ON a.id_alumno = m.id_alumno
        LEFT JOIN cursos c ON m.id_curso = c.id_curso
        WHERE
            (a.nombres LIKE ? OR a.dni LIKE ? OR a.codigo LIKE ?)
    """
    like = f"%{q}%"
    params = [like, like, like]

    # --- Filtro por sede para NO admin ---
    if rol != "admin":
        if sedes_ids:
            # Usuario con varias sedes
            placeholders = ",".join("?" * len(sedes_ids))
            sql += f" AND c.id_sede IN ({placeholders})"
            params.extend(sedes_ids)
        else:
            # Usuario con una sola sede
            sql += " AND c.id_sede = ?"
            params.append(id_sede_usuario)

    sql += """
        ORDER BY a.nombres ASC
        LIMIT 10
    """

    rows = cur.execute(sql, params).fetchall()
    con.close()

    return jsonify([
        {"id": r["id_alumno"], "nombre": r["nombres"]}
        for r in rows
    ])

#=====================================================
# 🔹 RUTA: Validar matrícula del alumno en curso
@app.route("/validar_matricula")
def validar_matricula():
    id_alumno = request.args.get("id_alumno")
    curso = request.args.get("curso")

    con = get_db_connection()
    row = con.execute("""
        SELECT 1 FROM matriculas m
        JOIN cursos c ON m.id_curso = c.id_curso
        WHERE m.id_alumno = ? AND c.nombre = ?
    """, (id_alumno, curso)).fetchone()
    con.close()

    return jsonify({"ok": row is not None})

#=====================================================
# 🔹 RUTA: Editar pago existente
@app.route("/pagos/editar/<int:id_pago>", methods=["GET", "POST"])
def editar_pago(id_pago):
    con = get_db_connection()
    cur = con.cursor()

    # Obtener el pago existente
    pago = cur.execute("""
        SELECT *
        FROM pagos
        WHERE id_pago = ?
    """, (id_pago,)).fetchone()

    # Lista de alumnos para el select
    alumnos = cur.execute("""
        SELECT id_alumno, nombres
        FROM alumnos
        ORDER BY nombres ASC
    """).fetchall()

    if request.method == "POST":
        # NO leemos id_alumno del form, usamos el de la BD
        id_alumno    = pago["id_alumno"]   # 👈 viene del SELECT inicial
        mes          = request.form.get("mes")
        monto        = float(request.form.get("monto") or 0)
        metodo_pago  = request.form.get("metodo_pago")
        tipo_pago    = request.form.get("tipo_pago")
        proximo_pago = request.form.get("proximo_pago") or None
        fecha_pago   = request.form.get("fecha_pago") or None
        curso        = request.form.get("curso") or None
        observacion  = request.form.get("observacion") or None

        cur.execute("""
            UPDATE pagos
            SET 
                id_alumno    = ?,         -- sigue estando, pero ya nunca es NULL
                mes          = ?,
                monto        = ?,
                metodo_pago  = ?,
                observacion  = ?,
                fecha_pago   = DATE('now'),
                curso        = ?,
                proximo_pago = ?,
                tipo_pago    = ?,
                fecha_registro = DATETIME('now','localtime')
            WHERE id_pago = ?
        """, (
            id_alumno,
            mes,
            monto,
            metodo_pago,
            observacion,
            curso,
            proximo_pago,
            tipo_pago,
            id_pago
        ))



        con.commit()
        con.close()
        return redirect(url_for("listar_pagos"))

    con.close()
    return render_template("editar_pago.html", pago=pago, alumnos=alumnos)

#=====================================================
# 🔹 RUTA: Obtener curso del alumno
@app.route("/obtener_curso_alumno")
def obtener_curso_alumno():
    id_alumno = request.args.get("id_alumno")

    con = get_db_connection()
    cur = con.cursor()

    cur.execute("""
        SELECT c.nombre
        FROM matriculas m
        JOIN cursos c ON c.id_curso = m.id_curso
        WHERE m.id_alumno = ?
    """, (id_alumno,))

    row = cur.fetchone()
    con.close()

    return jsonify({"curso": row[0] if row else None})

#=====================================================
# 🔹 RUTA: Eliminar pago
@app.route("/pagos/eliminar/<int:id_pago>")
def eliminar_pago(id_pago):
    con = get_db_connection()
    cur = con.cursor()

    cur.execute("DELETE FROM pagos WHERE id_pago = ?", (id_pago,))
    con.commit()
    con.close()

    return redirect(url_for("listar_pagos"))

#=====================================================
# 🔹 FUNCIONES AUXILIARES
def generar_qr_servidor(ip_local):
    """
    Genera un QR con la URL base del sistema: http://IP:5000
    y lo guarda en static/qrcodes/servidor.png
    """
    try:
        base_url = f"http://{ip_local}:5000"
        qr_folder = os.path.join(STATIC_DIR, "qrcodes")
        os.makedirs(qr_folder, exist_ok=True)

        qr_path = os.path.join(qr_folder, "servidor.png")

        img = qrcode.make(base_url)
        img.save(qr_path)

        print(f"✅ QR del servidor generado: {base_url} -> {qr_path}")
    except Exception as e:
        app.logger.exception("Error generando QR del servidor: %s", e)
        
#localhost para poder iniciar el cmd y la aplicacion 

if __name__ == "__main__":
    # Detectar si estamos ejecutando como ejecutable empaquetado
    es_ejecutable = getattr(sys, 'frozen', False)
    
    # Obtener IP local
    ip_local = obtener_ip_local()
    
    print("=" * 60)
    print("🚀 SISTEMA DE ASISTENCIAS QR")
    print("=" * 60)
    print(f"🌐 Servidor iniciando en: http://{ip_local}:5000")
    print(f"📱 Acceso desde red local: http://{ip_local}:5000")
    print("=" * 60)
    print()
    
    # 🔥 Generar QR del servidor
    generar_qr_servidor(ip_local)
    
    # Abrir navegador después de 2 segundos (dar tiempo al servidor)
    threading.Timer(2.0, abrir_navegador).start()
    
    # Usar waitress en producción (ejecutable) o Flask en desarrollo
    if es_ejecutable:
        print("✅ Modo PRODUCCIÓN (ejecutable)")
        print("⚠️  Para detener: Cierra la ventana del navegador y presiona Ctrl+C")
        print()
        serve(app, host="0.0.0.0", port=5000, threads=4)
    else:
        print("✅ Modo DESARROLLO")
        print()
        app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=True)

        
